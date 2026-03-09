import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
import os
import re
import io
from pathlib import Path
from datetime import date, timedelta
from typing import Dict, List, Optional, Tuple, Any, Union
from chart_utils import (
    CHART_STYLES, get_factory_color, get_process_color,
    apply_standard_layout, create_standard_line_chart, 
    create_standard_bar_chart, create_combo_chart
)

# --- 페이지 기본 설정 ---
st.set_page_config(layout="wide", page_title="지능형 생산 대시보드 2021~2022 전용", page_icon="👑")

# --- 화면 너비에 따른 동적 배율 조정 ---
st.markdown("""
<script>
function adjustZoom() {
    const baseWidth = 1920; // 기준 너비 (FHD)
    const currentWidth = window.innerWidth;
    
    // 기준 너비에서의 배율
    let scale = (currentWidth / baseWidth) * 0.8; 
    
    // 최소/최대 배율 제한
    scale = Math.max(0.65, Math.min(scale, 1.0)); 
    
    document.body.style.zoom = scale;
}

// 페이지 로드 및 창 크기 변경 시 함수 호출
window.addEventListener('load', adjustZoom);
window.addEventListener('resize', adjustZoom);
</script>
""", unsafe_allow_html=True)

# --- 데이터 로딩 및 캐싱 ---
@st.cache_data
def load_all_data() -> Dict[str, Tuple[pd.DataFrame, Optional[str]]]:
    """파일 로딩 및 데이터 전처리"""
    data_frames = {}
    keywords = {
        'target': '목표달성율', 
        'yield': '수율', 
        'utilization': '가동률', 
        'defect': ('불량실적현황', '최적화')
    }
    
    current_directory = '.'
    all_files_in_dir = os.listdir(current_directory)

    for key, keyword_info in keywords.items():
        try:
            csv_files = []
            xlsx_files = []

            for f in all_files_in_dir:
                filename_without_ext, ext = os.path.splitext(f)

                # CSV와 Excel 파일 모두 지원
                if ext.lower() not in ['.csv', '.xlsx', '.xls']:
                    continue

                normalized_name = filename_without_ext.replace("(", "").replace(")", "").replace(" ", "")

                # 키워드 매칭
                is_match = False
                if key == 'defect':
                    kw_base, kw_opt = keyword_info
                    if kw_base in normalized_name and kw_opt in normalized_name:
                        is_match = True
                else:
                    if keyword_info in normalized_name:
                        is_match = True

                # 파일 형식별로 분류
                if is_match:
                    if ext.lower() == '.csv':
                        csv_files.append(f)
                    else:
                        xlsx_files.append(f)

            # CSV 파일을 우선 선택
            relevant_files = csv_files if csv_files else xlsx_files

            if relevant_files:
                # (간편) 표기가 없는 최신 파일을 우선 선택
                non_simple_files = [f for f in relevant_files if '간편' not in f]
                search_pool = non_simple_files if non_simple_files else relevant_files

                # 21~22 전용 패키지에서는 통합 수율 파일을 우선 사용해 연도 누락을 방지
                if key == 'yield':
                    preferred_pattern = r'^생산실적현황\(수율\)\.(csv|xlsx|xls)$'
                    preferred_files = [f for f in search_pool if re.match(preferred_pattern, f)]
                    latest_file = preferred_files[0] if preferred_files else max(
                        search_pool, key=lambda f: os.path.getmtime(os.path.join(current_directory, f))
                    )
                else:
                    latest_file = max(search_pool, key=lambda f: os.path.getmtime(os.path.join(current_directory, f)))

                file_path = os.path.join(current_directory, latest_file)
                file_ext = os.path.splitext(latest_file)[1].lower()

                # 파일 형식에 따라 읽기
                if file_ext == '.csv':
                    # CSV 파일 읽기 (천단위 구분자 처리)
                    df = pd.read_csv(file_path, encoding='utf-8-sig', thousands=',', skip_blank_lines=True)
                    # Unnamed 컬럼 제거
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False)]
                    # 완전히 빈 컬럼 제거
                    df = df.dropna(axis=1, how='all')
                    # 컬럼명 공백 제거
                    df.columns = df.columns.str.strip()
                else:
                    # Excel 파일 읽기
                    try:
                        df = pd.read_excel(file_path, engine='openpyxl')
                    except Exception as excel_error:
                        # Excel 읽기 실패 시 CSV로 재시도
                        df = pd.read_csv(file_path, encoding='utf-8-sig', thousands=',', skip_blank_lines=True)
                        df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False)]
                        df = df.dropna(axis=1, how='all')
                        df.columns = df.columns.str.strip()
                
                for col in df.columns:
                    if df[col].dtype == 'object' and ('%' in str(df[col].iloc[0]) if not df[col].empty and df[col].iloc[0] is not None else False):
                        df[col] = df[col].astype(str).str.replace('%', '', regex=False).str.strip()
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                
                if key == 'defect':
                    cols = pd.Series(df.columns)
                    for dup in cols[cols.duplicated()].unique():
                        cols[cols[cols == dup].index.values.tolist()] = [f"{dup}_{i}" if i != 0 else dup for i in range(sum(cols == dup))]
                    df.columns = cols
                    
                    rename_dict = {}
                    if '불량수량(유형별)' in df.columns: rename_dict['불량수량(유형별)'] = '유형별_불량수량'
                    if '불량수량(전체)' in df.columns: rename_dict['불량수량(전체)'] = '총_불량수량'
                    elif '불량수량' in df.columns and '불량수량_1' in df.columns:
                        rename_dict['불량수량'] = '총_불량수량'
                        rename_dict['불량수량_1'] = '유형별_불량수량'
                    df = df.rename(columns=rename_dict)

                data_frames[key] = (df, latest_file)
            else:
                 data_frames[key] = (pd.DataFrame(), None)
        except Exception:
            data_frames[key] = (pd.DataFrame(), None)
    return data_frames

# --- AI 분석 엔진 ---
def analyze_target_data(df: pd.DataFrame) -> str:
    """목표 달성률 데이터 분석 브리핑 생성"""
    return (
        "#### AI Analyst 브리핑\n"
        "'양품 기반 달성률'을 기준으로 공장/공정별 성과를 비교하고, "
        "목표 대비 **양품 수량**의 차이가 큰 항목을 확인하여 "
        "품질 및 생산성 개선 포인트를 동시에 도출해야 합니다."
    )
def analyze_yield_data(df: pd.DataFrame) -> str:
    """수율 데이터 분석 브리핑 생성"""
    return (
        "#### AI Analyst 브리핑\n"
        "'수율'은 품질 경쟁력의 핵심 지표입니다. "
        "수율이 낮은 공정/품명을 식별하고, "
        "생산량 대비 양품 수량의 차이를 분석하여 원인을 개선해야 합니다."
    )
def analyze_utilization_data(df: pd.DataFrame) -> str:
    """가동률 데이터 분석 브리핑 생성"""
    return (
        "#### AI Analyst 브리핑\n"
        "'가동률'은 생산 효율성을 나타냅니다. "
        "이론적인 생산 능력(CAPA)과 실제 생산량의 차이를 분석하여, "
        "유휴 시간 및 비가동 손실을 최소화해야 합니다."
    )
def analyze_defect_data(df: pd.DataFrame) -> str:
    """불량 데이터 분석 브리핑 생성"""
    return (
        "#### AI Analyst 브리핑\n"
        "'파레토 분석'은 '80/20 법칙'에 기반하여, "
        "소수의 핵심 불량 원인이 전체 문제의 대부분을 차지한다고 봅니다. "
        "차트의 왼쪽에서부터 가장 큰 비중을 차지하는 불량 유형에 집중하여 "
        "개선 활동을 펼치면, 최소의 노력으로 최대의 품질 개선 효과를 얻을 수 있습니다."
    )

# --- 상수 정의 (chart_styles.json에서 로드) ---
PROCESS_MASTER_ORDER = ['[10] 사출조립', '[20] 분리', '[45] 하이드레이션/전면검사', '[55] 접착/멸균', '[80] 누수/규격검사']
FACTORY_COLOR_MAP = CHART_STYLES["colors"]["factory_colors"]  # JSON에서 로드
DEFAULT_FACTORY_COLOR = CHART_STYLES["colors"].get("default_color", "#888888")
CHART_HEIGHT = CHART_STYLES["chart_config"]["default_height"]  # JSON에서 로드
TEXT_FONT_SIZE = CHART_STYLES["chart_config"]["text_font_size"]  # JSON에서 로드

FACTORY_DEFINITIONS: Dict[str, str] = {
    "A관": "1공장",
    "C관": "2공장",
    "S관": "3공장",
}
FACTORY_DISPLAY_LABELS: Dict[str, str] = {
    code: f"{code} ({name})" for code, name in FACTORY_DEFINITIONS.items()
}
FACTORY_DISPLAY_CHOICES: List[str] = [FACTORY_DISPLAY_LABELS[code] for code in FACTORY_DEFINITIONS]
FACTORY_DISPLAY_TO_CODE: Dict[str, str] = {label: code for code, label in FACTORY_DISPLAY_LABELS.items()}

BASE_DIR = Path(__file__).resolve().parent
WORKFORCE_FILE_PATH = BASE_DIR / "workforce_master.xlsx"
EXCLUDED_WORKING_DAYS_PATH = BASE_DIR / "년월별 제외근무일수.csv"
WORKFORCE_SHEET_COLUMNS: Dict[str, List[str]] = {
    "배치운영": ["공장", "공정", "필요인원", "배치인원", "근무조", "신규투입", "이동계획"],
    "근태관리": ["날짜", "공장", "공정", "지각", "결근", "휴가", "특근", "비고"],
    "생산성": ["공장", "공정", "UPH", "UPPH", "평균작업시간(분)", "효율(%)", "잔업시간", "특근생산성"],
    "교육자격": ["공장", "이름", "부서", "교육명", "수료일", "만료일", "상태"],
    "수급계획": ["월", "공장", "예상수요", "가용인원", "외주/채용계획", "코멘트"],
    "비용관리": ["공장", "부서", "기본급합계", "잔업비", "특근비", "총인건비"],
    "현장이슈": ["날짜", "공장", "유형", "내용", "심각도", "조치현황"],
}
DEFAULT_WORKFORCE_DATA: Dict[str, List[Dict[str, Any]]] = {
    "배치운영": [
        {"공장": "A관", "공정": "사출", "필요인원": 24, "배치인원": 22, "근무조": "A/B", "신규투입": 2, "이동계획": "B조 1명 증원"},
        {"공장": "C관", "공정": "조립", "필요인원": 18, "배치인원": 18, "근무조": "주/야", "신규투입": 1, "이동계획": "야간 1명 교육"},
        {"공장": "S관", "공정": "검사", "필요인원": 14, "배치인원": 12, "근무조": "주간", "신규투입": 0, "이동계획": "라인 다기능화"},
        {"공장": "S관", "공정": "포장", "필요인원": 10, "배치인원": 11, "근무조": "2교대", "신규투입": 0, "이동계획": "여유 인원 검사 지원"},
    ],
    "근태관리": [
        {"날짜": "2025-01-02", "공장": "A관", "공정": "사출", "지각": 1, "결근": 0, "휴가": 2, "특근": 1, "비고": "설비 점검"},
        {"날짜": "2025-01-02", "공장": "C관", "공정": "조립", "지각": 0, "결근": 1, "휴가": 1, "특근": 0, "비고": ""},
        {"날짜": "2025-01-03", "공장": "S관", "공정": "검사", "지각": 0, "결근": 0, "휴가": 1, "특근": 1, "비고": "증가 요청"},
        {"날짜": "2025-01-03", "공장": "S관", "공정": "포장", "지각": 2, "결근": 0, "휴가": 0, "특근": 1, "비고": "폭설 영향"},
    ],
    "생산성": [
        {"공장": "A관", "공정": "사출", "UPH": 145, "UPPH": 6.2, "평균작업시간(분)": 48, "효율(%)": 92, "잔업시간": 1.5, "특근생산성": 138},
        {"공장": "C관", "공정": "조립", "UPH": 110, "UPPH": 5.1, "평균작업시간(분)": 54, "효율(%)": 88, "잔업시간": 2.0, "특근생산성": 120},
        {"공장": "S관", "공정": "검사", "UPH": 90, "UPPH": 4.8, "평균작업시간(분)": 60, "효율(%)": 95, "잔업시간": 0.5, "특근생산성": 98},
        {"공장": "S관", "공정": "포장", "UPH": 130, "UPPH": 5.5, "평균작업시간(분)": 52, "효율(%)": 89, "잔업시간": 1.0, "특근생산성": 133},
    ],
    "교육자격": [
        {"공장": "A관", "이름": "김현수", "부서": "사출", "교육명": "금형 안전", "수료일": "2024-11-05", "만료일": "2025-11-05", "상태": "정상"},
        {"공장": "S관", "이름": "이서연", "부서": "검사", "교육명": "품질 검사", "수료일": "2024-08-12", "만료일": "2025-08-12", "상태": "만료예정"},
        {"공장": "C관", "이름": "박지훈", "부서": "조립", "교육명": "라인 다기능", "수료일": "2023-12-01", "만료일": "2025-01-31", "상태": "갱신필요"},
    ],
    "수급계획": [
        {"월": "2025-01", "공장": "A관", "예상수요": 24, "가용인원": 22, "외주/채용계획": "야간 계약직 2명", "코멘트": "설비 증설 대응"},
        {"월": "2025-01", "공장": "C관", "예상수요": 23, "가용인원": 22, "외주/채용계획": "사내 다기능화", "코멘트": "주간 안정화"},
        {"월": "2025-01", "공장": "S관", "예상수요": 23, "가용인원": 22, "외주/채용계획": "단기 외주 1팀", "코멘트": "포장 캐파 확대"},
        {"월": "2025-02", "공장": "A관", "예상수요": 25, "가용인원": 24, "외주/채용계획": "계약연장 협의", "코멘트": "수요 증가 대비"},
        {"월": "2025-02", "공장": "C관", "예상수요": 24, "가용인원": 23, "외주/채용계획": "경력직 1명 채용", "코멘트": "신규 라인 준비"},
        {"월": "2025-02", "공장": "S관", "예상수요": 25, "가용인원": 24, "외주/채용계획": "주야 교대 보강", "코멘트": "성수기 대비"},
    ],
    "비용관리": [
        {"공장": "A관", "부서": "사출", "기본급합계": 28000, "잔업비": 4200, "특근비": 1800, "총인건비": 34000},
        {"공장": "C관", "부서": "조립", "기본급합계": 22000, "잔업비": 3800, "특근비": 2200, "총인건비": 28000},
        {"공장": "S관", "부서": "검사", "기본급합계": 16000, "잔업비": 2100, "특근비": 900, "총인건비": 19000},
        {"공장": "S관", "부서": "포장", "기본급합계": 14000, "잔업비": 1700, "특근비": 800, "총인건비": 16500},
    ],
    "현장이슈": [
        {"날짜": "2025-01-05", "공장": "A관", "유형": "이탈", "내용": "야간조 2명 퇴사 예정", "심각도": "높음", "조치현황": "면담 및 충원 진행"},
        {"날짜": "2025-01-06", "공장": "S관", "유형": "안전", "내용": "포장 라인 경미한 안전사고", "심각도": "중간", "조치현황": "현장 재교육"},
        {"날짜": "2025-01-07", "공장": "C관", "유형": "만족도", "내용": "조립 야간조 근무 만족도 하락", "심각도": "중간", "조치현황": "휴게 환경 개선"},
    ],
}
def ensure_workforce_master() -> None:
    WORKFORCE_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if WORKFORCE_FILE_PATH.exists():
        return

    with pd.ExcelWriter(WORKFORCE_FILE_PATH, engine="openpyxl") as writer:
        for sheet_name, columns in WORKFORCE_SHEET_COLUMNS.items():
            rows = DEFAULT_WORKFORCE_DATA.get(sheet_name, [])
            df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def load_workforce_data() -> Dict[str, pd.DataFrame]:
    ensure_workforce_master()
    data: Dict[str, pd.DataFrame] = {}
    workbook = pd.ExcelFile(WORKFORCE_FILE_PATH)
    for sheet_name, columns in WORKFORCE_SHEET_COLUMNS.items():
        if sheet_name in workbook.sheet_names:
            df = pd.read_excel(workbook, sheet_name=sheet_name)
        else:
            df = pd.DataFrame(columns=columns)
        for col in columns:
            if col not in df.columns:
                df[col] = pd.NA
        data[sheet_name] = df[columns]
    return data


def save_workforce_data(data: Dict[str, pd.DataFrame]) -> None:
    WORKFORCE_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(WORKFORCE_FILE_PATH, engine="openpyxl") as writer:
        for sheet_name, columns in WORKFORCE_SHEET_COLUMNS.items():
            df = data.get(sheet_name, pd.DataFrame(columns=columns)).copy()
            for col in columns:
                if col not in df.columns:
                    df[col] = pd.NA
            df = df[columns]
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def get_workforce_excel_bytes() -> bytes:
    ensure_workforce_master()
    return WORKFORCE_FILE_PATH.read_bytes()


def handle_workforce_upload(uploaded_file) -> Tuple[bool, str]:
    if uploaded_file is None:
        return False, "업로드된 파일이 없습니다."
    try:
        WORKFORCE_FILE_PATH.write_bytes(uploaded_file.getbuffer())
        return True, "엑셀 데이터를 갱신했습니다."
    except Exception as exc:
        return False, f"업로드 처리 중 오류가 발생했습니다: {exc}"


ensure_workforce_master()


@st.cache_data
def load_excluded_workdays() -> pd.DataFrame:
    if not EXCLUDED_WORKING_DAYS_PATH.exists():
        return pd.DataFrame(columns=["년", "월", "제외근무일수"])
    try:
        df = pd.read_csv(EXCLUDED_WORKING_DAYS_PATH, encoding="utf-8-sig")
    except Exception:
        df = pd.read_csv(EXCLUDED_WORKING_DAYS_PATH)
    expected_cols = {"년", "월", "제외근무일수"}
    if not expected_cols.issubset(df.columns):
        return pd.DataFrame(columns=["년", "월", "제외근무일수"])
    clean_df = df[list(expected_cols)].copy()
    clean_df["년"] = pd.to_numeric(clean_df["년"], errors="coerce").astype("Int64")
    clean_df["월"] = pd.to_numeric(clean_df["월"], errors="coerce").astype("Int64")
    clean_df["제외근무일수"] = pd.to_numeric(clean_df["제외근무일수"], errors="coerce").fillna(0).astype(int)
    clean_df = clean_df.dropna(subset=["년", "월"])
    return clean_df.reset_index(drop=True)




def normalize_process_codes(df: pd.DataFrame) -> pd.DataFrame:
    """공정 컬럼의 값을 표준화하고, 컬럼명을 '공정코드'로 통일하며, 안정성을 높입니다."""
    try:
        process_col_name = None
        if '공정코드' in df.columns:
            process_col_name = '공정코드'
        elif '공정' in df.columns:
            process_col_name = '공정'
        else:
            return df
            
        df[process_col_name] = df[process_col_name].astype(str).str.strip()
        process_map = {re.search(r'\[(\d+)\]', name).group(1): name for name in PROCESS_MASTER_ORDER}
        
        def map_process(process_name: str) -> str:
            if not isinstance(process_name, str):
                return process_name
            match = re.search(r'\[(\d+)\]', process_name)
            return process_map.get(match.group(1), process_name) if match else process_name
            
        df[process_col_name] = df[process_col_name].apply(map_process)
        
        if process_col_name == '공정':
            df = df.rename(columns={'공정': '공정코드'})
            
        return df
    except:
        return df

def get_process_order(df: pd.DataFrame, col_name: str = '공정코드') -> List[str]:
    if col_name not in df.columns: return []
    processes_in_df = df[col_name].unique()
    return [p for p in PROCESS_MASTER_ORDER if p in processes_in_df]

def add_date_column(df: pd.DataFrame, date_col_name: Optional[str] = None) -> pd.DataFrame:
    """다양한 날짜 컬럼명을 'date'로 통일하여 새 컬럼을 추가합니다."""
    try:
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            return df
            
        date_candidates = [date_col_name, '생산일자', '일자', '기간'] if date_col_name else ['생산일자', '일자', '기간']
        found_col = next((col for col in date_candidates if col in df.columns), None)
        
        if found_col:
            if found_col == '기간':
                df['date'] = pd.to_datetime(df[found_col].astype(str).str.split(' ~ ').str[0], errors='coerce')
            else:
                df['date'] = pd.to_datetime(df[found_col], errors='coerce')
        else:
            df['date'] = pd.NaT
            
        return df
    except:
        df['date'] = pd.NaT
        return df

def get_resampled_data(
    df: pd.DataFrame, 
    agg_level: str, 
    metrics_to_sum: List[str], 
    group_by_cols: List[str] = ['period', '공장', '공정코드']
) -> pd.DataFrame:
    if df.empty or 'date' not in df.columns or df['date'].isnull().all(): return pd.DataFrame()
    df_copy = df.copy().dropna(subset=['date'])
    if agg_level == '일별':
        df_copy['period'] = df_copy['date'].dt.strftime('%Y-%m-%d')
    elif agg_level == '주간별':
        start_of_week = df_copy['date'] - pd.to_timedelta(df_copy['date'].dt.dayofweek, unit='d')
        end_of_week = start_of_week + pd.to_timedelta(6, unit='d')
        df_copy['period'] = start_of_week.dt.strftime('%Y-%m-%d') + ' ~ ' + end_of_week.dt.strftime('%Y-%m-%d')
    elif agg_level == '월별':
        df_copy['period'] = df_copy['date'].dt.strftime('%Y-%m')
    elif agg_level == '분기별':
        df_copy['period'] = df_copy['date'].dt.year.astype(str) + '년 ' + df_copy['date'].dt.quarter.astype(str) + '분기'
    elif agg_level == '반기별':
        df_copy['period'] = df_copy['date'].dt.year.astype(str) + '년 ' + df_copy['date'].dt.month.apply(lambda m: '상반기' if m <= 6 else '하반기')
    elif agg_level == '년도별':
        df_copy['period'] = df_copy['date'].dt.strftime('%Y')
    else:
        df_copy['period'] = df_copy['date'].dt.strftime('%Y-%m-%d')
        
    valid_group_by_cols = [col for col in group_by_cols if col in df_copy.columns or col == 'period']
    agg_dict = {metric: 'sum' for metric in metrics_to_sum if metric in df_copy.columns}
    if not agg_dict:
        if 'period' not in df_copy.columns: return pd.DataFrame(columns=valid_group_by_cols)
        return df_copy[valid_group_by_cols].drop_duplicates()
    return df_copy.groupby(valid_group_by_cols).agg(agg_dict).reset_index()

def _normalize_personnel_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=PERSONNEL_COLUMNS)

    normalized = df.copy()
    # 컬럼명을 표준 이름으로 통일
    rename_map = {col: PERSONNEL_COLUMN_ALIASES[col] for col in normalized.columns if col in PERSONNEL_COLUMN_ALIASES}
    if rename_map:
        normalized = normalized.rename(columns=rename_map)

    for col in PERSONNEL_COLUMNS:
        if col not in normalized.columns:
            normalized[col] = pd.NA if col == "No." else ""

    normalized = normalized[PERSONNEL_COLUMNS].copy()
    # 숫자 열과 문자열 열을 구분해 정리
    normalized["No."] = pd.to_numeric(normalized["No."], errors="coerce")
    for col in PERSONNEL_COLUMNS:
        if col == "No.":
            continue
        normalized[col] = normalized[col].fillna("").astype(str).str.strip()

    # No.가 비어 있으면 순번 부여
    missing_no = normalized["No."].isna()
    if missing_no.any():
        current_max = normalized["No."].max(skipna=True)
        next_no = 1 if pd.isna(current_max) else int(current_max) + 1
        for idx in normalized[missing_no].index:
            normalized.at[idx, "No."] = next_no
            next_no += 1

    normalized["No."] = normalized["No."].astype("Int64")
    return normalized

def load_personnel_data() -> pd.DataFrame:
    if PERSONNEL_FILE_PATH.exists():
        try:
            df = pd.read_csv(PERSONNEL_FILE_PATH, encoding='utf-8-sig')
        except Exception:
            df = pd.DataFrame(columns=PERSONNEL_COLUMNS)
    else:
        df = pd.DataFrame(columns=PERSONNEL_COLUMNS)
    return _normalize_personnel_dataframe(df)

def save_personnel_data(df: pd.DataFrame) -> None:
    normalized = _normalize_personnel_dataframe(df)
    PERSONNEL_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    normalized.to_csv(PERSONNEL_FILE_PATH, index=False, encoding='utf-8-sig')

def load_personnel_history() -> pd.DataFrame:
    if PERSONNEL_HISTORY_FILE_PATH.exists():
        try:
            df = pd.read_csv(PERSONNEL_HISTORY_FILE_PATH, encoding='utf-8-sig')
        except Exception:
            df = pd.read_csv(PERSONNEL_HISTORY_FILE_PATH)
        rename_map = {
            col: PERSONNEL_HISTORY_COLUMN_ALIASES[col]
            for col in df.columns
            if col in PERSONNEL_HISTORY_COLUMN_ALIASES
        }
        if rename_map:
            df = df.rename(columns=rename_map)
        for col in PERSONNEL_HISTORY_COLUMNS:
            if col not in df.columns:
                df[col] = 0 if col == "등록인원" else ""
        df = df[PERSONNEL_HISTORY_COLUMNS]
        df["등록인원"] = pd.to_numeric(df["등록인원"], errors="coerce").fillna(0).astype(int)
        return df
    return pd.DataFrame(columns=PERSONNEL_HISTORY_COLUMNS)

def save_personnel_history(df: pd.DataFrame) -> None:
    normalized = df.copy()
    normalized = normalized[PERSONNEL_HISTORY_COLUMNS]
    PERSONNEL_HISTORY_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    normalized.to_csv(PERSONNEL_HISTORY_FILE_PATH, index=False, encoding='utf-8-sig')

def ensure_weekly_personnel_snapshot(registered_df: pd.DataFrame) -> None:
    if registered_df is None or registered_df.empty:
        return
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    snapshot_date = today.strftime("%Y-%m-%d")

    history_df = load_personnel_history()
    if not history_df.empty:
        history_dates = (
            pd.to_datetime(history_df["기준일"], errors="coerce")
            .dropna()
            .dt.date
        )
        history_week_starts = {
            d - timedelta(days=d.weekday())
            for d in history_dates
            if isinstance(d, date)
        }
        if week_start in history_week_starts:
            return

    summary = (
        registered_df.groupby(["상위부서명", "부서명"], dropna=False)
        .size()
        .reset_index(name="등록인원")
    )
    if summary.empty:
        return

    summary["기준일"] = snapshot_date
    summary = summary[["기준일", "상위부서명", "부서명", "등록인원"]]
    history_df = pd.concat([history_df, summary], ignore_index=True)
    save_personnel_history(history_df)

def build_monthly_headcount_table(history_df: pd.DataFrame, current_summary: pd.DataFrame) -> pd.DataFrame:
    """상위부서/부서 기준 월별 인원표를 생성합니다."""
    monthly_records = pd.DataFrame(columns=["기준월", "상위부서명", "부서명", "등록인원"])

    if history_df is not None and not history_df.empty:
        hist = history_df.copy()
        hist["기준일"] = pd.to_datetime(hist["기준일"], errors="coerce")
        hist = hist.dropna(subset=["기준일"])
        if not hist.empty:
            hist["기준월"] = hist["기준일"].dt.to_period("M").dt.to_timestamp()
            hist_monthly = (
                hist.groupby(["기준월", "상위부서명", "부서명"], dropna=False)["등록인원"]
                .last()
                .reset_index()
            )
            monthly_records = pd.concat([monthly_records, hist_monthly], ignore_index=True)

    if current_summary is not None and not current_summary.empty:
        current_month = pd.Timestamp.today().to_period("M").to_timestamp()
        latest = current_summary.copy()
        latest["기준월"] = current_month
        monthly_records = pd.concat([monthly_records, latest], ignore_index=True)

    if monthly_records.empty:
        return pd.DataFrame()

    monthly_records = monthly_records.drop_duplicates(
        subset=["기준월", "상위부서명", "부서명"], keep="last"
    )

    pivot = (
        monthly_records.pivot_table(
            index=["상위부서명", "부서명"],
            columns="기준월",
            values="등록인원",
            aggfunc="last",
            fill_value=0,
        )
        .reset_index()
    )

    def format_month(col):
        return col.strftime("%Y년%m월") if isinstance(col, pd.Timestamp) else col

    pivot.columns = [format_month(col) for col in pivot.columns]
    pivot = pivot.sort_values(["상위부서명", "부서명"])

    # 총합계 행 추가
    value_cols = [c for c in pivot.columns if c not in ("상위부서명", "부서명")]
    if value_cols:
        totals = pivot[value_cols].sum()
        total_row = {"상위부서명": "총합계", "부서명": ""}
        total_row.update({col: totals[col] for col in value_cols})
        pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

    # 상위부서명 반복 표시는 공백으로 처리해 계층감을 살림
    pivot["상위부서명"] = pivot["상위부서명"].fillna("")
    pivot["부서명"] = pivot["부서명"].fillna("")
    dup_mask = (pivot["상위부서명"] == pivot["상위부서명"].shift()) & (pivot["상위부서명"] != "총합계")
    pivot.loc[dup_mask, "상위부서명"] = ""

    return pivot

def generate_summary_text(df: pd.DataFrame, agg_level: str, factory_name: str = "전체", raw_data: pd.DataFrame = None) -> str:
    """고급 AI 분석 엔진을 활용한 종합 분석 브리핑 생성"""
    from datetime import datetime
    import calendar
    
    agg_map = {'일별': '일', '주간별': '주', '월별': '월', '분기별': '분기', '반기별': '반기', '년도별': '년'}
    period_text = agg_map.get(agg_level, '기간')
    title_prefix = f"{factory_name} " if factory_name != "전체" else ""
    
    # 데이터 부족 시 처리
    if df.empty or len(df) < 2:
        return f"""<div style="border: 1px solid #e0e0e0; border-radius: 8px; padding: 20px; margin-bottom: 20px; font-family: 'Malgun Gothic', sans-serif; background-color: #f9f9f9; line-height: 1.6;"><h4 style="margin-top:0; color: #1E88E5; font-size: 1.3em;">{title_prefix}AI Analyst 종합 분석 브리핑</h4><p style="font-size: 1.1em;">분석할 데이터가 부족하여 추이 분석을 제공할 수 없습니다. 최소 2개 이상의 {period_text}치 데이터를 선택해주세요.</p></div>"""
    
    df = df.copy().reset_index(drop=True)
    
    # 현재 시점 정보
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    current_day = current_date.day
    
    # 기본 통계
    start_period = df['period'].iloc[0]
    end_period = df['period'].iloc[-1]
    total_prod = df['총_생산수량'].sum()
    avg_prod = df['총_생산수량'].mean()
    avg_yield = df['종합수율(%)'].mean()
    
    # === 1. 시점 인식 분석 ===
    def analyze_timing_context():
        """현재 시점을 고려한 맥락 분석"""
        context_insights = []
        
        # 월별 데이터인 경우 진행중인 월 식별
        if agg_level == '월별':
            latest_period = end_period
            try:
                if '2025' in latest_period and f'{current_month:02d}' in latest_period:
                    if current_day <= 15:  # 월 중순 이전
                        context_insights.append(f"⚠️ <strong>진행중 데이터 주의:</strong> {latest_period}은 현재 진행중인 월로, {current_day}일 현재까지의 부분 데이터입니다.")
                    else:  # 월 중순 이후
                        context_insights.append(f"📊 <strong>거의 완성된 데이터:</strong> {latest_period}은 {current_day}일 현재까지 누적된 데이터로, 월말 예상치에 근접합니다.")
            except:
                pass
        
        return context_insights
    
    # === 2. 고급 트렌드 분석 ===
    def analyze_advanced_trends():
        """선형 회귀, 가속도, 변곡점을 활용한 트렌드 분석"""
        insights = []
        
        if len(df) >= 3:
            # 선형 회귀를 통한 트렌드 분석
            x_vals = np.arange(len(df))
            
            # 생산량 트렌드
            prod_slope, prod_intercept = np.polyfit(x_vals, df['총_생산수량'], 1)
            prod_r_squared = np.corrcoef(x_vals, df['총_생산수량'])[0, 1] ** 2
            
            # 수율 트렌드  
            yield_slope, yield_intercept = np.polyfit(x_vals, df['종합수율(%)'], 1)
            yield_r_squared = np.corrcoef(x_vals, df['종합수율(%)'])[0, 1] ** 2
            
            # 트렌드 강도 판단
            def get_trend_strength(r_squared):
                if r_squared >= 0.8: return "매우 강한"
                elif r_squared >= 0.6: return "강한" 
                elif r_squared >= 0.4: return "중간"
                elif r_squared >= 0.2: return "약한"
                else: return "불규칙한"
            
            prod_trend_strength = get_trend_strength(prod_r_squared)
            yield_trend_strength = get_trend_strength(yield_r_squared)
            
            # 생산량 트렌드 인사이트
            if abs(prod_slope) > avg_prod * 0.05:  # 평균의 5% 이상 기울기
                trend_direction = "상승" if prod_slope > 0 else "하락"
                insights.append(f"📈 <strong>생산량 {trend_direction} 트렌드:</strong> {prod_trend_strength} {trend_direction} 추세 (결정계수: {prod_r_squared:.2f})")
            
            # 수율 트렌드 인사이트
            if abs(yield_slope) > 1.0:  # 1% 이상 기울기
                trend_direction = "개선" if yield_slope > 0 else "악화"
                insights.append(f"⚙️ <strong>수율 {trend_direction} 트렌드:</strong> {yield_trend_strength} {trend_direction} 추세 (결정계수: {yield_r_squared:.2f})")
            
            # 가속도 분석 (최근 3개 구간)
            if len(df) >= 4:
                recent_prod = df['총_생산수량'].tail(3).values
                recent_changes = np.diff(recent_prod)
                if len(recent_changes) >= 2:
                    acceleration = recent_changes[-1] - recent_changes[-2]
                    if abs(acceleration) > avg_prod * 0.1:
                        accel_text = "가속화" if acceleration > 0 else "둔화"
                        insights.append(f"🚀 <strong>변화 가속도:</strong> 최근 생산량 변화가 {accel_text}되고 있습니다.")
        
        return insights
    
    # === 3. 계절성 및 패턴 분석 ===
    def analyze_seasonality():
        """월별, 계절별 패턴 분석"""
        insights = []
        
        if agg_level == '월별' and len(df) >= 6:
            # 월별 평균 성과 계산
            month_performance = {}
            for idx, row in df.iterrows():
                try:
                    period_str = str(row['period'])
                    if len(period_str.split('.')) >= 2:
                        month = int(period_str.split('.')[1])
                        if month not in month_performance:
                            month_performance[month] = []
                        month_performance[month].append(row['종합수율(%)'])
                except:
                    continue
            
            if len(month_performance) >= 3:
                month_avg = {month: np.mean(values) for month, values in month_performance.items()}
                best_month = max(month_avg, key=month_avg.get)
                worst_month = min(month_avg, key=month_avg.get)
                
                month_names = ['', '1월', '2월', '3월', '4월', '5월', '6월', 
                              '7월', '8월', '9월', '10월', '11월', '12월']
                
                insights.append(f"📅 <strong>계절성 패턴:</strong> {month_names[best_month]}이 평균 수율이 가장 높고({month_avg[best_month]:.1f}%), {month_names[worst_month]}이 가장 낮습니다({month_avg[worst_month]:.1f}%)")
        
        return insights
    
    # === 4. 이상 징후 탐지 ===
    def detect_anomalies():
        """Z-score를 활용한 이상 징후 탐지"""
        insights = []
        
        # 생산량 이상치
        prod_mean = df['총_생산수량'].mean()
        prod_std = df['총_생산수량'].std()
        
        if prod_std > 0:
            df['prod_zscore'] = (df['총_생산수량'] - prod_mean) / prod_std
            extreme_prod = df[abs(df['prod_zscore']) > 2]  # 2시그마 이상
            
            if not extreme_prod.empty:
                for idx, row in extreme_prod.iterrows():
                    anomaly_type = "급증" if row['prod_zscore'] > 0 else "급감"
                    insights.append(f"⚡ <strong>생산량 이상 징후:</strong> {row['period']}에 생산량 {anomaly_type} (Z-score: {row['prod_zscore']:.1f})")
        
        # 수율 이상치
        yield_mean = df['종합수율(%)'].mean()
        yield_std = df['종합수율(%)'].std()
        
        if yield_std > 0:
            df['yield_zscore'] = (df['종합수율(%)'] - yield_mean) / yield_std
            extreme_yield = df[abs(df['yield_zscore']) > 2]
            
            if not extreme_yield.empty:
                for idx, row in extreme_yield.iterrows():
                    anomaly_type = "급상승" if row['yield_zscore'] > 0 else "급하락"
                    insights.append(f"⚡ <strong>수율 이상 징후:</strong> {row['period']}에 수율 {anomaly_type} (Z-score: {row['yield_zscore']:.1f})")
        
        return insights
    
    # === 5. 예측 및 전망 ===
    def generate_forecast():
        """현재 트렌드 기반 다음 기간 예측"""
        insights = []
        
        if len(df) >= 3:
            # 최근 3개 기간 기준 단순 선형 예측
            recent_df = df.tail(3)
            x_vals = np.arange(len(recent_df))
            
            try:
                # 생산량 예측
                prod_slope, prod_intercept = np.polyfit(x_vals, recent_df['총_생산수량'], 1)
                next_prod = prod_slope * len(recent_df) + prod_intercept
                
                # 수율 예측
                yield_slope, yield_intercept = np.polyfit(x_vals, recent_df['종합수율(%)'], 1)
                next_yield = yield_slope * len(recent_df) + yield_intercept
                
                current_prod = df['총_생산수량'].iloc[-1]
                current_yield = df['종합수율(%)'].iloc[-1]
                
                prod_change_pred = (next_prod - current_prod) / current_prod * 100
                yield_change_pred = next_yield - current_yield
                
                if abs(prod_change_pred) > 5:  # 5% 이상 변화 예상
                    trend_text = "증가" if prod_change_pred > 0 else "감소"
                    insights.append(f"🔮 <strong>다음 기간 전망:</strong> 현재 추세 유지 시 생산량 {abs(prod_change_pred):.1f}% {trend_text} 예상")
                
                if abs(yield_change_pred) > 1:  # 1% 이상 변화 예상
                    trend_text = "개선" if yield_change_pred > 0 else "악화"
                    insights.append(f"🔮 <strong>수율 전망:</strong> 현재 추세라면 수율 {abs(yield_change_pred):.1f}%p {trend_text} 예상")
                
            except:
                pass
        
        return insights
    
    # === 6. 실행 가능한 비즈니스 인사이트 ===
    def generate_actionable_insights():
        """구체적이고 실행 가능한 개선 방안 제시"""
        insights = []
        
        # 최고 성과 기간 분석
        best_period = df.loc[df['종합수율(%)'].idxmax()]
        worst_period = df.loc[df['종합수율(%)'].idxmin()]
        
        yield_gap = best_period['종합수율(%)'] - worst_period['종합수율(%)']
        
        if yield_gap > 5:  # 5% 이상 격차
            insights.append(f"🎯 <strong>개선 포텐셜:</strong> 최고 성과({best_period['period']}: {best_period['종합수율(%)']:.1f}%)와 최저 성과 간 {yield_gap:.1f}%p 격차로, 표준화를 통한 개선 여지가 큽니다.")
        
        # 생산량-수율 관계 분석
        correlation = df['총_생산수량'].corr(df['종합수율(%)'])
        
        if correlation > 0.6:
            insights.append(f"✅ <strong>긍정적 시너지:</strong> 생산량 증대와 수율 향상이 동시에 달성 가능한 구조입니다. (상관계수: {correlation:.2f})")
        elif correlation < -0.6:
            insights.append(f"⚠️ <strong>트레이드오프 관리:</strong> 생산량 증대 시 수율 저하 위험이 있어 품질관리 강화가 필요합니다. (상관계수: {correlation:.2f})")
        
        # 최근 성과 평가
        recent_yield = df['종합수율(%)'].tail(2).mean()
        overall_yield = df['종합수율(%)'].mean()
        
        if recent_yield > overall_yield + 2:
            insights.append(f"📈 <strong>개선 모멘텀:</strong> 최근 성과가 전체 평균보다 {recent_yield - overall_yield:.1f}%p 높아 긍정적 추세입니다.")
        elif recent_yield < overall_yield - 2:
            insights.append(f"📉 <strong>주의 신호:</strong> 최근 성과가 전체 평균보다 {overall_yield - recent_yield:.1f}%p 낮아 원인 분석이 필요합니다.")
        
        return insights
    
    # 모든 분석 실행
    timing_insights = analyze_timing_context()
    trend_insights = analyze_advanced_trends()  
    seasonal_insights = analyze_seasonality()
    anomaly_insights = detect_anomalies()
    forecast_insights = generate_forecast()
    actionable_insights = generate_actionable_insights()
    
    # 모든 인사이트 통합
    all_insights = timing_insights + trend_insights + seasonal_insights + anomaly_insights + forecast_insights + actionable_insights
    
    # 기본 통계 정리
    max_prod_row = df.loc[df['총_생산수량'].idxmax()]
    min_prod_row = df.loc[df['총_생산수량'].idxmin()]
    max_yield_row = df.loc[df['종합수율(%)'].idxmax()]
    min_yield_row = df.loc[df['종합수율(%)'].idxmin()]
    
    # 브리핑 HTML 생성
    insights_html = ""
    if all_insights:
        insights_html = "<ul style='list-style-type: none; padding-left: 0; margin: 15px 0;'>"
        for insight in all_insights[:6]:  # 최대 6개 인사이트
            insights_html += f"<li style='margin-bottom: 8px; padding: 8px; background: rgba(30, 136, 229, 0.1); border-left: 3px solid #1E88E5; border-radius: 4px;'>{insight}</li>"
        insights_html += "</ul>"
    
    summary = f"""
<div style="border: 1px solid #e0e0e0; border-radius: 8px; padding: 20px; margin-bottom: 20px; font-family: 'Malgun Gothic', sans-serif; background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); line-height: 1.6; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
    <h4 style="margin-top:0; color: #1E88E5; font-size: 1.4em; font-weight: bold; border-bottom: 2px solid #1E88E5; padding-bottom: 10px;">🤖 {title_prefix}AI Analyst 종합 분석 브리핑 ({agg_level})</h4>
    
    <div style="background: white; padding: 15px; border-radius: 6px; margin: 15px 0; border: 1px solid #dee2e6;">
        <p style="font-size: 1.0em; margin-bottom: 15px;"><strong>📊 분석 기간:</strong> {start_period} ~ {end_period}</p>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin-bottom: 15px;">
            <div style="background: #f8f9fa; padding: 12px; border-radius: 4px; border-left: 4px solid #28a745;">
                <div style="font-size: 0.9em; color: #6c757d;">총 생산량</div>
                <div style="font-size: 1.3em; font-weight: bold; color: #28a745;">{total_prod:,.0f}개</div>
                <div style="font-size: 0.8em; color: #6c757d;">{period_text} 평균: {avg_prod:,.0f}개</div>
            </div>
            <div style="background: #f8f9fa; padding: 12px; border-radius: 4px; border-left: 4px solid #ffc107;">
                <div style="font-size: 0.9em; color: #6c757d;">평균 종합수율</div>
                <div style="font-size: 1.3em; font-weight: bold; color: #e67e22;">{avg_yield:.2f}%</div>
                <div style="font-size: 0.8em; color: #6c757d;">최고: {max_yield_row['종합수율(%)']:.1f}% ({max_yield_row['period']})</div>
            </div>
        </div>
        
        <div style="font-size: 0.9em; color: #6c757d; margin-bottom: 10px;">
            📈 <strong>생산량:</strong> 최고 {max_prod_row['period']} ({max_prod_row['총_생산수량']:,.0f}개) | 최저 {min_prod_row['period']} ({min_prod_row['총_생산수량']:,.0f}개)<br>
            ⚙️ <strong>수율:</strong> 최고 {max_yield_row['period']} ({max_yield_row['종합수율(%)']:.1f}%) | 최저 {min_yield_row['period']} ({min_yield_row['종합수율(%)']:.1f}%)
        </div>
    </div>
    
    <div style="margin-top: 20px;">
        <h5 style="color: #dc3545; font-size: 1.2em; margin-bottom: 15px; font-weight: bold;">🔍 핵심 인사이트 & 액션 아이템</h5>
        {insights_html if all_insights else '<p style="color: #6c757d; font-style: italic;">추가 데이터가 필요하여 고급 분석을 수행할 수 없습니다.</p>'}
    </div>
    
    <div style="margin-top: 20px; padding: 15px; background: rgba(220, 53, 69, 0.1); border: 1px solid #dc3545; border-radius: 6px;">
        <p style="font-size: 0.9em; color: #495057; margin: 0;">
            <strong>💡 AI 분석 기준:</strong> 트렌드 분석(선형회귀), 이상징후 탐지(Z-score ≥2σ), 계절성 패턴, 예측 모델링을 종합하여 생성된 브리핑입니다.
        </p>
    </div>
</div>
"""
    return summary

def create_line_chart(df: pd.DataFrame, x: str, y: str, color: Optional[str] = None, 
                     title: str = "", markers: bool = True, text: Optional[str] = None) -> go.Figure:
    """공통 라인 차트 생성 함수"""
    try:
        fig = px.line(df, x=x, y=y, color=color, title=title, markers=markers, text=text)
        if text:
            fig.update_traces(
                texttemplate='%{text:.2f}%', 
                textposition='top center', 
                textfont=dict(size=TEXT_FONT_SIZE, color='black')
            )
        fig.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(df[x].unique()))
        return fig
    except:
        return go.Figure().update_layout(title=title)

def create_bar_chart(df: pd.DataFrame, x: str, y: str, color: Optional[str] = None, 
                    title: str = "", orientation: str = 'v') -> go.Figure:
    """공통 막대 차트 생성 함수"""
    try:
        fig = px.bar(df, x=x, y=y, color=color, title=title, orientation=orientation, height=CHART_HEIGHT)
        fig.update_traces(textposition='auto')
        return fig
    except:
        return go.Figure().update_layout(title=title)

def plot_pareto_chart(df: pd.DataFrame, title: str, defect_qty_col: str = '유형별_봨0량수량') -> go.Figure:
    """파레토 차트 생성"""
    if df.empty or defect_qty_col not in df.columns: 
        st.info("차트를 그릴 데이터가 없습니다.")
        return go.Figure().update_layout(title=title)
    df_agg = df.groupby('불량명')[defect_qty_col].sum().reset_index()
    df_agg = df_agg.sort_values(by=defect_qty_col, ascending=False)
    df_agg = df_agg[df_agg[defect_qty_col] > 0] 
    if df_agg.empty: 
        st.info("선택된 항목에 보고된 불량이 없습니다.")
        return
    df_agg['누적합계'] = df_agg[defect_qty_col].cumsum()
    df_agg['누적비율'] = (df_agg['누적합계'] / df_agg[defect_qty_col].sum()) * 100
    
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    
    fig.add_trace(go.Bar(
        x=df_agg['불량명'], 
        y=df_agg[defect_qty_col], 
        name='불량 수량', 
        text=df_agg[defect_qty_col], 
        texttemplate='%{text:,.0f}', 
        textposition='outside',
        textfont=dict(size=18, family="Arial, sans-serif", color="black")
    ), secondary_y=False)
    
    fig.add_trace(go.Scatter(
        x=df_agg['불량명'], 
        y=df_agg['누적비율'], 
        name='누적 비율', 
        mode='lines+markers+text',
        text=df_agg['누적비율'], 
        texttemplate='%{text:.1f}%', 
        textposition='top center',
        textfont=dict(size=16, color='black') 
    ), secondary_y=True)
    
    fig.update_layout(height=600, title_text=f'<b>{title}</b>', margin=dict(t=120), legend=dict(orientation="h", yanchor="bottom", y=1.10, xanchor="right", x=1))
    fig.update_yaxes(title_text="<b>불량 수량 (개)</b>", secondary_y=False)
    fig.update_yaxes(title_text="<b>누적 비율 (%)</b>", secondary_y=True, range=[0, 105])
    fig.update_xaxes(title_text="<b>불량 유형</b>")
    st.plotly_chart(fig, use_container_width=True)

def get_year_boundaries(reference_date: Optional[date], min_data_date: date, max_data_date: date) -> Tuple[date, date]:
    """주어진 참조일이 속한 연도의 시작/끝을 데이터 범위에 맞춰 반환합니다."""
    target_date = reference_date or date.today()
    year_start = date(target_date.year, 1, 1)
    year_end = date(target_date.year, 12, 31)
    start = max(min_data_date, year_start)
    end = min(max_data_date, year_end)
    if start > end:
        start, end = min_data_date, max_data_date
    return start, end


def reset_filters(min_data_date, max_data_date, reference_date=None):
    """집계 기준을 월별로 두고, 조회 연도를 기준으로 기간을 재설정합니다."""
    start, end = get_year_boundaries(
        reference_date or st.session_state.get('range_reference_date'),
        min_data_date,
        max_data_date,
    )
    st.session_state.date_range = (start, end)
    st.session_state.agg_level = '월별'
    st.session_state.range_reference_date = end


def set_maximum_period(min_data_date, max_data_date):
    """데이터의 전체 기간으로 조회 범위를 확장합니다."""
    st.session_state.date_range = (min_data_date, max_data_date)
    st.session_state.agg_level = '월별'
    st.session_state.range_reference_date = max_data_date

# --- 대시보드 UI 시작 ---
st.title("👑 지능형 생산 대시보드 2021~2022 전용")

all_data = load_all_data()
df_target_orig, target_filename = all_data.get('target', (pd.DataFrame(), None)); df_yield_orig, yield_filename = all_data.get('yield', (pd.DataFrame(), None)); df_utilization_orig, util_filename = all_data.get('utilization', (pd.DataFrame(), None)); df_defect_orig, defect_filename = all_data.get('defect', (pd.DataFrame(), None))

if not df_target_orig.empty: 
    df_target_orig = normalize_process_codes(add_date_column(df_target_orig))
    # 배합 공정 제외
    if '공정코드' in df_target_orig.columns:
        df_target_orig = df_target_orig[~df_target_orig['공정코드'].str.contains('배합', na=False)]
if not df_yield_orig.empty: 
    df_yield_orig = normalize_process_codes(add_date_column(df_yield_orig))
    # 배합 공정 제외
    if '공정코드' in df_yield_orig.columns:
        df_yield_orig = df_yield_orig[~df_yield_orig['공정코드'].str.contains('배합', na=False)]
if not df_utilization_orig.empty: 
    df_utilization_orig = normalize_process_codes(add_date_column(df_utilization_orig))
    # 배합 공정 제외 (불필요한 데이터)
    df_utilization_orig = df_utilization_orig[~df_utilization_orig['공정코드'].str.contains('배합', na=False)]
    # 이론상 생산량이 0인 레코드 처리 (실제 생산이 있으면 가동률을 별도 계산)
    mask_zero_theory = df_utilization_orig['이론상_총_생산량'] == 0
    df_utilization_orig.loc[mask_zero_theory & (df_utilization_orig['총_생산수량'] > 0), '가동률(%)'] = pd.NA
    # 가동률이 계산 불가능한 경우 필터링 옵션을 위해 표시 컬럼 추가
    df_utilization_orig['계산가능'] = ~mask_zero_theory
if not df_defect_orig.empty: 
    df_defect_orig = normalize_process_codes(add_date_column(df_defect_orig))
    # 배합 공정 제외
    if '공정코드' in df_defect_orig.columns:
        df_defect_orig = df_defect_orig[~df_defect_orig['공정코드'].str.contains('배합', na=False)]

# === 일일 생산 현황 보고 탭 전용 데이터 생성 (완전 독립) ===
@st.cache_data
def create_daily_report_dataset():
    """일일 생산 현황 보고 탭 전용 데이터셋 생성 - 다른 탭들과 완전 분리"""
    if df_target_orig.empty or df_yield_orig.empty:
        return pd.DataFrame(), pd.DataFrame()
    
    # 전체 기간 설정 (고정)
    all_dates_daily = pd.concat([df_target_orig['date'], df_yield_orig['date']]).dropna()
    if all_dates_daily.empty:
        return pd.DataFrame(), pd.DataFrame()
    
    daily_start_date = all_dates_daily.min().date() 
    daily_end_date = all_dates_daily.max().date()
    
    # 목표 데이터 필터링 (전체 기간)
    mask_target = (df_target_orig['date'].dt.date >= daily_start_date) & (df_target_orig['date'].dt.date <= daily_end_date)
    daily_target_data = df_target_orig[mask_target].copy()
    
    # 수율 데이터 필터링 (전체 기간)  
    mask_yield = (df_yield_orig['date'].dt.date >= daily_start_date) & (df_yield_orig['date'].dt.date <= daily_end_date)
    daily_yield_data = df_yield_orig[mask_yield].copy()
    
    return daily_target_data, daily_yield_data

# 일일 보고서 전용 데이터 생성 (프로그램 시작 시 한 번만)
daily_report_target_data, daily_report_yield_data = create_daily_report_dataset()

if 'date_range' not in st.session_state or 'agg_level' not in st.session_state:
    all_dfs = [df_target_orig, df_yield_orig, df_utilization_orig, df_defect_orig]
    all_dates = pd.concat([d['date'] for d in all_dfs if d is not None and not d.empty and 'date' in d.columns]).dropna()
    min_date_global, max_date_global = (all_dates.min().date(), all_dates.max().date()) if not all_dates.empty else (date.today(), date.today())
    if 'date_range' not in st.session_state: st.session_state.date_range = (min_date_global, max_date_global)
    if 'agg_level' not in st.session_state: st.session_state.agg_level = '월별'

if 'range_reference_date' not in st.session_state:
    default_reference = st.session_state.date_range[1] if 'date_range' in st.session_state else date.today()
    st.session_state.range_reference_date = default_reference

st.sidebar.header("로딩된 파일 정보")
st.sidebar.info(f"수율: {yield_filename}" if yield_filename else "수율: 파일 없음")
st.sidebar.info(f"목표: {target_filename}" if target_filename else "목표: 미사용 (21~22 전용 구성)")
st.sidebar.info(f"가동률: {util_filename}" if util_filename else "가동률: 미사용 (21~22 전용 구성)")
st.sidebar.info(f"불량: {defect_filename}" if defect_filename else "불량: 미사용 (21~22 전용 구성)")

tab_list = ["수율 분석", "종합 분석", "생산실적 상세조회"]
selected_tab = st.radio("메인 네비게이션", tab_list, key='main_tab_selector', horizontal=True, label_visibility='collapsed')

# === 탭 전환 감지 및 설정 보정 시스템 ===
def manage_tab_transitions():
    """일일 생산 현황 보고 탭 전환 시 기간/집계기준 설정 보정"""
    current_tab = selected_tab
    previous_tab = st.session_state.get('previous_tab', None)
    
    # 전체 데이터 기간 계산
    all_dfs = [df_target_orig, df_yield_orig, df_utilization_orig, df_defect_orig]
    all_dates = pd.concat([d['date'] for d in all_dfs if d is not None and not d.empty and 'date' in d.columns]).dropna()
    if not all_dates.empty:
        full_start_date, full_end_date = all_dates.min().date(), all_dates.max().date()
    else:
        full_start_date, full_end_date = date.today(), date.today()

    reference_date_for_year = (
        st.session_state.get('daily_reference_date')
        or st.session_state.get('range_reference_date')
        or full_end_date
    )

    # 패턴 1: 프로그램 시작 또는 일일 생산 현황 보고 탭에서 다른 탭으로 이동
    if (previous_tab is None or previous_tab == "📊 일일 생산 현황 보고") and current_tab != "📊 일일 생산 현황 보고":
        # 저장된 설정이 있으면 복원, 없으면 조회 연도 전체로 설정
        if 'saved_date_range' in st.session_state and 'saved_agg_level' in st.session_state:
            st.session_state.date_range = st.session_state.saved_date_range
            st.session_state.agg_level = st.session_state.saved_agg_level
            if isinstance(st.session_state.saved_date_range, (list, tuple)) and len(st.session_state.saved_date_range) == 2:
                st.session_state.range_reference_date = st.session_state.saved_date_range[1]
        else:
            reset_filters(full_start_date, full_end_date, reference_date_for_year)
    
    # 패턴 2: 다른 탭에서 일일 보고서 탭으로 이동 (현재 설정 저장)
    elif current_tab == "📊 일일 생산 현황 보고" and previous_tab != "📊 일일 생산 현황 보고" and previous_tab is not None:
        # 현재 설정을 저장 (나중에 복원용)
        if 'date_range' in st.session_state and 'agg_level' in st.session_state:
            st.session_state.saved_date_range = st.session_state.date_range
            st.session_state.saved_agg_level = st.session_state.agg_level
    
    # 현재 탭을 이전 탭으로 저장 (다음 번 비교용)
    st.session_state.previous_tab = current_tab

# 탭 전환 관리 실행
manage_tab_transitions()

def render_personnel_section() -> None:
    if PERSONNEL_FEEDBACK_KEY in st.session_state:
        st.success(st.session_state[PERSONNEL_FEEDBACK_KEY])
        del st.session_state[PERSONNEL_FEEDBACK_KEY]

    personnel_df = load_personnel_data()
    for col in PERSONNEL_COLUMNS:
        if col != "No." and col in personnel_df.columns:
            personnel_df[col] = personnel_df[col].fillna("").astype(str).str.strip()

    name_series = personnel_df["성명"].fillna("").astype(str).str.strip() if "성명" in personnel_df else pd.Series(dtype=str)
    registered_mask = name_series.astype(bool)
    registered_df = personnel_df[registered_mask].copy()
    if not registered_df.empty:
        registered_df.loc[:, "성명"] = name_series[registered_mask].values
        for col in ["상위부서명", "부서명", "사번", "직위", "직책"]:
            if col in registered_df.columns:
                registered_df.loc[:, col] = registered_df[col].fillna("").astype(str).str.strip()

    ensure_weekly_personnel_snapshot(registered_df)

    history_df = load_personnel_history()
    history_monthly = pd.DataFrame(columns=["월", "상위부서명", "부서명", "등록인원", "기준일"])
    if not history_df.empty:
        history_proc = history_df.copy()
        history_proc["기준일"] = pd.to_datetime(history_proc["기준일"], errors="coerce")
        history_proc = history_proc.dropna(subset=["기준일"])
        if not history_proc.empty:
            for col in ["상위부서명", "부서명"]:
                if col in history_proc.columns:
                    history_proc[col] = history_proc[col].fillna("").astype(str).str.strip()
            history_proc["등록인원"] = pd.to_numeric(history_proc["등록인원"], errors="coerce").fillna(0).astype(int)
            history_proc = history_proc.sort_values("기준일")
            history_proc["월"] = history_proc["기준일"].dt.to_period("M").dt.to_timestamp()
            history_monthly = history_proc.drop_duplicates(subset=["월", "상위부서명", "부서명"], keep="last")
            history_monthly = history_monthly[["월", "상위부서명", "부서명", "등록인원", "기준일"]]

    current_summary = pd.DataFrame(columns=["상위부서명", "부서명", "등록인원"])
    if not registered_df.empty:
        current_summary = (
            registered_df.groupby(["상위부서명", "부서명"], dropna=False)
            .size()
            .reset_index(name="등록인원")
        )
        current_summary["상위부서명"] = current_summary["상위부서명"].fillna("").astype(str).str.strip()
        current_summary["부서명"] = current_summary["부서명"].fillna("").astype(str).str.strip()

    monthly_headcount_table = build_monthly_headcount_table(history_df, current_summary)

    upper_options = sorted({str(name).strip() for name in personnel_df["상위부서명"].tolist() if str(name).strip()})
    dept_options = sorted({str(name).strip() for name in personnel_df["부서명"].tolist() if str(name).strip()})
    title_options = sorted({str(name).strip() for name in personnel_df["직위"].tolist() if str(name).strip()}) if "직위" in personnel_df else []
    duty_options = sorted({str(name).strip() for name in personnel_df["직책"].tolist() if str(name).strip()}) if "직책" in personnel_df else []

    st.markdown("#### 부서별 인원 현황 (월별)")

    if monthly_headcount_table.empty:
        st.info("등록된 인원 데이터 또는 이력이 없어 월별 인원표를 만들 수 없습니다. 파일을 업데이트하면 자동으로 반영됩니다.")
    else:
        st.dataframe(monthly_headcount_table, use_container_width=True, hide_index=True)
        st.caption("상위부서명-부서명 기준 월별 인원 현황입니다. 최신 월은 현재 파일 기준으로 갱신됩니다.")

    st.markdown("#### 현재 인원 현황")

    if personnel_df.empty:
        st.info("등록된 인원 정보가 없습니다.")
        return

    utility_cols = st.columns(2)
    with utility_cols[0]:
        summary_export = current_summary.copy()
        export_bytes = None
        export_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        export_name = "인원현황.xlsx"

        for engine in ("xlsxwriter", "openpyxl"):
            buffer = io.BytesIO()
            try:
                with pd.ExcelWriter(buffer, engine=engine) as writer:
                    personnel_df.to_excel(writer, index=False, sheet_name="인원현황")
                    if not summary_export.empty:
                        summary_export.to_excel(writer, index=False, sheet_name="상위부서_부서별")
                    if not monthly_headcount_table.empty:
                        monthly_headcount_table.to_excel(writer, index=False, sheet_name="월별현황")
                export_bytes = buffer.getvalue()
                break
            except Exception:
                continue

        if export_bytes is None:
            export_bytes = personnel_df.to_csv(index=False).encode("utf-8-sig")
            export_mime = "text/csv"
            export_name = "인원현황.csv"

        st.download_button(
            label="인원현황 다운로드",
            data=export_bytes,
            file_name=export_name,
            mime=export_mime
        )
    with utility_cols[1]:
        st.metric("등록 인원 수", f"{len(registered_df)}명")
        st.caption("성명이 입력된 인원만 집계합니다.")

    st.markdown("##### 상위부서·부서별 현재 인원")
    if current_summary.empty:
        st.info("등록된 인원이 없어 현황을 표시할 수 없습니다. 성명을 입력하면 자동으로 반영됩니다.")
    else:
        current_display = current_summary.rename(columns={"등록인원": "현재 인원"}).sort_values(["상위부서명", "부서명"])
        st.dataframe(current_display, use_container_width=True, hide_index=True)

    st.markdown("##### 인원 이력 트렌드 (월별)")
    if history_monthly.empty:
        st.info("이력 데이터가 없습니다. 신규 데이터가 추가되면 최신 월 기준으로 자동 기록됩니다.")
    else:
        filters = st.columns(2)
        with filters[0]:
            upper_options_hist = sorted({team for team in history_monthly["상위부서명"].unique() if team})
            upper_select = st.selectbox(
                "상위부서 선택",
                ["전체"] + upper_options_hist,
                key="personnel_history_upper"
            )

        filtered_monthly = history_monthly.copy()
        if upper_select != "전체":
            filtered_monthly = filtered_monthly[filtered_monthly["상위부서명"] == upper_select]

        with filters[1]:
            dept_options_hist = sorted({proc for proc in history_monthly["부서명"].unique() if proc})
            dept_select = st.selectbox(
                "부서 선택",
                ["전체"] + dept_options_hist,
                key="personnel_history_dept"
            )
        if dept_select != "전체":
            filtered_monthly = filtered_monthly[filtered_monthly["부서명"] == dept_select]

        chart_placeholder = st.empty()
        if filtered_monthly.empty:
            chart_placeholder.info("선택한 조건에 해당하는 이력 데이터가 없습니다.")
        else:
            if upper_select == "전체":
                chart_df = (
                    filtered_monthly.groupby(["월", "상위부서명"], dropna=False)["등록인원"]
                    .sum()
                    .reset_index()
                )
                if chart_df.empty:
                    chart_placeholder.info("표시할 데이터가 없습니다.")
                else:
                    chart_df = chart_df.sort_values("월")
                    fig_trend = px.line(
                        chart_df,
                        x="월",
                        y="등록인원",
                        color="상위부서명",
                        markers=True,
                        title="상위부서별 월간 인원 추이"
                    )
                    fig_trend.update_layout(
                        height=420,
                        xaxis_title="월",
                        yaxis_title="등록 인원(명)",
                        legend_title="상위부서명"
                    )
                    chart_placeholder.plotly_chart(fig_trend, use_container_width=True)
            else:
                team_monthly = filtered_monthly.copy()
                chart_df = (
                    team_monthly.groupby(["월", "부서명"], dropna=False)["등록인원"]
                    .sum()
                    .reset_index()
                )
                chart_df = chart_df.sort_values("월")
                fig_trend = px.line(
                    chart_df,
                    x="월",
                    y="등록인원",
                    color="부서명",
                    markers=True,
                    title=f"{upper_select} 부서별 월간 인원 추이"
                )
                fig_trend.update_layout(
                    height=420,
                    xaxis_title="월",
                    yaxis_title="등록 인원(명)",
                    legend_title="부서명"
                )
                chart_placeholder.plotly_chart(fig_trend, use_container_width=True)

            latest_month = filtered_monthly["월"].max()
            if pd.notna(latest_month):
                latest_summary = (
                    filtered_monthly[filtered_monthly["월"] == latest_month]
                    .groupby(["상위부서명", "부서명"], dropna=False)["등록인원"]
                    .sum()
                    .reset_index()
                    .sort_values(["상위부서명", "부서명"])
                )
                display_table = filtered_monthly.copy().sort_values(["월", "상위부서명", "부서명"])
                display_table["월"] = display_table["월"].dt.strftime("%Y-%m")
                latest_month_str = latest_month.strftime("%Y-%m")
                st.caption(f"• 최신 기록 기준 ({latest_month_str}) 상위부서/부서별 등록 인원")
                st.dataframe(latest_summary, use_container_width=True, hide_index=True)
                with st.expander("월별 상세 내역", expanded=False):
                    st.dataframe(
                        display_table[["월", "상위부서명", "부서명", "등록인원"]],
                        use_container_width=True,
                        hide_index=True
                    )

    upper_names = sorted({name for name in personnel_df["상위부서명"].tolist() if str(name).strip()})
    if upper_names:
        markdown_lines = "\n".join(f"- {team}" for team in upper_names)
        st.markdown("**등록된 상위부서 목록**\n" + markdown_lines)

    if registered_df.empty:
        st.info("등록된 인원 정보가 없습니다. 아래 인원 등록 영역을 활용해 인원을 추가해 주세요.")
        filtered_df = pd.DataFrame(columns=PERSONNEL_COLUMNS)
        edited_df = None
    else:
        search_name = st.text_input("성명 검색", key="personnel_search_name", placeholder="찾을 성명을 입력하세요.")
        filtered_df = registered_df.copy()
        if search_name:
            filtered_df = filtered_df[
                filtered_df["성명"].str.contains(search_name, case=False, na=False)
            ].copy()
            if filtered_df.empty:
                st.info(f"'{search_name}' 성명을 가진 인원을 찾지 못했습니다.")

        if filtered_df.empty:
            edited_df = None
        else:
            delete_view_df = filtered_df.copy()
            delete_view_df.insert(0, "삭제", False)

            edited_df = st.data_editor(
                delete_view_df,
                column_config={
                    "삭제": st.column_config.CheckboxColumn(
                        "삭제",
                        help="삭제할 인원을 선택하세요.",
                        default=False
                    )
                },
                disabled=PERSONNEL_COLUMNS,
                hide_index=True,
                use_container_width=True,
                key="personnel_delete_editor"
            )

    if st.button("선택 인원 삭제"):
        if isinstance(edited_df, pd.DataFrame):
            rows_to_delete = edited_df[edited_df["삭제"]]
            if rows_to_delete.empty:
                st.warning("삭제할 인원을 선택해 주세요.")
            else:
                delete_records = rows_to_delete[PERSONNEL_COLUMNS].apply(tuple, axis=1).tolist()
                base_records = personnel_df[PERSONNEL_COLUMNS].apply(tuple, axis=1)
                remaining_mask = ~base_records.isin(delete_records)
                remaining_df = personnel_df[remaining_mask].reset_index(drop=True)
                save_personnel_data(remaining_df)
                st.session_state.pop("personnel_delete_editor", None)
                st.session_state[PERSONNEL_FEEDBACK_KEY] = f"{len(rows_to_delete)}건의 인원 정보를 삭제했습니다."
                st.experimental_rerun()
        else:
            st.error("표 데이터를 불러오지 못했습니다. 다시 시도해 주세요.")

    with st.expander("인원 등록", expanded=True):
        with st.form("personnel_entry_form"):
            col1, col2 = st.columns(2)

            upper_select_options = [PERSONNEL_NEW_OPTION_LABEL] + upper_options
            upper_select = col1.selectbox(
                "상위부서명",
                upper_select_options,
                key="personnel_upper_option"
            )
            upper_input = col1.text_input(
                "상위부서명 입력",
                key="personnel_upper_input",
                disabled=upper_select != PERSONNEL_NEW_OPTION_LABEL,
                placeholder="상위부서를 입력하세요."
            )
            upper_name = upper_input.strip() if upper_select == PERSONNEL_NEW_OPTION_LABEL else upper_select.strip()

            if upper_select != PERSONNEL_NEW_OPTION_LABEL and upper_select:
                dept_candidates = personnel_df[personnel_df["상위부서명"] == upper_select]
            else:
                dept_candidates = personnel_df
            dept_seed = sorted({str(name).strip() for name in dept_candidates["부서명"].tolist() if str(name).strip()})

            dept_select_options = [PERSONNEL_NEW_OPTION_LABEL] + dept_seed
            dept_select = col2.selectbox(
                "부서명",
                dept_select_options,
                key="personnel_dept_option"
            )
            dept_input = col2.text_input(
                "부서명 입력",
                key="personnel_dept_input",
                disabled=dept_select != PERSONNEL_NEW_OPTION_LABEL,
                placeholder="부서명을 입력하세요."
            )
            dept_name = dept_input.strip() if dept_select == PERSONNEL_NEW_OPTION_LABEL else dept_select.strip()

            col3, col4 = st.columns(2)
            emp_id = col3.text_input("사번", key="personnel_emp_id", placeholder="선택 입력")
            name_value = col4.text_input("성명", key="personnel_name_input")

            col5, col6 = st.columns(2)
            title_select_options = [PERSONNEL_NEW_OPTION_LABEL] + title_options
            title_select = col5.selectbox(
                "직위",
                title_select_options,
                key="personnel_title_option"
            )
            title_input = col5.text_input(
                "직위 입력",
                key="personnel_title_input",
                disabled=title_select != PERSONNEL_NEW_OPTION_LABEL,
                placeholder="예: 사원, 대리"
            )
            title_value = title_input.strip() if title_select == PERSONNEL_NEW_OPTION_LABEL else title_select.strip()

            duty_select_options = [PERSONNEL_NEW_OPTION_LABEL] + duty_options
            duty_select = col6.selectbox(
                "직책",
                duty_select_options,
                key="personnel_role_option"
            )
            duty_input = col6.text_input(
                "직책 입력",
                key="personnel_role_input",
                disabled=duty_select != PERSONNEL_NEW_OPTION_LABEL,
                placeholder="예: 팀장, 담당"
            )
            duty_value = duty_input.strip() if duty_select == PERSONNEL_NEW_OPTION_LABEL else duty_select.strip()

            submitted = st.form_submit_button("저장", type="primary")

            if submitted:
                inputs = {
                    "No.": pd.NA,
                    "사번": emp_id.strip(),
                    "상위부서명": upper_name,
                    "부서명": dept_name,
                    "성명": name_value.strip(),
                    "직위": title_value,
                    "직책": duty_value,
                }

                required = {"상위부서명": upper_name, "부서명": dept_name, "성명": name_value.strip()}
                missing_fields = [label for label, value in required.items() if not value]
                if missing_fields:
                    st.warning("상위부서명, 부서명, 성명은 필수입니다.")
                else:
                    if inputs["사번"]:
                        duplicate_mask = personnel_df["사번"] == inputs["사번"]
                    else:
                        duplicate_mask = (
                            (personnel_df["상위부서명"] == inputs["상위부서명"]) &
                            (personnel_df["부서명"] == inputs["부서명"]) &
                            (personnel_df["성명"] == inputs["성명"])
                        )

                    if duplicate_mask.any():
                        st.info("이미 동일한 인원 정보가 등록되어 있습니다.")
                    else:
                        personnel_df = pd.concat([personnel_df, pd.DataFrame([inputs])], ignore_index=True)
                        save_personnel_data(personnel_df)
                        st.session_state["personnel_upper_option"] = PERSONNEL_NEW_OPTION_LABEL
                        st.session_state["personnel_dept_option"] = PERSONNEL_NEW_OPTION_LABEL
                        st.session_state["personnel_title_option"] = PERSONNEL_NEW_OPTION_LABEL
                        st.session_state["personnel_role_option"] = PERSONNEL_NEW_OPTION_LABEL
                        st.session_state["personnel_upper_input"] = ""
                        st.session_state["personnel_dept_input"] = ""
                        st.session_state["personnel_title_input"] = ""
                        st.session_state["personnel_role_input"] = ""
                        st.session_state["personnel_emp_id"] = ""
                        st.session_state["personnel_name_input"] = ""
                        st.session_state.pop("personnel_delete_editor", None)
                        st.session_state[PERSONNEL_FEEDBACK_KEY] = "인원 정보를 저장했습니다."
                        st.experimental_rerun()

def render_masterdata_tab() -> None:
    st.markdown("## 기초 정보 관리")
    st.caption("생산 관련 기초 데이터를 한 곳에서 관리합니다.")

    master_tabs = st.tabs(["인원 관리", "기타 기초정보"])

    with master_tabs[0]:
        render_personnel_section()

    with master_tabs[1]:
        st.info("추가 기초 정보를 이 영역에 확장할 수 있습니다.")

def dataframe_to_html_table(df, font_size=18, highlight_col=None):
    """DataFrame을 HTML 테이블로 변환하는 함수 (글꼴 크기 조절 가능)"""
    
    # 테이블 시작
    html = f"""
    <table style="width: 100%; border-collapse: collapse; font-size: {font_size}px !important;">
    <thead>
    <tr style="background-color: #f8f9fa; border-bottom: 2px solid #dee2e6;">
    """
    
    # 헤더 생성
    for col in df.columns:
        html += f'<th style="padding: 8px; text-align: center; font-weight: bold; font-size: {font_size}px !important; border: 1px solid #dee2e6;">{col}</th>'
    html += "</tr></thead><tbody>"
    
    # 데이터 행 생성
    for idx, row in df.iterrows():
        # 마지막 행(전체 합계)인지 확인
        is_total_row = '전체 합계' in str(row.iloc[0]) if len(row) > 0 else False
        
        if is_total_row:
            html += f'<tr style="background-color: #e9ecef; font-weight: bold; border-top: 2px solid #6c757d;">'
        else:
            html += '<tr>'
        
        for col_idx, (col_name, value) in enumerate(row.items()):
            # 셀 스타일 설정
            cell_style = f"padding: 8px; text-align: center; font-size: {font_size}px !important; border: 1px solid #dee2e6;"
            
            # 달성율 컬럼에 색상 하이라이팅 적용
            if highlight_col and col_name == highlight_col and '%' in str(value):
                try:
                    numeric_val = float(str(value).replace('%', ''))
                    if numeric_val >= 100:
                        cell_style += " background-color: #d4edda; color: #155724;"  # 녹색 (100% 이상)
                    elif numeric_val >= 90:
                        cell_style += " background-color: #fff3cd; color: #856404;"  # 노란색 (90-100%)
                    else:
                        cell_style += " background-color: #f8d7da; color: #721c24;"  # 빨간색 (90% 미만)
                except:
                    pass
            
            if is_total_row:
                cell_style += " font-weight: bold;"
            
            html += f'<td style="{cell_style}">{value}</td>'
        html += '</tr>'
    
    html += "</tbody></table>"
    return html

def daily_dataframe_to_html_table(df, font_size=14):
    """일별 현황 DataFrame을 HTML 테이블로 변환하는 함수"""
    
    html = f"""
    <table style="width: 100%; border-collapse: collapse; font-size: {font_size}px !important;">
    <thead>
    <tr style="background-color: #f8f9fa; border-bottom: 2px solid #dee2e6;">
    """
    
    # 헤더 생성
    for col in df.columns:
        html += f'<th style="padding: 8px; text-align: center; font-weight: bold; font-size: {font_size}px !important; border: 1px solid #dee2e6;">{col}</th>'
    html += "</tr></thead><tbody>"
    
    # 데이터 행 생성 (주말 하이라이팅 포함)
    for idx, row in df.iterrows():
        # 주말인지 확인 (토요일, 일요일)
        try:
            date_str = str(row['생산일자'])  # MM/DD 형식
            # 실제 날짜로 변환하여 요일 확인
            import datetime
            current_year = datetime.datetime.now().year
            month, day = date_str.split('/')
            check_date = datetime.date(current_year, int(month), int(day))
            is_weekend = check_date.weekday() >= 5  # 토요일(5), 일요일(6)
        except:
            is_weekend = False
            
        if is_weekend:
            html += '<tr style="background-color: #f0f0f0; color: #666;">'
        else:
            html += '<tr>'
        
        for col_name, value in row.items():
            cell_style = f"padding: 8px; text-align: center; font-size: {font_size}px !important; border: 1px solid #dee2e6;"
            
            # 주말 배경색 유지
            if is_weekend:
                cell_style += " background-color: #f0f0f0; color: #666;"
            
            html += f'<td style="{cell_style}">{value}</td>'
        html += '</tr>'
    
    html += "</tbody></table>"
    return html

def create_download_section(df, tab_name, agg_level, start_date, end_date):
    """필터링된 데이터를 다운로드할 수 있는 섹션 생성"""
    if df.empty:
        return
    
    st.markdown("---")
    st.subheader("📥 데이터 다운로드", anchor=False)
    
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        # Excel 다운로드
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='필터링된_데이터')
        excel_data = excel_buffer.getvalue()
        
        filename_excel = f"{tab_name}_{agg_level}_{start_date}_{end_date}.xlsx"
        st.download_button(
            label="📊 Excel 다운로드",
            data=excel_data,
            file_name=filename_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col2:
        # CSV 다운로드
        csv_data = df.to_csv(index=False, encoding='utf-8-sig')
        filename_csv = f"{tab_name}_{agg_level}_{start_date}_{end_date}.csv"
        st.download_button(
            label="📄 CSV 다운로드", 
            data=csv_data,
            file_name=filename_csv,
            mime="text/csv",
            use_container_width=True
        )
    
    with col3:
        st.info(f"**데이터 정보**\n- 기간: {start_date} ~ {end_date}\n- 집계: {agg_level}\n- 행 수: {len(df):,}개")

def render_workforce_management_tab() -> None:
    """생산기획팀 인력 운영 대시보드"""
    st.header("인력 운영 센터", anchor=False)
    st.caption("배치·근태·생산성·교육·수급·비용·이슈까지 한 화면에서 관리합니다. 엑셀을 통해 세부 데이터를 지속적으로 유지할 수 있습니다.")

    workforce_data = load_workforce_data()
    deployment_df = workforce_data.get("배치운영", pd.DataFrame())
    attendance_df = workforce_data.get("근태관리", pd.DataFrame())
    productivity_df = workforce_data.get("생산성", pd.DataFrame())
    training_df = workforce_data.get("교육자격", pd.DataFrame())
    supply_df = workforce_data.get("수급계획", pd.DataFrame())
    cost_df = workforce_data.get("비용관리", pd.DataFrame())
    issue_df = workforce_data.get("현장이슈", pd.DataFrame())

    selected_display = st.multiselect(
        "관리 대상 공장",
        FACTORY_DISPLAY_CHOICES,
        default=FACTORY_DISPLAY_CHOICES,
        help="필요한 공장만 선택하면 모든 지표가 해당 공장 기준으로 필터링됩니다."
    )
    if selected_display:
        selected_factories = [FACTORY_DISPLAY_TO_CODE[label] for label in selected_display]
    else:
        selected_factories = list(FACTORY_DEFINITIONS.keys())
    st.caption("선택된 공장: " + ", ".join(FACTORY_DISPLAY_LABELS[code] for code in selected_factories))

    def apply_factory_filter(df: pd.DataFrame) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()
        df_copy = df.copy()
        if df_copy.empty or "공장" not in df_copy.columns:
            return df_copy
        if len(selected_factories) == len(FACTORY_DEFINITIONS):
            return df_copy
        return df_copy[df_copy["공장"].isin(selected_factories)].copy()

    deployment = apply_factory_filter(deployment_df)
    attendance = apply_factory_filter(attendance_df)
    productivity = apply_factory_filter(productivity_df)
    training = apply_factory_filter(training_df)
    supply = apply_factory_filter(supply_df)
    cost = apply_factory_filter(cost_df)
    issue = apply_factory_filter(issue_df)

    control_cols = st.columns([2, 3])
    with control_cols[0]:
        st.download_button(
            "관리용 엑셀 다운로드",
            data=get_workforce_excel_bytes(),
            file_name="workforce_master.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with control_cols[1]:
        uploaded = st.file_uploader(
            "갱신된 엑셀 업로드",
            type=["xlsx"],
            key="workforce_excel_uploader",
            help="시트 구조를 유지한 상태로 데이터를 수정한 뒤 업로드하면 즉시 반영됩니다."
        )
        if uploaded is not None:
            success, message = handle_workforce_upload(uploaded)
            if success:
                st.success(message)
                st.experimental_rerun()
            else:
                st.error(message)

    def _numeric(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce").fillna(0)

    required_sum = _numeric(deployment["필요인원"]).sum() if "필요인원" in deployment.columns else 0
    assigned_sum = _numeric(deployment["배치인원"]).sum() if "배치인원" in deployment.columns else 0
    gap = assigned_sum - required_sum
    tardy = _numeric(attendance["지각"]).sum() if "지각" in attendance.columns else 0
    absence = _numeric(attendance["결근"]).sum() if "결근" in attendance.columns else 0
    avg_eff = _numeric(productivity["효율(%)"]).mean() if "효율(%)" in productivity.columns and not productivity.empty else 0
    total_cost = _numeric(cost["총인건비"]).sum() if "총인건비" in cost.columns else 0
    issue_count = len(issue.index)

    kpi_cols = st.columns(4)
    with kpi_cols[0]:
        st.metric("현재 배치 인원", f"{assigned_sum:,.0f}명", delta=f"{gap:+.0f}명 vs 필요")
    with kpi_cols[1]:
        st.metric("근태 이슈", f"{(tardy + absence):.0f}건", delta=f"지각 {tardy:.0f} · 결근 {absence:.0f}")
    with kpi_cols[2]:
        st.metric("평균 효율", f"{avg_eff:,.1f}%", delta=f"{avg_eff - 95:+.1f}p (목표 95%)")
    with kpi_cols[3]:
        st.metric("누적 인건비(만원)", f"{total_cost:,.0f}", delta=f"현장 이슈 {issue_count}건")

    st.markdown("### 1. 인력 배치 · 운영")
    if deployment.empty:
        st.info("배치운영 데이터가 없습니다. 엑셀 파일을 업데이트해 주세요.")
    else:
        deploy = deployment.copy()
        for col in ["필요인원", "배치인원", "신규투입"]:
            if col in deploy.columns:
                deploy[col] = _numeric(deploy[col])
        deploy["증감"] = deploy.get("배치인원", 0) - deploy.get("필요인원", 0)
        plot_df = deploy.copy()
        x_field = "공정"
        if "공장" in plot_df.columns:
            plot_df["공정(공장)"] = plot_df["공정"].astype(str) + " (" + plot_df["공장"].astype(str) + ")"
            x_field = "공정(공장)"
        id_cols = [x_field, "근무조"]
        if "공장" in plot_df.columns:
            id_cols.append("공장")
        melt_df = plot_df.melt(
            id_vars=id_cols,
            value_vars=[col for col in ["필요인원", "배치인원"] if col in plot_df.columns],
            var_name="구분",
            value_name="인원"
        )
        if not melt_df.empty:
            fig = px.bar(
                melt_df,
                x=x_field,
                y="인원",
                color="구분",
                barmode="group",
                hover_data=["근무조"] + (["공장"] if "공장" in plot_df.columns else []),
                text_auto=True
            )
            fig.update_layout(height=380, legend_title="구분", yaxis_title="인원(명)")
            st.plotly_chart(fig, use_container_width=True)
        if "공장" in deploy.columns:
            factory_summary = deploy.groupby("공장")[['필요인원', '배치인원']].sum().reset_index()
            factory_melt = factory_summary.melt(id_vars=["공장"], value_vars=["필요인원", "배치인원"], var_name="구분", value_name="인원")
            fig_factory = px.bar(factory_melt, x="공장", y="인원", color="구분", barmode="group", text_auto=True, title="공장별 총배치 현황")
            fig_factory.update_layout(height=300, yaxis_title="인원(명)")
            st.plotly_chart(fig_factory, use_container_width=True)
        display_columns = []
        if "공장" in deploy.columns:
            display_columns.append("공장")
        display_columns += [col for col in ["공정", "필요인원", "배치인원", "증감", "근무조", "신규투입", "이동계획"] if col in deploy.columns]
        st.dataframe(deploy[display_columns], use_container_width=True, hide_index=True)

    st.markdown("### 2. 근태 모니터링")
    if attendance.empty:
        st.info("근태 데이터가 없습니다.")
    else:
        attendance["날짜"] = pd.to_datetime(attendance["날짜"], errors="coerce")
        metric_cols = [col for col in ["지각", "결근", "휴가", "특근"] if col in attendance.columns]
        for col in metric_cols:
            attendance[col] = _numeric(attendance[col])
        att_daily = attendance.groupby("날짜")[metric_cols].sum().reset_index()
        if not att_daily.empty:
            melted_att = att_daily.melt(id_vars=["날짜"], value_vars=metric_cols, var_name="구분", value_name="건수")
            fig_att = px.area(melted_att, x="날짜", y="건수", color="구분", title="일자별 근태 추세")
            fig_att.update_layout(height=320)
            st.plotly_chart(fig_att, use_container_width=True)
        if "공장" in attendance.columns and metric_cols:
            att_factory = attendance.groupby("공장")[metric_cols].sum().reset_index()
            st.dataframe(att_factory, use_container_width=True, hide_index=True)
        latest = attendance.sort_values("날짜", ascending=False)
        st.dataframe(latest, use_container_width=True, hide_index=True)

    st.markdown("### 3. 생산성 · 효율")
    if productivity.empty:
        st.info("생산성 데이터가 없습니다.")
    else:
        prod = productivity.copy()
        prod_cols = ["UPH", "UPPH", "평균작업시간(분)", "효율(%)", "잔업시간", "특근생산성"]
        for col in prod_cols:
            if col in prod.columns:
                prod[col] = _numeric(prod[col])
        melt_prod = prod.melt(id_vars=["공정"], value_vars=["UPH", "UPPH"], var_name="지표", value_name="값")
        fig_uph = px.bar(melt_prod, x="공정", y="값", color="지표", text_auto=True, title="공정별 인당 생산량")
        fig_uph.update_layout(height=360, yaxis_title="단위/시간")
        st.plotly_chart(fig_uph, use_container_width=True)
        if "효율(%)" in prod.columns:
            fig_eff = px.line(prod, x="공정", y="효율(%)", markers=True, title="공정별 효율")
            fig_eff.update_yaxes(range=[0, max(110, prod["효율(%)"].max() + 5)])
            st.plotly_chart(fig_eff, use_container_width=True)
        st.dataframe(prod, use_container_width=True, hide_index=True)

    st.markdown("### 4. 교육 · 자격 만료 현황")
    if training.empty:
        st.info("교육/자격 데이터가 없습니다.")
    else:
        training["만료일"] = pd.to_datetime(training["만료일"], errors="coerce")
        training["수료일"] = pd.to_datetime(training["수료일"], errors="coerce")
        upcoming_limit = pd.Timestamp.today() + pd.Timedelta(days=45)
        upcoming = training[training["만료일"] <= upcoming_limit].sort_values("만료일")
        if upcoming.empty:
            st.success("45일 내 만료 예정인 교육이 없습니다.")
        else:
            st.warning("만료 예정 인원 확인이 필요합니다.")
            st.dataframe(upcoming, use_container_width=True, hide_index=True)
        with st.expander("전체 교육 이력 보기", expanded=False):
            st.dataframe(training.sort_values("만료일"), use_container_width=True, hide_index=True)

    st.markdown("### 5. 인력 수급 계획")
    if supply.empty:
        st.info("수급 계획 데이터가 없습니다.")
    else:
        supply["월"] = pd.to_datetime(supply["월"], errors="coerce")
        for col in ["예상수요", "가용인원"]:
            if col in supply.columns:
                supply[col] = _numeric(supply[col])
        supply["과부족"] = supply["가용인원"] - supply["예상수요"]
        monthly_totals = supply.groupby("월")[['예상수요', '가용인원']].sum().reset_index()
        line_fig = px.line(monthly_totals.sort_values("월"), x="월", y=["예상수요", "가용인원"], markers=True, title="월별 인력 수요 vs 가용 인원")
        line_fig.update_layout(height=360, yaxis_title="인원(명)")
        st.plotly_chart(line_fig, use_container_width=True)
        if "공장" in supply.columns:
            shortage_chart = px.bar(supply.sort_values("월"), x="월", y="과부족", color="공장", text_auto=True, title="공장별 과부족 인원 추이")
        else:
            shortage_chart = px.bar(supply.sort_values("월"), x="월", y="과부족", text_auto=True, title="과부족 인원 추이")
        shortage_chart.update_layout(height=260, yaxis_title="가용 - 수요")
        st.plotly_chart(shortage_chart, use_container_width=True)
        st.dataframe(supply, use_container_width=True, hide_index=True)

    st.markdown("### 6. 비용 구조")
    if cost.empty:
        st.info("비용관리 데이터가 없습니다.")
    else:
        for col in ["기본급합계", "잔업비", "특근비", "총인건비"]:
            if col in cost.columns:
                cost[col] = _numeric(cost[col])
        if "총인건비" not in cost.columns or cost["총인건비"].isna().all():
            cost["총인건비"] = cost.get("기본급합계", 0) + cost.get("잔업비", 0) + cost.get("특근비", 0)
        bar_fig = px.bar(cost, x="공장", y="총인건비", color="부서", text_auto=True, barmode="group", title="공장/부서별 인건비")
        bar_fig.update_layout(height=360, yaxis_title="인건비(만원)")
        st.plotly_chart(bar_fig, use_container_width=True)
        st.dataframe(cost, use_container_width=True, hide_index=True)

    st.markdown("### 7. 현장 이슈 모니터링")
    if issue.empty:
        st.info("등록된 현장 이슈가 없습니다.")
    else:
        issue["날짜"] = pd.to_datetime(issue["날짜"], errors="coerce")
        issue = issue.sort_values("날짜", ascending=False)
        st.dataframe(issue, use_container_width=True, hide_index=True)

    st.markdown("---")
    with st.expander("데이터 편집 및 엑셀 반영", expanded=False):
        sheet_options = list(WORKFORCE_SHEET_COLUMNS.keys())
        selected_sheet = st.selectbox("편집할 시트 선택", sheet_options, key="workforce_sheet_select")
        editable_df = workforce_data.get(selected_sheet, pd.DataFrame(columns=WORKFORCE_SHEET_COLUMNS[selected_sheet]))
        st.caption("필요 시 행을 추가하거나 삭제하고, 저장 버튼을 누르면 엑셀 파일로 즉시 반영됩니다.")
        editor_key = f"workforce_editor_{selected_sheet}"
        edited_df = st.data_editor(editable_df, num_rows="dynamic", use_container_width=True, key=editor_key)
        if st.button("변경 사항 저장", type="primary", key=f"workforce_save_{selected_sheet}"):
            workforce_data[selected_sheet] = edited_df
            save_workforce_data(workforce_data)
            st.success("엑셀 파일에 반영했습니다.")
            st.experimental_rerun()


def create_shared_filter_controls(df_for_current_tab):
    """
    모든 탭에서 공유되는 필터 컨트롤을 생성하고 필터링된 데이터프레임을 반환합니다.
    일일 생산 현황 보고 탭은 완전히 독립적으로 처리됩니다.
    """
    all_dfs = [df_target_orig, df_yield_orig, df_utilization_orig, df_defect_orig]
    all_dates = pd.concat([d['date'] for d in all_dfs if d is not None and not d.empty and 'date' in d.columns]).dropna()
    min_date_global, max_date_global = (all_dates.min().date(), all_dates.max().date()) if not all_dates.empty else (date(2000, 1, 1), date.today())

    # 일일 생산 현황 보고 탭은 이 함수를 사용하지 않음 (완전 독립적 처리)
    if selected_tab == "📊 일일 생산 현황 보고":
        # 이 탭은 별도의 고정 설정을 사용하므로 이 함수 호출하지 말아야 함
        # 혹시 호출되더라도 기본값만 반환
        return df_for_current_tab, min_date_global, max_date_global, '월별'

    # date_input 키(`date_range`)에 이전 실행의 범위 밖 값이 남아 있으면 StreamlitAPIException이 발생할 수 있어
    # 위젯 렌더링 전에 현재 데이터 범위로 강제 보정한다.
    def _to_date_safe(value, fallback):
        if isinstance(value, date):
            return value
        try:
            parsed = pd.to_datetime(value, errors='coerce')
            if pd.isna(parsed):
                return fallback
            return parsed.date()
        except Exception:
            return fallback

    raw_range = st.session_state.get('date_range')
    if isinstance(raw_range, (list, tuple)) and len(raw_range) == 2:
        raw_start, raw_end = raw_range
    elif isinstance(raw_range, date):
        raw_start, raw_end = raw_range, raw_range
    else:
        raw_start, raw_end = min_date_global, max_date_global

    safe_start = _to_date_safe(raw_start, min_date_global)
    safe_end = _to_date_safe(raw_end, max_date_global)
    if safe_start > safe_end:
        safe_start, safe_end = safe_end, safe_start

    safe_start = max(min_date_global, min(safe_start, max_date_global))
    safe_end = max(min_date_global, min(safe_end, max_date_global))
    if safe_start > safe_end:
        safe_start, safe_end = min_date_global, max_date_global

    st.session_state.date_range = (safe_start, safe_end)

    header_cols = st.columns([1, 1])
    with header_cols[0]:
        header_title = selected_tab
        if "분석" not in selected_tab: header_title = f"{selected_tab} 분석"
        st.header(header_title, anchor=False)

    filter_cols = st.columns([5.4, 1.1, 1.1, 3.4])
    with filter_cols[0]:
        st.date_input(
            "조회할 기간을 선택하세요",
            min_value=min_date_global,
            max_value=max_date_global,
            key='date_range'
        )
    with filter_cols[1]:
        st.markdown("<div style='padding-top: 28px;'></div>", unsafe_allow_html=True)
        st.button(
            "기간 초기화",
            on_click=reset_filters,
            args=(min_date_global, max_date_global),
            help="현재 조회 기준일이 속한 연도의 모든 데이터를 기준으로 월별 집계합니다."
        )
    with filter_cols[2]:
        st.markdown("<div style='padding-top: 28px;'></div>", unsafe_allow_html=True)
        st.button(
            "최대 기간",
            on_click=set_maximum_period,
            args=(min_date_global, max_date_global),
            help="데이터가 존재하는 전체 기간으로 조회 범위를 확장하고 월별로 집계합니다."
        )
    with filter_cols[3]:
        st.markdown("<div style='padding-top: 28px;'></div>", unsafe_allow_html=True)
        st.radio("집계 기준", options=['일별', '주간별', '월별', '분기별', '반기별', '년도별'], key='agg_level', horizontal=True)

    date_range_value = st.session_state.get('date_range')
    agg_level = st.session_state.get('agg_level', '월별')

    if isinstance(date_range_value, (list, tuple)) and len(date_range_value) == 2:
        start_date, end_date = date_range_value
    else:
        start_date, end_date = min_date_global, max_date_global

    final_start_date = max(start_date, min_date_global)
    final_end_date = min(end_date, max_date_global)
    st.session_state.range_reference_date = final_end_date

    with header_cols[1]:
        st.markdown(f"<p style='text-align: right; margin-top: 1.2rem; font-size: 1.1rem; color: grey;'>({final_start_date.strftime('%Y-%m-%d')} ~ {final_end_date.strftime('%Y-%m-%d')})</p>", unsafe_allow_html=True)
    
    if df_for_current_tab.empty or 'date' not in df_for_current_tab.columns or df_for_current_tab['date'].isnull().all():
        return pd.DataFrame(), final_start_date, final_end_date, agg_level
        
    mask = (df_for_current_tab['date'].dt.date >= final_start_date) & (df_for_current_tab['date'].dt.date <= final_end_date)
    return df_for_current_tab[mask].copy(), final_start_date, final_end_date, agg_level

def aggregate_overall_data(df, analysis_type):
    if df.empty: return pd.DataFrame()
    group_cols = ['공장', '공정코드']
    metrics_map = {'target': {'sums': ['목표_총_생산량', '총_양품수량'], 'rate': '달성률(%)'}, 'yield': {'sums': ['총_생산수량', '총_양품수량'], 'rate': '평균_수율'}, 'utilization': {'sums': ['총_생산수량', '이론상_총_생산량'], 'rate': '평균_가동률'}}
    metrics = metrics_map.get(analysis_type);
    if not metrics: return pd.DataFrame()
    agg_dict = {col: 'sum' for col in metrics['sums'] if col in df.columns};
    if not agg_dict: return pd.DataFrame()
    agg_df = df.groupby(group_cols).agg(agg_dict).reset_index()
    rate_name, sums = metrics['rate'], metrics['sums']
    c1, c2 = sums if analysis_type != 'utilization' else (sums[1], sums[0])
    with pd.option_context('mode.use_inf_as_na', True): agg_df[rate_name] = (100 * agg_df[c2] / agg_df[c1]).fillna(0)
    return agg_df

def plot_horizontal_bar_chart_all_processes(df, analysis_info, all_factories, all_processes):
    rate_col, y_axis_title, chart_title = analysis_info['rate_col'], analysis_info['y_axis_title'], analysis_info['chart_title']
    all_combinations = pd.DataFrame([(f, p) for f in all_factories for p in all_processes], columns=['공장', '공정코드'])
    df_complete = pd.merge(all_combinations, df, on=['공장', '공정코드'], how='left')
    df_complete[rate_col] = df_complete[rate_col].fillna(0)
    st.divider(); st.subheader("공장/공정별 현황 (전체 기간 집계)")
    
    # 그래프 설정 옵션
    with st.expander("📊 차트 설정", expanded=False):
        col_set1, col_set2, col_set3, col_set4 = st.columns(4)
        with col_set1:
            yield_bar_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="yield_bar_label_size")
        with col_set2:
            yield_bar_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="yield_bar_axis_title_size")
        with col_set3:
            yield_bar_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="yield_bar_axis_tick_size")
        with col_set4:
            yield_bar_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="yield_bar_chart_height")
    
    df_complete['공정코드'] = pd.Categorical(df_complete['공정코드'], categories=all_processes, ordered=True)
    df_complete = df_complete.sort_values(by=['공장', '공정코드']); category_orders = {'공정코드': all_processes}
    fig = px.bar(df_complete, x=rate_col, y='공정코드', color='공장', text=rate_col, title=f'<b>{chart_title}</b>', orientation='h', facet_row="공장", height=yield_bar_chart_height, facet_row_spacing=0.05, category_orders=category_orders)
    fig.update_traces(texttemplate='%{text:.2f}%', textposition='auto', textfont_size=yield_bar_label_size); fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1])); fig.update_yaxes(title=y_axis_title, title_font_size=yield_bar_axis_title_size, tickfont_size=yield_bar_axis_tick_size); fig.update_xaxes(title_font_size=yield_bar_axis_title_size, tickfont_size=yield_bar_axis_tick_size); fig.update_layout(title_font_size=yield_bar_axis_title_size)
    st.plotly_chart(fig, use_container_width=True)

# --- 탭별 UI 구현 ---
if selected_tab == "📊 일일 생산 현황 보고":
    st.header("📊 일일 생산 현황 보고", anchor=False)
    st.markdown("*※ 이 보고서는 다른 탭과 완전히 독립적으로 운영됩니다.*")
    
    # === 일일 생산 현황 보고 탭 전용 데이터 사용 (완전 독립) ===
    # 전용 데이터 검증
    if daily_report_target_data.empty or daily_report_yield_data.empty:
        st.warning("일일 생산 현황 보고를 위해서는 '목표달성율'과 '수율' 데이터가 모두 필요합니다.")
    else:
        from datetime import datetime
        import calendar
        
        # 스마트한 데이터 기간 선택 (전용 데이터 사용)
        def get_latest_available_month():
            """가장 최근 데이터가 있는 월을 반환 - 일일 보고서 전용 데이터 사용"""
            # 전용 데이터에서 가장 최근 월 찾기 (완전 독립)
            latest_target_date = daily_report_target_data['date'].max() if not daily_report_target_data.empty else None
            latest_yield_date = daily_report_yield_data['date'].max() if not daily_report_yield_data.empty else None
            
            if latest_target_date and latest_yield_date:
                # 둘 중 더 이른 날짜를 기준으로 (두 데이터가 모두 있는 월)
                latest_common_date = min(latest_target_date, latest_yield_date)
                return latest_common_date.year, latest_common_date.month
            elif latest_target_date or latest_yield_date:
                # 하나라도 있으면 해당 데이터 사용
                latest_date = latest_target_date or latest_yield_date
                return latest_date.year, latest_date.month
            else:
                # 데이터가 없으면 현재 월 반환
                current_date = datetime.now()
                return current_date.year, current_date.month
        
        # 가장 최근 가용 월 선택
        latest_year, latest_month = get_latest_available_month()
        current_date = datetime.now()
        
        # 현재 월과 최신 데이터 월이 다른 경우 안내
        is_current_month = (current_date.year == latest_year and current_date.month == latest_month)
        
        # 1. 핵심 KPI 카드 섹션
        if is_current_month:
            st.subheader("📊 핵심 경영 지표", anchor=False)
        else:
            st.subheader(f"📊 핵심 경영 지표 (기준: {latest_year}년 {latest_month}월)", anchor=False)
            st.info(f"현재월({current_date.year}년 {current_date.month}월) 데이터가 없어 가장 최근 데이터({latest_year}년 {latest_month}월)를 표시합니다.")
        
        # 선택된 월의 데이터 필터링
        month_start = datetime(latest_year, latest_month, 1).date()
        if latest_month == 12:
            next_month_start = datetime(latest_year + 1, 1, 1).date()
        else:
            next_month_start = datetime(latest_year, latest_month + 1, 1).date()
        
        # 목표 데이터 필터링 (전용 데이터 사용)
        mask_target_current = (daily_report_target_data['date'].dt.date >= month_start) & (daily_report_target_data['date'].dt.date < next_month_start)
        df_target_current = daily_report_target_data[mask_target_current].copy()
        
        # 수율 데이터 필터링 (전용 데이터 사용)
        mask_yield_current = (daily_report_yield_data['date'].dt.date >= month_start) & (daily_report_yield_data['date'].dt.date < next_month_start)
        df_yield_current = daily_report_yield_data[mask_yield_current].copy()

        reference_date_for_range = month_start
        if not df_yield_current.empty:
            latest_record = df_yield_current['date'].max()
            if pd.notnull(latest_record):
                reference_date_for_range = latest_record.date()
        elif not df_target_current.empty:
            latest_record = df_target_current['date'].max()
            if pd.notnull(latest_record):
                reference_date_for_range = latest_record.date()
        st.session_state.range_reference_date = reference_date_for_range
        st.session_state.daily_reference_date = reference_date_for_range

        if not df_target_current.empty and not df_yield_current.empty:
            # 목표달성률 탭과 동일한 데이터 처리 방식 적용
            key_cols = ['date', '공장', '공정코드']
            
            # 목표 데이터 집계 (목표달성률 탭과 동일)
            target_agg = df_target_current.groupby(key_cols).agg(목표_총_생산량=('목표_총_생산량', 'sum')).reset_index()
            
            # 수율 데이터 집계 (목표달성률 탭과 동일)  
            yield_agg = df_yield_current.groupby(key_cols).agg(
                총_생산수량=('총_생산수량', 'sum'), 
                총_양품수량=('총_양품수량', 'sum')
            ).reset_index()
            
            # 데이터 병합
            df_merged = pd.merge(target_agg, yield_agg, on=key_cols, how='outer')
            df_merged.fillna({'총_양품수량': 0, '총_생산수량': 0, '목표_총_생산량': 0}, inplace=True)
            
            # 완제품 제조 기준 ([80] 누수/규격검사) - 목표달성률 탭과 동일
            df_kpi_base = df_merged[df_merged['공정코드'] == '[80] 누수/규격검사']
            
            if not df_kpi_base.empty:
                # 공장별 집계
                df_kpi_agg = df_kpi_base.groupby('공장').agg(
                    목표_총_생산량=('목표_총_생산량', 'sum'), 
                    총_양품수량=('총_양품수량', 'sum')
                ).reset_index()
                
                # KPI 계산 (양품 기준 달성률)
                total_target = df_kpi_agg['목표_총_생산량'].sum()
                total_good_production = df_kpi_agg['총_양품수량'].sum()
                achievement_rate = (total_good_production / total_target * 100) if total_target > 0 else 0
                
                # 일평균 생산량 (양품 기준)
                working_days = len(df_yield_current['date'].dt.date.unique()) if not df_yield_current.empty else 1
                daily_avg = total_good_production / working_days if working_days > 0 else 0
                
                # 당일 실적 (최신 날짜 기준, 양품 기준)
                latest_date = df_yield_current['date'].max() if not df_yield_current.empty else current_date
                daily_production_data = df_yield_current[
                    (df_yield_current['date'] == latest_date) & 
                    (df_yield_current['공정코드'] == '[80] 누수/규격검사')
                ]
                daily_production = daily_production_data['총_양품수량'].sum() if not daily_production_data.empty else 0
            
                # 월말까지 목표 갭 (선택된 월 기준)
                days_in_month = calendar.monthrange(latest_year, latest_month)[1]
                if is_current_month:
                    # 현재 월이면 남은 일수 계산
                    remaining_days = max(0, days_in_month - current_date.day)
                else:
                    # 과거 월이면 해당 월의 전체 일수
                    remaining_days = days_in_month - working_days
                
                remaining_target = max(0, total_target - total_good_production)
                daily_needed = remaining_target / max(1, remaining_days) if remaining_days > 0 else 0
            else:
                # KPI 기준 데이터가 없는 경우
                achievement_rate = 0
                daily_avg = 0
                daily_production = 0
                remaining_target = 0
                daily_needed = 0
                total_target = 0
                total_good_production = 0
            
            # KPI 카드 표시
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                label = f"📊 {latest_year}년 {latest_month}월 현재 진도율"
                st.metric(
                    label=label,
                    value=f"{achievement_rate:.1f}%",
                    delta=f"{achievement_rate - 100:.1f}%p" if achievement_rate != 0 else None
                )
            
            with col2:
                st.metric(
                    label="📈 일평균 생산량",
                    value=f"{daily_avg:,.0f}개",
                    help=f"{latest_year}년 {latest_month}월 일평균 생산량"
                )
            
            with col3:
                st.metric(
                    label="⚡ 최신 일일실적",
                    value=f"{daily_production:,.0f}개",
                    help=f"기준일: {latest_date.strftime('%Y-%m-%d') if pd.notnull(latest_date) else '데이터 없음'}"
                )
            
            with col4:
                label = "🎯 목표 부족분"
                if is_current_month:
                    help_text = f"월 목표 대비 부족한 총량"
                else:
                    help_text = f"해당 월 목표 대비 부족량"
                
                st.metric(
                    label=label,
                    value=f"{remaining_target:,.0f}개",
                    help=help_text
                )
        
        else:
            st.info("당월 데이터를 불러올 수 없습니다.")
        
        st.divider()
        
        # 2. 월별 경영실적 요약표
        st.subheader("📅 월별 경영실적 요약", anchor=False)
        
        # 테이블 글꼴 크기 고정
        table_font_size = 18
        
        # 전체 기간 월별 데이터 집계
        df_target_monthly = df_target_orig.copy()
        df_yield_monthly = df_yield_orig.copy()
        
        if not df_target_monthly.empty and not df_yield_monthly.empty:
            # 월별 그룹화
            df_target_monthly['년월'] = df_target_monthly['date'].dt.to_period('M')
            df_yield_monthly['년월'] = df_yield_monthly['date'].dt.to_period('M')
            
            # 목표달성률 탭과 동일한 방식으로 데이터 처리
            key_cols_monthly = ['date', '공장', '공정코드']
            
            # 목표 데이터 집계
            target_agg_monthly = df_target_monthly.groupby(key_cols_monthly).agg(
                목표_총_생산량=('목표_총_생산량', 'sum')
            ).reset_index()
            target_agg_monthly['년월'] = target_agg_monthly['date'].dt.to_period('M')
            
            # 수율 데이터 집계  
            yield_agg_monthly = df_yield_monthly.groupby(key_cols_monthly).agg(
                총_생산수량=('총_생산수량', 'sum'), 
                총_양품수량=('총_양품수량', 'sum')
            ).reset_index()
            yield_agg_monthly['년월'] = yield_agg_monthly['date'].dt.to_period('M')
            
            # 데이터 병합
            df_merged_monthly = pd.merge(target_agg_monthly, yield_agg_monthly, on=key_cols_monthly + ['년월'], how='outer')
            df_merged_monthly.fillna({'총_양품수량': 0, '총_생산수량': 0, '목표_총_생산량': 0}, inplace=True)
            
            # 완제품 제조 기준만 필터링
            df_monthly_kpi = df_merged_monthly[df_merged_monthly['공정코드'] == '[80] 누수/규격검사']
            
            if not df_monthly_kpi.empty:
                # 월별 요약 (양품 기준)
                monthly_summary = df_monthly_kpi.groupby('년월').agg({
                    '목표_총_생산량': 'sum',
                    '총_양품수량': 'sum', 
                    'date': 'nunique'  # 작업일수
                }).reset_index()

                current_year = date.today().year
                monthly_summary = monthly_summary[monthly_summary['년월'].dt.year == current_year]
                
                excluded_days_df = load_excluded_workdays()
                if not excluded_days_df.empty:
                    excluded_days_df = excluded_days_df.copy()
                    excluded_days_df['년월'] = pd.PeriodIndex(year=excluded_days_df['년'], month=excluded_days_df['월'], freq='M')
                    monthly_summary = monthly_summary.merge(
                        excluded_days_df[['년월', '제외근무일수']],
                        on='년월',
                        how='left'
                    )
                else:
                    monthly_summary['제외근무일수'] = 0
                monthly_summary['제외근무일수'] = monthly_summary['제외근무일수'].fillna(0).astype(int)
                calendar_days = monthly_summary['년월'].dt.to_timestamp().dt.daysinmonth
                monthly_summary['계획작업일수'] = np.maximum(calendar_days - monthly_summary['제외근무일수'], 0)
                
                monthly_summary.rename(columns={
                    '목표_총_생산량': '목표수량',
                    '총_양품수량': '총_생산수량',  # 양품수량을 생산실적으로 표시
                    'date': '데이터일수'
                }, inplace=True)
                
                # 생산이 있는 일자만 집계하여 작업일수 재계산
                daily_productivity = (
                    df_monthly_kpi.groupby(['년월', 'date'])
                    .agg({
                        '목표_총_생산량': 'sum',
                        '총_생산수량': 'sum',
                        '총_양품수량': 'sum'
                    })
                    .reset_index()
                )
                daily_productivity[['총_생산수량', '총_양품수량']] = daily_productivity[['총_생산수량', '총_양품수량']].fillna(0)
                daily_productivity['has_production'] = daily_productivity[['총_생산수량', '총_양품수량']].max(axis=1) > 0
                
                productive_days = (
                    daily_productivity.loc[daily_productivity['has_production']]
                    .groupby('년월')['date']
                    .nunique()
                    .rename('작업일수')
                )
                
                monthly_summary = monthly_summary.merge(productive_days, on='년월', how='left')
                monthly_summary['데이터일수'] = monthly_summary['데이터일수'].fillna(0)
                monthly_summary['작업일수'] = monthly_summary['작업일수'].fillna(0)
                monthly_summary['작업일수'] = np.minimum(monthly_summary['작업일수'], monthly_summary['계획작업일수'])
                
                valid_mask = (monthly_summary['작업일수'] > 0) & (monthly_summary['데이터일수'] > 0)
                monthly_summary['목표수량'] = np.where(
                    valid_mask,
                    (monthly_summary['목표수량'] / monthly_summary['데이터일수']) * monthly_summary['작업일수'],
                    0
                )
                
                monthly_summary['데이터일수'] = monthly_summary['데이터일수'].astype(int)
                monthly_summary['작업일수'] = monthly_summary['작업일수'].astype(int)
                
                # 계산 컬럼 추가
                monthly_summary['차이'] = monthly_summary['총_생산수량'] - monthly_summary['목표수량']
                with np.errstate(divide='ignore', invalid='ignore'):
                    monthly_achievement_rate = np.where(
                        monthly_summary['목표수량'] > 0,
                        (monthly_summary['총_생산수량'] / monthly_summary['목표수량']) * 100,
                        0
                    )
                monthly_summary['달성율'] = np.round(monthly_achievement_rate, 1)
                # 표시용 데이터 준비
                display_data = monthly_summary.copy()
                display_data['구분'] = display_data['년월'].astype(str)
                
                # 컬럼 재정렬 및 이름 변경
                display_cols = {
                    '구분': '구분',
                    '목표수량': '생산목표', 
                    '총_생산수량': '생산실적',
                    '차이': '차이',
                    '달성율': '달성율(%)',
                    '작업일수': '작업일수',
                    '제외근무일수': '휴일수'
                }
                
                display_summary = display_data[list(display_cols.keys())].rename(columns=display_cols)
                
                # 수치 포맷팅
                for col in ['생산목표', '생산실적', '차이']:
                    display_summary[col] = display_summary[col].apply(lambda x: f"{x:,.0f}" if pd.notnull(x) else "0")
                
                for col in ['달성율(%)']:
                    display_summary[col] = display_summary[col].apply(lambda x: f"{x:.1f}%" if pd.notnull(x) else "0.0%")
                
                display_summary['작업일수'] = display_summary['작업일수'].apply(lambda x: f"{x:.0f}일" if pd.notnull(x) else "0일")
                display_summary['휴일수'] = display_summary['휴일수'].apply(lambda x: f"{x:.0f}일" if pd.notnull(x) else "0일")
                
                # 합계 행 추가
                if len(display_summary) > 1:
                    total_achievement = (monthly_summary['총_생산수량'].sum() / monthly_summary['목표수량'].sum() * 100) if monthly_summary['목표수량'].sum() > 0 else 0
                    
                    total_row = pd.DataFrame([{
                        '구분': '전체 합계',
                        '생산목표': f"{monthly_summary['목표수량'].sum():,.0f}",
                        '생산실적': f"{monthly_summary['총_생산수량'].sum():,.0f}",
                        '차이': f"{monthly_summary['차이'].sum():,.0f}",
                        '달성율(%)': f"{total_achievement:.1f}%",
                        '작업일수': f"{monthly_summary['작업일수'].sum():.0f}일",
                        '휴일수': f"{monthly_summary['제외근무일수'].sum():.0f}일"
                    }])
                    
                    display_with_total = pd.concat([display_summary, total_row], ignore_index=True)
                    
                    # HTML 테이블로 표시 (글꼴 크기 조절 적용)
                    html_table = dataframe_to_html_table(
                        display_with_total, 
                        font_size=table_font_size, 
                        highlight_col='달성율(%)'
                    )
                    st.markdown(html_table, unsafe_allow_html=True)
                else:
                    # HTML 테이블로 표시 (글꼴 크기 조절 적용)
                    html_table = dataframe_to_html_table(
                        display_summary, 
                        font_size=table_font_size, 
                        highlight_col='달성율(%)'
                    )
                    st.markdown(html_table, unsafe_allow_html=True)
            else:
                st.info("목표 데이터를 찾을 수 없습니다.")
        else:
            st.info("월별 요약을 위한 데이터가 부족합니다.")
        
        st.divider()
        
        # 3. 일별 공장별 현황표
        if is_current_month:
            st.subheader("🏭 당월 일별 공장별 현황", anchor=False)
        else:
            st.subheader(f"🏭 {latest_year}년 {latest_month}월 일별 공장별 현황", anchor=False)
        
        if not df_yield_current.empty and not df_target_current.empty:
            # 목표달성률 탭과 동일한 데이터 처리 (완제품 기준)
            key_cols_daily = ['date', '공장', '공정코드']
            
            # 완제품 기준 ([80] 누수/규격검사) 데이터만 필터링
            df_yield_final = df_yield_current[df_yield_current['공정코드'] == '[80] 누수/규격검사'].copy()
            df_target_final = df_target_current[df_target_current['공정코드'] == '[80] 누수/규격검사'].copy()
            
            if not df_yield_final.empty and not df_target_final.empty:
                # 실제 공장명을 A관, C관, S관으로 매핑 (목표달성률 탭 기준)
                unique_factories = sorted(df_yield_final['공장'].unique())
                factory_mapping = {}
                
                # 실제 공장 순서에 따라 매핑 (1공장->A관(1공장), 2공장->C관(2공장), 3공장->S관(3공장))
                for i, factory in enumerate(unique_factories):
                    if '1공장' in factory or '1' in factory:
                        factory_mapping[factory] = 'A관(1공장)'
                    elif '2공장' in factory or '2' in factory:
                        factory_mapping[factory] = 'C관(2공장)' 
                    elif '3공장' in factory or '3' in factory:
                        factory_mapping[factory] = 'S관(3공장)'
                    else:
                        # 순서대로 할당
                        mapping_order = ['A관(1공장)', 'C관(2공장)', 'S관(3공장)']
                        factory_mapping[factory] = mapping_order[i % 3]
                
                # 공장명 매핑 적용
                df_yield_mapped = df_yield_final.copy()
                df_yield_mapped['공장_매핑'] = df_yield_mapped['공장'].map(factory_mapping)
            
                # 일별 공장별 집계 (양품 기준)
                daily_factory_summary = df_yield_mapped.groupby([
                    df_yield_mapped['date'].dt.date, '공장_매핑'
                ]).agg({
                    '총_양품수량': 'sum'  # 목표달성률 탭과 동일하게 양품 기준
                }).reset_index()
                daily_factory_summary.rename(columns={'date': '생산일자', '총_양품수량': '총_생산수량'}, inplace=True)
            
                # 피벗 테이블로 변환 (일자별로 각 공장의 생산량을 컬럼으로)
                pivot_daily = daily_factory_summary.pivot(
                    index='생산일자', 
                    columns='공장_매핑', 
                    values='총_생산수량'
                ).fillna(0).reset_index()
                
                # 컬럼 정렬 (A관(1공장), C관(2공장), S관(3공장) 순서)
                available_factories = [f for f in ['A관(1공장)', 'C관(2공장)', 'S관(3공장)'] if f in pivot_daily.columns]
                pivot_daily = pivot_daily[['생산일자'] + available_factories]
                
                # 합계 컬럼 추가
                pivot_daily['합계'] = pivot_daily[available_factories].sum(axis=1)
                
                # 목표 대비 달성률 계산 (일별 목표가 있는 경우)
                if not df_target_final.empty:
                    daily_targets = df_target_final.groupby(df_target_final['date'].dt.date)['목표_총_생산량'].sum().reset_index()
                    daily_targets.rename(columns={'date': '생산일자', '목표_총_생산량': '목표수량'}, inplace=True)
                
                    # 목표 데이터와 병합
                    pivot_with_target = pd.merge(pivot_daily, daily_targets, on='생산일자', how='left')
                    pivot_with_target['목표수량'] = pivot_with_target['목표수량'].fillna(0)
                    pivot_with_target['달성율'] = (
                        pivot_with_target['합계'] / pivot_with_target['목표수량'] * 100
                    ).round(1)
                    pivot_with_target['달성율'] = pivot_with_target['달성율'].replace([float('inf'), float('-inf')], 0)
                else:
                    pivot_with_target = pivot_daily.copy()
                    pivot_with_target['달성율'] = 0
                
                # 표시용 데이터 준비
                display_daily = pivot_with_target.copy()
                
                # 날짜 포맷팅
                display_daily['생산일자'] = pd.to_datetime(display_daily['생산일자']).dt.strftime('%m/%d')
                
                # 수치 포맷팅
                for col in available_factories + ['합계']:
                    display_daily[col] = display_daily[col].apply(lambda x: f"{x:,.0f}" if x > 0 else "-")
                
                display_daily['달성율'] = display_daily['달성율'].apply(lambda x: f"{x:.1f}%" if x > 0 else "-")
                
                # 달성율에 따른 색상 및 아이콘 추가
                def add_status_icon(val):
                    try:
                        if '%' in str(val) and val != '-':
                            numeric_val = float(str(val).replace('%', ''))
                            if numeric_val >= 100:
                                return f"✅ {val}"
                            elif numeric_val >= 80:
                                return f"⚠️ {val}"
                            else:
                                return f"❌ {val}"
                    except:
                        pass
                    return val
                
                display_daily['상태'] = display_daily['달성율'].apply(add_status_icon)
                
                # 최종 표시 컬럼 선택
                final_columns = ['생산일자'] + available_factories + ['합계', '상태']
                display_final = display_daily[final_columns]
                
                # 누적 실적 요약 추가 (양품 기준)
                total_production_current = pivot_daily['합계'].sum()
                total_target_current = df_target_final['목표_총_생산량'].sum() if not df_target_final.empty else 0
                overall_achievement = (total_production_current / total_target_current * 100) if total_target_current > 0 else 0
                
                col_summary1, col_summary2, col_summary3 = st.columns(3)
                with col_summary1:
                    month_label = "당월" if is_current_month else f"{latest_year}년 {latest_month}월"
                    st.metric(f"🎯 {month_label} 목표량", f"{total_target_current:,.0f}개")
                with col_summary2:
                    st.metric(f"🏭 {month_label} 총 생산량", f"{total_production_current:,.0f}개")
                with col_summary3:
                    st.metric(f"📊 {month_label} 달성률", f"{overall_achievement:.1f}%")
                
                st.markdown("#### 📅 일별 상세 현황")
                
                # 일별 테이블 글꼴 크기 고정
                daily_font_size = 14
                
                # HTML 테이블로 표시 (일별 테이블 글꼴 크기 조절 적용)
                html_daily_table = daily_dataframe_to_html_table(display_final, font_size=daily_font_size)
                st.markdown(html_daily_table, unsafe_allow_html=True)
                
                # 범례 설명
                st.markdown("""
                **상태 범례:**
                - ✅ 100% 이상 달성
                - ⚠️ 80-100% 달성  
                - ❌ 80% 미만
                - 회색 배경: 주말
                """)
            else:
                month_label = "당월" if is_current_month else f"{latest_year}년 {latest_month}월"
                st.info(f"{month_label} 완제품 생산 데이터([80] 누수/규격검사)가 부족합니다.")
        else:
            month_label = "당월" if is_current_month else f"{latest_year}년 {latest_month}월"
            st.info(f"{month_label} 일별 현황을 표시할 데이터가 부족합니다.")
        
        st.divider()
        
        # 4. 다운로드 및 인쇄 기능
        st.subheader("📄 보고서 다운로드", anchor=False)
        
        col_download1, col_download2 = st.columns(2)
        
        with col_download1:
            if st.button("📊 일일 생산 현황 보고 Excel 다운로드", use_container_width=True):
                try:
                    # Excel 파일 생성
                    import io
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    from datetime import datetime
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "일일생산현황보고"
                    
                    # ============= 전문 보고서 스타일 설정 =============
                    # 색상 정의
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # 진한 파란색
                    sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # 연한 파란색
                    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 녹색
                    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 노란색
                    danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 빨간색
                    
                    # 경계선 스타일
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    # ============= 1. 보고서 헤더 =============
                    ws.merge_cells('A1:H2')
                    ws['A1'] = "일일 생산 현황 보고서"
                    ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
                    ws['A1'].fill = header_fill
                    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 보고서 정보
                    current_date = datetime.now()
                    ws.merge_cells('A3:D3')
                    ws['A3'] = f"보고 기준일: {current_date.strftime('%Y년 %m월 %d일')}"
                    ws['A3'].font = Font(size=12)
                    
                    ws.merge_cells('E3:H3') 
                    ws['E3'] = f"작성일시: {current_date.strftime('%Y-%m-%d %H:%M')}"
                    ws['E3'].font = Font(size=12)
                    ws['E3'].alignment = Alignment(horizontal='right')
                    
                    row = 5
                    
                    # ============= 2. 핵심 성과 지표 (Executive Summary) =============
                    ws.merge_cells(f'A{row}:H{row}')
                    ws[f'A{row}'] = "📊 핵심 성과 지표 (Executive Summary)"
                    ws[f'A{row}'].font = Font(size=16, bold=True)
                    ws[f'A{row}'].fill = sub_header_fill
                    ws[f'A{row}'].alignment = Alignment(horizontal='center')
                    row += 2
                    
                    if not df_target_current.empty and not df_yield_current.empty:
                        # KPI 카드 형태로 구성
                        kpi_data = [
                            ["지표명", "수치", "단위", "평가"],
                            ["당월 목표달성률", f"{achievement_rate:.1f}", "%", "우수" if achievement_rate >= 100 else "양호" if achievement_rate >= 90 else "개선필요"],
                            ["일평균 생산량", f"{daily_avg:,.0f}", "개", ""],
                            ["최근 일일실적", f"{daily_production:,.0f}", "개", ""],
                            ["목표달성 필요량", f"{daily_needed:,.0f}", "개/일", ""]
                        ]
                        
                        for i, kpi_row in enumerate(kpi_data):
                            for j, value in enumerate(kpi_row):
                                cell = ws.cell(row + i, j + 1, value)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                                if i == 0:  # 헤더
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = header_fill
                                elif i == 1:  # 목표달성률
                                    if j == 3:  # 평가 컬럼
                                        if "우수" in value:
                                            cell.fill = good_fill
                                        elif "양호" in value:
                                            cell.fill = warning_fill
                                        else:
                                            cell.fill = danger_fill
                        
                        row += len(kpi_data) + 2
                    
                    # ============= 3. 월별 생산 실적 추이 =============
                    ws.merge_cells(f'A{row}:H{row}')
                    ws[f'A{row}'] = "📈 월별 생산 실적 추이"
                    ws[f'A{row}'].font = Font(size=16, bold=True)
                    ws[f'A{row}'].fill = sub_header_fill
                    ws[f'A{row}'].alignment = Alignment(horizontal='center')
                    row += 2
                    
                    # 월별 요약 데이터가 있는 경우
                    if 'monthly_summary' in locals() and not monthly_summary.empty:
                        # 테이블 헤더
                        headers = ['구분', '생산목표', '생산실적', '차이', '달성율', '작업일수', '휴일수']
                        
                        for i, header in enumerate(headers):
                            cell = ws.cell(row, i + 1, header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = header_fill
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                        
                        row += 1
                        
                        # 데이터 행
                        for _, data_row in monthly_summary.iterrows():
                            cells_data = [
                                str(data_row['년월']),
                                f"{int(data_row['목표수량']):,}",
                                f"{int(data_row['총_생산수량']):,}",
                                f"{int(data_row['차이']):,}",
                                f"{data_row['달성율']:.1f}%",
                                f"{int(data_row['작업일수'])}일",
                                f"{int(data_row['제외근무일수'])}일"
                            ]
                            
                            for i, cell_value in enumerate(cells_data):
                                cell = ws.cell(row, i + 1, cell_value)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center')
                                
                                # 달성률에 따른 색상 적용
                                if i == 4:  # 달성율 컬럼 (0부터 시작하므로 4번째)
                                    try:
                                        rate = float(cell_value.replace('%', ''))
                                        if rate >= 100:
                                            cell.fill = good_fill
                                        elif rate >= 90:
                                            cell.fill = warning_fill
                                        else:
                                            cell.fill = danger_fill
                                    except:
                                        pass
                            row += 1
                        
                        # 전체 합계 행 추가
                        excel_total_achievement = (monthly_summary['총_생산수량'].sum() / monthly_summary['목표수량'].sum() * 100) if monthly_summary['목표수량'].sum() > 0 else 0
                        
                        total_data = [
                            "전체 합계",
                            f"{monthly_summary['목표수량'].sum():,.0f}",
                            f"{monthly_summary['총_생산수량'].sum():,.0f}",
                            f"{monthly_summary['차이'].sum():,.0f}",
                            f"{excel_total_achievement:.1f}%",
                            f"{monthly_summary['작업일수'].sum():.0f}일",
                            f"{monthly_summary['제외근무일수'].sum():.0f}일"
                        ]
                        
                        for i, cell_value in enumerate(total_data):
                            cell = ws.cell(row, i + 1, cell_value)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                        
                        row += 3
                    
                    # ============= 4. 생산 현황 분석 =============
                    ws.merge_cells(f'A{row}:H{row}')
                    ws[f'A{row}'] = "📋 생산 현황 분석 및 개선 방안"
                    ws[f'A{row}'].font = Font(size=16, bold=True)
                    ws[f'A{row}'].fill = sub_header_fill
                    ws[f'A{row}'].alignment = Alignment(horizontal='center')
                    row += 2
                    
                    # 분석 내용
                    analysis_content = [
                        "1. 현황 분석",
                        f"   • 당월 목표달성률: {achievement_rate:.1f}%",
                        f"   • 평가: {'목표를 상회하는 우수한 성과' if achievement_rate >= 100 else '목표 달성을 위한 추가 노력 필요' if achievement_rate >= 90 else '생산량 증대 대책 시급'}",
                        "",
                        "2. 주요 이슈 및 개선 방안",
                        f"   • {'현재 생산 수준 유지 및 품질 관리 강화' if achievement_rate >= 100 else '생산 효율성 개선을 통한 목표 달성' if achievement_rate >= 90 else '생산 프로세스 전반 점검 및 개선 필요'}",
                        f"   • 일일 목표달성을 위한 필요 생산량: {daily_needed:,.0f}개",
                        "",
                        "3. 향후 계획",
                        "   • 지속적인 생산량 모니터링",
                        "   • 품질 관리 체계 강화",
                        "   • 생산 효율성 개선 방안 검토"
                    ]
                    
                    for content in analysis_content:
                        ws[f'A{row}'] = content
                        if content.startswith(("1.", "2.", "3.")):
                            ws[f'A{row}'].font = Font(bold=True, size=12)
                        row += 1
                    
                    # ============= 5. 컬럼 너비 자동 조정 =============
                    for col in range(1, 9):
                        max_length = 0
                        column = get_column_letter(col)
                        for row_cells in ws[f'{column}1:{column}{row}']:
                            for cell in row_cells:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                        adjusted_width = min(max_length + 2, 25)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    # 파일 저장
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    
                    st.download_button(
                        label="📁 보고서 다운로드",
                        data=buffer.getvalue(),
                        file_name=f"일일생산현황보고서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                except Exception as e:
                    st.error(f"Excel 파일 생성 중 오류가 발생했습니다: {str(e)}")
        
        with col_download2:
            if st.button("🖨️ 인쇄용 레이아웃 보기", use_container_width=True):
                st.info("브라우저의 인쇄 기능(Ctrl+P)을 사용하여 현재 화면을 인쇄할 수 있습니다.")

elif selected_tab == "목표 달성률":
    if df_target_orig.empty or df_yield_orig.empty: st.info("해당 분석을 위해서는 '목표달성율'과 '수율' 데이터가 모두 필요합니다.")
    else:
        df_target_filtered, start_date, end_date, agg_level = create_shared_filter_controls(df_target_orig)
        if df_target_filtered.empty: st.info("선택된 기간에 목표 데이터가 없습니다.")
        else:
            mask_yield = (df_yield_orig['date'].dt.date >= start_date) & (df_yield_orig['date'].dt.date <= end_date); df_yield_filtered = df_yield_orig.loc[mask_yield].copy()
            if df_yield_filtered.empty: st.info("선택된 기간에 수율 데이터가 없어, 양품 기반 달성률을 계산할 수 없습니다.")
            else:
                key_cols = ['date', '공장', '공정코드']; target_agg_day = df_target_filtered.groupby(key_cols).agg(목표_총_생산량=('목표_총_생산량', 'sum')).reset_index(); yield_agg_day = df_yield_filtered.groupby(key_cols).agg(총_생산수량=('총_생산수량', 'sum'), 총_양품수량=('총_양품수량', 'sum')).reset_index()
                df_merged = pd.merge(target_agg_day, yield_agg_day, on=key_cols, how='outer'); df_merged.fillna({'총_양품수량': 0, '총_생산수량': 0, '목표_총_생산량': 0}, inplace=True); main_col, side_col = st.columns([2.8, 1])
                with main_col:
                    st.subheader("핵심 지표 요약 (완제품 제조 기준, 양품 기반 달성률)"); df_kpi_base = df_merged[df_merged['공정코드'] == '[80] 누수/규격검사']
                    if not df_kpi_base.empty:
                        df_kpi_agg_factory = df_kpi_base.groupby('공장').agg(목표_총_생산량=('목표_총_생산량', 'sum'), 총_양품수량=('총_양품수량', 'sum')).reset_index()
                        with pd.option_context('mode.use_inf_as_na', True): df_kpi_agg_factory['달성률(%)'] = (100 * df_kpi_agg_factory['총_양품수량'] / df_kpi_agg_factory['목표_총_생산량']).fillna(0)
                        target_kpi, good_kpi = df_kpi_agg_factory['목표_총_생산량'].sum(), df_kpi_agg_factory['총_양품수량'].sum(); rate_kpi = (good_kpi / target_kpi * 100) if target_kpi > 0 else 0
                        kpi1, kpi2, kpi3 = st.columns(3); kpi1.metric("완제품 목표", f"{target_kpi:,.0f} 개"); kpi2.metric("완제품 양품 실적", f"{good_kpi:,.0f} 개"); kpi3.metric("완제품 달성률", f"{rate_kpi:.2f} %")
                        st.divider(); st.markdown("##### 공장별 최종 완제품 달성률 (양품 기준)"); factory_kpi_cols = st.columns(len(df_kpi_agg_factory) or [1])
                        for i, row in df_kpi_agg_factory.iterrows():
                            with factory_kpi_cols[i]: st.metric(label=row['공장'], value=f"{row['달성률(%)']:.2f}%"); st.markdown(f"<p style='font-size:0.8rem;color:grey;margin-top:-8px;'>목표:{row['목표_총_생산량']:,.0f}<br>양품실적:{row['총_양품수량']:,.0f}</p>", unsafe_allow_html=True)
                    st.divider(); st.subheader(f"{agg_level} 완제품 달성률 추이 (양품 기준)")
                    
                    # 차트 설정 옵션
                    with st.expander("📊 차트 설정", expanded=False):
                        chart_setting_cols = st.columns(4)
                        with chart_setting_cols[0]:
                            trend_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="trend_label_size")
                        with chart_setting_cols[1]:
                            trend_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="trend_axis_title_size")
                        with chart_setting_cols[2]:
                            trend_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="trend_axis_tick_size")
                        with chart_setting_cols[3]:
                            trend_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="trend_chart_height")
                    
                    df_resampled = get_resampled_data(df_merged, agg_level, ['목표_총_생산량', '총_양품수량']); df_trend = df_resampled[df_resampled['공정코드'] == '[80] 누수/규격검사'].copy()
                    if not df_trend.empty:
                        with pd.option_context('mode.use_inf_as_na', True): df_trend['달성률(%)'] = (100 * df_trend['총_양품수량'] / df_trend['목표_총_생산량']).fillna(0)
                        
                        # 라벨 겹침 방지 로직
                        df_trend = df_trend.sort_values(['period', '달성률(%)'], ascending=[True, False])
                        positions = ['top center', 'bottom center', 'middle right', 'middle left', 'top right', 'bottom right']
                        df_trend['text_position'] = df_trend.groupby('period').cumcount().apply(lambda i: positions[i % len(positions)])
                        
                        fig_trend = go.Figure()

                        for factory_name in sorted(df_trend['공장'].unique()):
                            df_factory = df_trend[df_trend['공장'] == factory_name].sort_values('period')
                            factory_color = next((color for key, color in FACTORY_COLOR_MAP.items() if key in factory_name), '#888888')

                            fig_trend.add_trace(go.Scatter(
                                x=df_factory['period'], y=df_factory['달성률(%)'], name=f'{factory_name} 달성률',
                                mode='lines+markers+text', text=df_factory['달성률(%)'], texttemplate='%{text:.2f}%',
                                textposition=df_factory['text_position'], 
                                line=dict(color=factory_color), legendgroup=factory_name,
                                textfont=dict(size=trend_label_size, color='black')
                            ))

                        fig_trend.update_layout(height=trend_chart_height, title_text=f'<b>{agg_level} 완제품 달성률 추이 (양품 기준)</b>', margin=dict(t=120), legend=dict(orientation="h", yanchor="bottom", y=1.10, xanchor="right", x=1))
                        fig_trend.update_yaxes(title_text="<b>달성률 (%)</b>", autorange=True, title_font_size=trend_axis_title_size, tickfont_size=trend_axis_tick_size) # Y축 범위 자동 조정
                        fig_trend.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(df_trend['period'].unique()), title_text=f"<b>{agg_level.replace('별','')}</b>", title_font_size=trend_axis_title_size, tickfont_size=trend_axis_tick_size)
                        
                        # 자동 라벨 겹침 방지 기능 활성화
                        fig_trend.update_traces(textfont_size=trend_label_size, textposition='top center')
                        fig_trend.update_layout(uniformtext_minsize=max(8, trend_label_size-4), uniformtext_mode='hide')

                        st.plotly_chart(fig_trend, use_container_width=True)
                    
                    # --- 신규 차트: 공장/공정별 상세 분석 ---
                    st.divider()
                    st.subheader("공장/공정별 상세 달성률 분석")

                    all_factories_detail = ['전체'] + sorted(df_merged['공장'].unique())

                    if len(all_factories_detail) > 1:
                        filter_cols_detail = st.columns(2)
                        with filter_cols_detail[0]:
                            selected_factory_detail = st.selectbox("공장 선택", options=all_factories_detail, key="detail_target_factory_select")

                        # 선택된 공장에 따라 동적으로 공정 목록 필터링
                        df_for_processes = df_merged if selected_factory_detail == '전체' else df_merged[df_merged['공장'] == selected_factory_detail]
                        
                        # 달성률 계산이 가능한(목표가 있는) 공정만 필터링
                        df_for_processes_with_target = df_for_processes[df_for_processes['목표_총_생산량'] > 0]
                        all_processes_detail_list = get_process_order(df_for_processes_with_target)
                        all_processes_detail = ['전체'] + all_processes_detail_list
                        
                        with filter_cols_detail[1]:
                            selected_process_detail = st.selectbox("공정 선택", options=all_processes_detail, key="detail_target_process_select")

                        df_detail_filtered = df_merged.copy()
                        
                        if selected_factory_detail == '전체' and selected_process_detail == '전체':
                            st.info("상세 분석을 위해 '공장' 또는 '공정'을 하나 이상 선택해주세요. (현재 위 차트와 동일한 완제품 기준 데이터가 표시됩니다.)")
                            df_detail_filtered = df_detail_filtered[df_detail_filtered['공정코드'] == '[80] 누수/규격검사']
                        else:
                            if selected_factory_detail != '전체':
                                df_detail_filtered = df_detail_filtered[df_detail_filtered['공장'] == selected_factory_detail]
                            
                            if selected_process_detail != '전체':
                                df_detail_filtered = df_detail_filtered[df_detail_filtered['공정코드'] == selected_process_detail]

                        group_by_cols = ['period']
                        color_col = None
                        barmode = 'relative'
                        title_factory = selected_factory_detail
                        title_process = selected_process_detail

                        if selected_factory_detail == '전체' and selected_process_detail != '전체':
                            group_by_cols.append('공장')
                            color_col = '공장'
                            barmode = 'group'
                        elif selected_factory_detail != '전체' and selected_process_detail == '전체':
                            group_by_cols.append('공정코드')
                            color_col = '공정코드'
                            barmode = 'group'
                        elif selected_factory_detail == '전체' and selected_process_detail == '전체':
                            group_by_cols.append('공장')
                            color_col = '공장'
                            barmode = 'group'
                            title_process = '완제품'


                        df_detail_resampled = get_resampled_data(df_detail_filtered, agg_level, ['목표_총_생산량', '총_양품수량'], group_by_cols=group_by_cols)

                        if not df_detail_resampled.empty:
                            with pd.option_context('mode.use_inf_as_na', True): 
                                df_detail_resampled['달성률(%)'] = (100 * df_detail_resampled['총_양품수량'] / df_detail_resampled['목표_총_생산량']).fillna(0)
                            
                            df_detail_resampled = df_detail_resampled[df_detail_resampled['목표_총_생산량'] > 0].copy()

                            if not df_detail_resampled.empty:
                                # 차트 설정 및 축 범위 조절
                                setting_cols = st.columns(2)
                                with setting_cols[0]:
                                    with st.expander("📊 차트 설정", expanded=False):
                                        detail_setting_cols = st.columns(4)
                                        with detail_setting_cols[0]:
                                            detail_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="detail_label_size")
                                        with detail_setting_cols[1]:
                                            detail_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="detail_axis_title_size")
                                        with detail_setting_cols[2]:
                                            detail_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="detail_axis_tick_size")
                                        with detail_setting_cols[3]:
                                            detail_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="detail_chart_height")
                                
                                with setting_cols[1]:
                                    with st.expander("달성률 축 범위 조절", expanded=False):
                                        rate_values = pd.to_numeric(df_detail_resampled['달성률(%)'], errors='coerce')
                                        rate_values = rate_values[np.isfinite(rate_values)]
                                        if rate_values.empty:
                                            min_rate_val, max_rate_val = 0.0, 100.0
                                        else:
                                            min_rate_val = float(rate_values.min())
                                            max_rate_val = float(rate_values.max())

                                        buffer = (max_rate_val - min_rate_val) * 0.1 if max_rate_val > min_rate_val else 5.0
                                        slider_min = max(0.0, min_rate_val - buffer)
                                        slider_max = max_rate_val + buffer
                                        if not np.isfinite(slider_min):
                                            slider_min = 0.0
                                        if not np.isfinite(slider_max):
                                            slider_max = 150.0
                                        rate_slider_max = float(max(150.0, round(slider_max, -1)))
                                        if (not np.isfinite(rate_slider_max)) or rate_slider_max <= 0:
                                            rate_slider_max = 150.0
                                        slider_min = float(max(0.0, min(slider_min, rate_slider_max)))
                                        slider_max = float(max(0.0, min(slider_max, rate_slider_max)))
                                        if slider_min > slider_max:
                                            slider_min, slider_max = 0.0, min(100.0, rate_slider_max)

                                        rate_range = st.slider(
                                            "달성률(%) Y축 범위 선택",
                                            min_value=0.0,
                                            max_value=rate_slider_max,
                                            value=(float(slider_min), float(slider_max)),
                                            step=1.0,
                                            format="%.0f%%",
                                            key="detail_rate_range_slider"
                                        )

                                # 그래프 표시 여부 토글
                                toggle_cols = st.columns(2)
                                with toggle_cols[0]:
                                    show_bar_chart = st.toggle("막대그래프 표시(양품 실적)", value=True, key="show_bar_chart_detail")
                                with toggle_cols[1]:
                                    show_line_chart = st.toggle("꺾은선그래프 표시(달성률)", value=True, key="show_line_chart_detail")

                                # 라벨 겹침 방지 로직
                                if color_col in df_detail_resampled.columns:
                                    df_detail_resampled = df_detail_resampled.sort_values(['period', '달성률(%)'], ascending=[True, False])
                                    positions = ['top center', 'bottom center', 'middle right', 'middle left', 'top right', 'bottom right']
                                    df_detail_resampled['text_position'] = df_detail_resampled.groupby('period').cumcount().apply(lambda i: positions[i % len(positions)])
                                else:
                                    df_detail_resampled['text_position'] = 'top center'

                                if not show_bar_chart and not show_line_chart:
                                    st.warning("차트를 보려면 '막대그래프' 또는 '꺾은선그래프' 중 하나 이상을 선택해주세요.")
                                else:
                                    fig_detail = make_subplots(specs=[[{"secondary_y": True}]])
                                    process_color_map = {p: px.colors.qualitative.Plotly[i % len(px.colors.qualitative.Plotly)] for i, p in enumerate(all_processes_detail_list)}

                                    if color_col:
                                        unique_items = sorted(df_detail_resampled[color_col].unique())
                                        for item_name in unique_items:
                                            df_item = df_detail_resampled[df_detail_resampled[color_col] == item_name].sort_values('period')
                                            
                                            item_color = process_color_map.get(item_name, '#888888')
                                            if color_col == '공장':
                                                item_color = next((color for key, color in FACTORY_COLOR_MAP.items() if key in item_name), '#888888')
                                            
                                            if show_bar_chart:
                                                fig_detail.add_trace(go.Bar(
                                                    x=df_item['period'], y=df_item['총_양품수량'], name=f'{item_name} 양품 실적',
                                                    marker_color=item_color, legendgroup=item_name, text=df_item['총_양품수량'],
                                                    texttemplate='%{text:,.0f}', textposition='outside',
                                                    textfont=dict(size=detail_label_size, color='black')
                                                ), secondary_y=False)
                                            
                                            if show_line_chart:
                                                fig_detail.add_trace(go.Scatter(
                                                    x=df_item['period'], y=df_item['달성률(%)'], name=f'{item_name} 달성률',
                                                    mode='lines+markers+text', text=df_item['달성률(%)'], texttemplate='%{text:.2f}%',
                                                    textposition=df_item['text_position'], line=dict(color=item_color),
                                                    legendgroup=item_name, textfont=dict(size=detail_label_size, color='black')
                                                ), secondary_y=True)
                                    else:
                                        if show_bar_chart:
                                            fig_detail.add_trace(go.Bar(
                                                x=df_detail_resampled['period'], y=df_detail_resampled['총_양품수량'], name='양품 실적',
                                                text=df_detail_resampled['총_양품수량'], texttemplate='%{text:,.0f}',
                                                textposition='outside', marker_color='#85c1e9', textfont=dict(size=detail_label_size, color='black')
                                            ), secondary_y=False)
                                        if show_line_chart:
                                            fig_detail.add_trace(go.Scatter(
                                                x=df_detail_resampled['period'], y=df_detail_resampled['달성률(%)'], name='달성률',
                                                mode='lines+markers+text', text=df_detail_resampled['달성률(%)'], texttemplate='%{text:.2f}%',
                                                textposition=df_detail_resampled['text_position'], line=dict(color='#2874a6'),
                                                textfont=dict(size=detail_label_size, color='black')
                                            ), secondary_y=True)

                                    # 차트 레이아웃 설정 및 출력
                                    max_bar_val_detail = df_detail_resampled['총_양품수량'].max() * 1.2 if not df_detail_resampled.empty else 0
                                    fig_detail.update_layout(barmode=barmode, height=detail_chart_height, title_text=f"<b>{agg_level} {title_factory} {title_process} 실적 및 달성률</b>", margin=dict(t=120), legend=dict(orientation="h", yanchor="bottom", y=1.10, xanchor="right", x=1))
                                    fig_detail.update_yaxes(title_text="<b>양품 실적 (개)</b>", secondary_y=False, range=[0, max_bar_val_detail], visible=show_bar_chart, title_font_size=detail_axis_title_size, tickfont_size=detail_axis_tick_size)
                                    fig_detail.update_yaxes(title_text="<b>달성률 (%)</b>", secondary_y=True, range=rate_range, visible=show_line_chart, title_font_size=detail_axis_title_size, tickfont_size=detail_axis_tick_size)
                                    fig_detail.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(df_detail_resampled['period'].unique()), title_text=f"<b>{agg_level.replace('별','')}</b>", title_font_size=detail_axis_title_size, tickfont_size=detail_axis_tick_size)
                                    st.plotly_chart(fig_detail, use_container_width=True)
                            else:
                                st.info("선택된 조건에 대한 데이터가 없습니다.")
                        else:
                            st.warning("상세 분석을 위한 공장 또는 공정 데이터가 부족합니다.")
                    
                    df_total_agg = df_merged.groupby(['공장', '공정코드']).agg(목표_총_생산량=('목표_총_생산량', 'sum'), 총_양품수량=('총_양품수량', 'sum')).reset_index()
                    with pd.option_context('mode.use_inf_as_na', True): df_total_agg['달성률(%)'] = (100 * df_total_agg['총_양품수량'] / df_total_agg['목표_총_생산량']).fillna(0)
                    df_total_agg = df_total_agg[df_total_agg['목표_총_생산량'] > 0]; st.divider(); st.subheader("공장/공정별 현황 (전체 기간 집계)")
                    chart_process_order = get_process_order(df_total_agg)
                    # 그래프 설정 옵션
                    with st.expander("📊 차트 설정", expanded=False):
                        col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                        with col_set1:
                            target_bar_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="target_bar_label_size")
                        with col_set2:
                            target_bar_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="target_bar_axis_title_size")
                        with col_set3:
                            target_bar_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="target_bar_axis_tick_size")
                        with col_set4:
                            target_bar_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="target_bar_chart_height")

                    df_total_agg['공정코드'] = pd.Categorical(df_total_agg['공정코드'], categories=chart_process_order, ordered=True); df_total_agg = df_total_agg.sort_values(by=['공장', '공정코드']); category_orders = {'공정코드': chart_process_order}
                    fig_bar = px.bar(df_total_agg, x='달성률(%)', y='공정코드', color='공장', text='달성률(%)', title='<b>공장/공정별 달성률 현황 (양품 기준)</b>', orientation='h', facet_row="공장", height=target_bar_chart_height, facet_row_spacing=0.05, category_orders=category_orders)
                    fig_bar.update_traces(texttemplate='%{text:.2f}%', textposition='auto', textfont_size=target_bar_label_size); fig_bar.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1])); fig_bar.update_yaxes(title="공정", title_font_size=target_bar_axis_title_size, tickfont_size=target_bar_axis_tick_size); fig_bar.update_xaxes(title_font_size=target_bar_axis_title_size, tickfont_size=target_bar_axis_tick_size); fig_bar.update_layout(title_font_size=target_bar_axis_title_size); st.plotly_chart(fig_bar, use_container_width=True)

                # 심층 분석 섹션 추가
                st.divider()
                st.subheader("🔍 심층 분석")
                
                analysis_tabs = st.tabs(["📊 상세 통계", "📈 추세 분석", "⚠️ 이상치 분석", "🎯 성과 매트릭스"])
                
                with analysis_tabs[0]:  # 상세 통계
                    st.markdown("##### 목표달성률 상세 통계")
                    
                    # 달성률 계산
                    with pd.option_context('mode.use_inf_as_na', True):
                        df_merged['달성률(%)'] = (100 * df_merged['총_양품수량'] / df_merged['목표_총_생산량']).fillna(0)
                    
                    stats_col1, stats_col2 = st.columns(2)
                    
                    with stats_col1:
                        st.markdown("**전체 기간 통계**")
                        avg_achievement = df_merged['달성률(%)'].mean()
                        median_achievement = df_merged['달성률(%)'].median()
                        std_achievement = df_merged['달성률(%)'].std()
                        min_achievement = df_merged['달성률(%)'].min()
                        max_achievement = df_merged['달성률(%)'].max()
                        
                        st.metric("평균 달성률", f"{avg_achievement:.2f}%")
                        st.metric("표준편차", f"{std_achievement:.2f}%")
                        st.metric("최고 달성률", f"{max_achievement:.2f}%")
                        st.metric("최저 달성률", f"{min_achievement:.2f}%")
                    
                    with stats_col2:
                        st.markdown("**공장별 성과**")
                        factory_stats = df_merged.groupby('공장')['달성률(%)'].agg(['mean', 'count', 'sum']).round(2)
                        factory_stats.columns = ['평균달성률(%)', '측정일수', '총달성률합계']
                        st.dataframe(factory_stats, use_container_width=True)
                        
                        # 공정별 성과
                        st.markdown("**공정별 성과**")
                        process_stats = df_merged.groupby('공정코드')['달성률(%)'].agg(['mean', 'count']).round(2)
                        process_stats.columns = ['평균달성률(%)', '측정일수']
                        st.dataframe(process_stats, use_container_width=True)
                
                with analysis_tabs[1]:  # 추세 분석
                    st.markdown("##### 목표달성률 추세 분석")
                    
                    trend_data = get_resampled_data(df_merged, agg_level, ['목표_총_생산량', '총_양품수량'], 
                                                   group_by_cols=['period', '공장', '공정코드'])
                    
                    if not trend_data.empty:
                        with pd.option_context('mode.use_inf_as_na', True):
                            trend_data['달성률(%)'] = (100 * trend_data['총_양품수량'] / trend_data['목표_총_생산량']).fillna(0)
                        
                        # 전체 추세
                        overall_trend = trend_data.groupby('period').agg({
                            '목표_총_생산량': 'sum',
                            '총_양품수량': 'sum'
                        }).reset_index()
                        
                        with pd.option_context('mode.use_inf_as_na', True):
                            overall_trend['전체달성률(%)'] = (100 * overall_trend['총_양품수량'] / overall_trend['목표_총_생산량']).fillna(0)
                        
                        # 그래프 설정 옵션
                        with st.expander("📊 차트 설정", expanded=False):
                            col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                            with col_set1:
                                trend_analysis_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="trend_analysis_label_size")
                            with col_set2:
                                trend_analysis_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="trend_analysis_axis_title_size")
                            with col_set3:
                                trend_analysis_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="trend_analysis_axis_tick_size")
                            with col_set4:
                                trend_analysis_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="trend_analysis_chart_height")

                        fig_trend = px.line(overall_trend.sort_values('period'), 
                                          x='period', y='전체달성률(%)', 
                                          title='<b>전체 목표달성률 추세</b>',
                                          markers=True, text='전체달성률(%)', height=trend_analysis_chart_height)
                        fig_trend.update_traces(texttemplate='%{text:.1f}%', textposition='top center', textfont=dict(size=trend_analysis_label_size, color='black'))
                        fig_trend.update_xaxes(type='category', title_font_size=trend_analysis_axis_title_size, tickfont_size=trend_analysis_axis_tick_size)
                        fig_trend.update_yaxes(title_font_size=trend_analysis_axis_title_size, tickfont_size=trend_analysis_axis_tick_size)
                        fig_trend.update_layout(title_font_size=trend_analysis_axis_title_size)
                        st.plotly_chart(fig_trend, use_container_width=True)
                        
                        # 공장별 추세 비교
                        factory_trend = trend_data.groupby(['period', '공장']).agg({
                            '목표_총_생산량': 'sum',
                            '총_양품수량': 'sum'
                        }).reset_index()
                        
                        with pd.option_context('mode.use_inf_as_na', True):
                            factory_trend['달성률(%)'] = (100 * factory_trend['총_양품수량'] / factory_trend['목표_총_생산량']).fillna(0)
                        
                        fig_factory_trend = px.line(factory_trend.sort_values('period'), 
                                                  x='period', y='달성률(%)', color='공장',
                                                  title='<b>공장별 목표달성률 추세 비교</b>',
                                                  markers=True, text='달성률(%)', height=trend_analysis_chart_height)
                        fig_factory_trend.update_traces(texttemplate='%{text:.1f}%', textposition='top center', textfont=dict(size=trend_analysis_label_size, color='black'))
                        fig_factory_trend.update_xaxes(type='category', title_font_size=trend_analysis_axis_title_size, tickfont_size=trend_analysis_axis_tick_size)
                        fig_factory_trend.update_yaxes(title_font_size=trend_analysis_axis_title_size, tickfont_size=trend_analysis_axis_tick_size)
                        fig_factory_trend.update_layout(title_font_size=trend_analysis_axis_title_size)
                        st.plotly_chart(fig_factory_trend, use_container_width=True)
                
                with analysis_tabs[2]:  # 이상치 분석
                    st.markdown("##### 이상치 및 변동성 분석")
                    
                    # 그래프 설정 옵션
                    with st.expander("📊 차트 설정", expanded=False):
                        col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                        with col_set1:
                            outlier_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="outlier_label_size")
                        with col_set2:
                            outlier_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="outlier_axis_title_size")
                        with col_set3:
                            outlier_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="outlier_axis_tick_size")
                        with col_set4:
                            outlier_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="outlier_chart_height")

                    # 달성률 분포
                    fig_hist = px.histogram(df_merged, x='달성률(%)', nbins=30,
                                          title='<b>목표달성률 분포</b>', height=outlier_chart_height)
                    fig_hist.update_layout(showlegend=False, xaxis_title='달성률 (%)', yaxis_title='빈도', title_font_size=outlier_axis_title_size)
                    fig_hist.update_traces(texttemplate='%{y}', textposition='outside', textfont=dict(size=outlier_label_size, color='black'))
                    fig_hist.update_xaxes(title_font_size=outlier_axis_title_size, tickfont_size=outlier_axis_tick_size)
                    fig_hist.update_yaxes(title_font_size=outlier_axis_title_size, tickfont_size=outlier_axis_tick_size)
                    st.plotly_chart(fig_hist, use_container_width=True)
                    
                    # 이상치 식별
                    Q1 = df_merged['달성률(%)'].quantile(0.25)
                    Q3 = df_merged['달성률(%)'].quantile(0.75)
                    IQR = Q3 - Q1
                    lower_bound = Q1 - 1.5 * IQR
                    upper_bound = Q3 + 1.5 * IQR
                    
                    outliers = df_merged[(df_merged['달성률(%)'] < lower_bound) | 
                                       (df_merged['달성률(%)'] > upper_bound)]
                    
                    if not outliers.empty:
                        st.markdown("**이상치 데이터**")
                        outliers_display = outliers[['date', '공장', '공정코드', '달성률(%)', '목표_총_생산량', '총_양품수량']].copy()
                        outliers_display['date'] = outliers_display['date'].dt.strftime('%Y-%m-%d')
                        st.dataframe(outliers_display.sort_values('달성률(%)', ascending=False), use_container_width=True)
                    else:
                        st.info("통계적 이상치가 발견되지 않았습니다.")
                    
                    # 변동성 분석
                    volatility_analysis = df_merged.groupby(['공장', '공정코드'])['달성률(%)'].agg(['std', 'mean']).reset_index()
                    volatility_analysis['변동계수'] = volatility_analysis['std'] / volatility_analysis['mean']
                    volatility_analysis = volatility_analysis.sort_values('변동계수', ascending=False)
                    
                    st.markdown("**변동성이 높은 공정 (변동계수 기준)**")
                    st.dataframe(volatility_analysis.round(3), use_container_width=True)
                
                with analysis_tabs[3]:  # 성과 매트릭스
                    st.markdown("##### 목표 vs 실적 성과 매트릭스")
                    
                    # 산점도 분석
                    matrix_data = df_merged.groupby(['공장', '공정코드']).agg({
                        '목표_총_생산량': 'sum',
                        '총_양품수량': 'sum'
                    }).reset_index()
                    
                    with pd.option_context('mode.use_inf_as_na', True):
                        matrix_data['달성률(%)'] = (100 * matrix_data['총_양품수량'] / matrix_data['목표_총_생산량']).fillna(0)
                    
                    # 그래프 설정 옵션
                    with st.expander("📊 차트 설정", expanded=False):
                        col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                        with col_set1:
                            matrix_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="matrix_label_size")
                        with col_set2:
                            matrix_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="matrix_axis_title_size")
                        with col_set3:
                            matrix_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="matrix_axis_tick_size")
                        with col_set4:
                            matrix_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="matrix_chart_height")

                    fig_scatter = px.scatter(matrix_data, 
                                           x='목표_총_생산량', y='총_양품수량',
                                           color='공장', symbol='공정코드',
                                           size='달성률(%)',
                                           title='<b>목표 vs 실적 성과 매트릭스</b>',
                                           hover_data=['달성률(%)'],
                                           text='달성률(%)', height=matrix_chart_height)
                    fig_scatter.update_traces(texttemplate='%{text:.1f}%', textposition='middle center', textfont=dict(size=matrix_label_size, color='white'))
                    fig_scatter.update_xaxes(title_font_size=matrix_axis_title_size, tickfont_size=matrix_axis_tick_size)
                    fig_scatter.update_yaxes(title_font_size=matrix_axis_title_size, tickfont_size=matrix_axis_tick_size)
                    fig_scatter.update_layout(title_font_size=matrix_axis_title_size)
                    
                    # 45도 기준선 추가 (목표=실적)
                    max_val = max(matrix_data['목표_총_생산량'].max(), matrix_data['총_양품수량'].max())
                    fig_scatter.add_shape(type="line", x0=0, y0=0, x1=max_val, y1=max_val,
                                        line=dict(color="red", width=2, dash="dash"))
                    
                    st.plotly_chart(fig_scatter, use_container_width=True)
                    
                    # 성과 매트릭스 테이블
                    st.markdown("**성과 매트릭스 상세**")
                    perf_matrix = matrix_data.copy()
                    perf_matrix['목표대비차이'] = perf_matrix['총_양품수량'] - perf_matrix['목표_총_생산량']
                    perf_matrix['성과등급'] = pd.cut(perf_matrix['달성률(%)'], 
                                                bins=[0, 80, 90, 100, 110, float('inf')],
                                                labels=['미달', '보통', '양호', '우수', '초과달성'])
                    
                    grade_summary = perf_matrix['성과등급'].value_counts()
                    st.dataframe(grade_summary, use_container_width=True)

                with side_col:
                    st.markdown(analyze_target_data(df_merged)); st.divider(); st.subheader("데이터 원본 (일별 집계)"); df_display = df_merged.copy();
                    with pd.option_context('mode.use_inf_as_na', True): df_display['달성률(%)'] = (100 * df_display['총_양품수량'] / df_display['목표_총_생산량']).fillna(0)
                    df_display = df_display.rename(columns={'date': '일자', '목표_총_생산량': '목표 생산량', '총_양품수량': '총 양품수량'}); st.dataframe(df_display[['일자', '공장', '공정코드', '목표 생산량', '총 양품수량', '달성률(%)']].sort_values(by=['일자', '공장', '공정코드']), use_container_width=True, height=500)
        
        # 다운로드 섹션 추가
        create_download_section(df_target_filtered, "목표달성률", agg_level, start_date, end_date)

elif selected_tab == "수율 분석":
    df_filtered, start_date, end_date, agg_level = create_shared_filter_controls(df_yield_orig)
    if not df_filtered.empty:
        main_col, side_col = st.columns([2.8, 1])
        with main_col:
            # --- 공장별 종합 수율 추이 ---
            df_resampled_factory = get_resampled_data(df_filtered, agg_level, ['총_생산수량', '총_양품수량'], group_by_cols=['period', '공장', '공정코드'])
            if not df_resampled_factory.empty:
                st.subheader(f"{agg_level} 공장별 종합 수율 추이")
                
                # 그래프 설정 옵션
                with st.expander("📊 차트 설정", expanded=False):
                    col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                    with col_set1:
                        yield_factory_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="yield_factory_label_size")
                    with col_set2:
                        yield_factory_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="yield_factory_axis_title_size")
                    with col_set3:
                        yield_factory_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="yield_factory_axis_tick_size")
                    with col_set4:
                        yield_factory_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="yield_factory_chart_height")
                
                with pd.option_context('mode.use_inf_as_na', True): df_resampled_factory['개별수율'] = (df_resampled_factory['총_양품수량'] / df_resampled_factory['총_생산수량']).fillna(1.0)
                factory_yield_trend = df_resampled_factory.groupby(['period', '공장'])['개별수율'].prod().reset_index()
                factory_yield_trend['종합수율(%)'] = factory_yield_trend.pop('개별수율') * 100
                fig_factory_trend = px.line(factory_yield_trend.sort_values('period'), x='period', y='종합수율(%)', color='공장', title=f'<b>{agg_level} 공장별 종합 수율 추이</b>', markers=True, text='종합수율(%)', height=yield_factory_chart_height)
                fig_factory_trend.update_traces(texttemplate='%{text:.2f}%', textposition='top center', textfont=dict(size=yield_factory_label_size, color='black'))
                fig_factory_trend.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(factory_yield_trend['period'].unique()), title_font_size=yield_factory_axis_title_size, tickfont_size=yield_factory_axis_tick_size)
                fig_factory_trend.update_yaxes(title_font_size=yield_factory_axis_title_size, tickfont_size=yield_factory_axis_tick_size)
                fig_factory_trend.update_layout(title_font_size=yield_factory_axis_title_size)
                st.plotly_chart(fig_factory_trend, use_container_width=True)

            st.divider()
            
            # --- 공정별 수율 추이 분석 ---
            # 제목 영역을 별도로 분리하여 겹침 방지
            st.markdown(f"### {agg_level} 공정별 수율 추이 분석")
            st.markdown("---")  # 구분선 추가
            
            # 동적 필터 - 드롭다운 multiselect 형태로 변경
            filter_col1, filter_col2 = st.columns(2)
            
            with filter_col1:
                # 공장 필터 - multiselect 드롭다운
                all_factories_for_process = sorted(df_filtered['공장'].unique())
                selected_factories_process = st.multiselect(
                    "공장 선택",
                    options=all_factories_for_process,
                    default=all_factories_for_process,
                    key="process_yield_factory_multiselect",
                    help="분석할 공장을 선택합니다. 복수 선택 가능합니다."
                )
            
            with filter_col2:
                # 공정 필터 - multiselect 드롭다운
                all_processes_for_filter = [p for p in PROCESS_MASTER_ORDER if p in df_filtered['공정코드'].unique()]
                process_options = []
                process_mapping = {}
                
                for process in all_processes_for_filter:
                    display_name = process.split('] ')[1] if '] ' in process else process
                    process_options.append(display_name)
                    process_mapping[display_name] = process
                
                selected_process_names = st.multiselect(
                    "공정 선택",
                    options=process_options,
                    default=process_options,
                    key="process_yield_process_multiselect",
                    help="분석할 공정을 선택합니다. 복수 선택 가능합니다."
                )
                
                # 선택된 표시명을 실제 공정 코드로 변환
                selected_processes = [process_mapping[name] for name in selected_process_names]
            
            # 표시 방식 선택 옵션
            st.markdown("**표시 방식**")
            display_mode = st.radio(
                "데이터 표시 방식을 선택하세요:",
                options=["공장별로 구분하여 표시", "전체 공장 합산하여 표시"],
                index=0,
                key="process_yield_display_mode",
                help="공장별 구분: 각 공장-공정 조합별로 별도 라인 표시 / 전체 합산: 모든 공장의 동일 공정 데이터를 합산하여 표시",
                horizontal=True
            )
            
            if selected_factories_process and selected_processes:
                # 선택된 조건에 따라 데이터 필터링
                df_process_filtered = df_filtered[
                    (df_filtered['공장'].isin(selected_factories_process)) &
                    (df_filtered['공정코드'].isin(selected_processes))
                ].copy()
                
                if not df_process_filtered.empty:
                    # 공정별 시간대별 수율 계산
                    df_resampled_process = get_resampled_data(
                        df_process_filtered, 
                        agg_level, 
                        ['총_생산수량', '총_양품수량'], 
                        group_by_cols=['period', '공장', '공정코드']
                    )
                    
                    if not df_resampled_process.empty:
                        # 수율 계산
                        df_resampled_process['총_생산수량'] = df_resampled_process['총_생산수량'].replace(0, pd.NA)
                        with pd.option_context('mode.use_inf_as_na', True):
                            df_resampled_process['수율(%)'] = (
                                df_resampled_process['총_양품수량'] / df_resampled_process['총_생산수량'] * 100
                            ).fillna(0)
                        
                        # 선택된 표시 방식에 따라 데이터 그룹화
                        if display_mode == "공장별로 구분하여 표시":
                            # 공장-공정 조합별로 그룹화
                            process_yield_data = df_resampled_process.groupby(['period', '공장', '공정코드'])['수율(%)'].mean().reset_index()
                            # 공장-공정 조합 컬럼 생성
                            process_yield_data['공장_공정'] = process_yield_data['공장'].astype(str) + ' - ' + process_yield_data['공정코드'].str.split('] ').str[1].fillna(process_yield_data['공정코드'])
                            color_column = '공장_공정'
                        else:
                            # 전체 공장 합산 - 공정별로만 그룹화
                            process_yield_data = df_resampled_process.groupby(['period', '공정코드'])['수율(%)'].mean().reset_index()
                            color_column = '공정코드'
                        
                        # 그래프 설정 옵션
                        with st.expander("📊 차트 설정", expanded=False):
                            col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                            with col_set1:
                                process_yield_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="process_yield_label_size")
                            with col_set2:
                                process_yield_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="process_yield_axis_title_size")
                            with col_set3:
                                process_yield_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="process_yield_axis_tick_size")
                            with col_set4:
                                process_yield_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="process_yield_chart_height")
                            
                            # 수율 범위 설정
                            st.markdown("**수율 범위 설정**")
                            # 데이터의 최소값과 최대값을 기본값으로 설정
                            min_yield_value = max(0, int(process_yield_data['수율(%)'].min()) - 5)
                            max_yield_value = min(120, int(process_yield_data['수율(%)'].max()) + 5)
                            
                            # 범위 슬라이더 (하나의 컨트롤로 최소/최대값 동시 조정)
                            yield_range = st.slider(
                                "수율(%) 축 범위",
                                min_value=0, 
                                max_value=120, 
                                value=(min_yield_value, max_yield_value), 
                                step=1, 
                                key="process_yield_range",
                                help="차트 Y축의 수율 범위를 조정합니다. 왼쪽 핸들은 최소값, 오른쪽 핸들은 최대값입니다."
                            )
                            yield_min_range, yield_max_range = yield_range
                        
                        # 꺾은선 그래프 생성
                        if not process_yield_data.empty:
                            # 제목 설정
                            if display_mode == "공장별로 구분하여 표시":
                                chart_title = f'<b>{agg_level} 공장별 공정별 수율 추이</b>'
                            else:
                                chart_title = f'<b>{agg_level} 공정별 수율 추이 (전체 공장 합산)</b>'
                            
                            fig_process_yield = px.line(
                                process_yield_data.sort_values('period'), 
                                x='period', 
                                y='수율(%)', 
                                color=color_column,
                                title="",  # 차트 제목 제거하여 겹침 방지
                                markers=True,
                                text='수율(%)',
                                height=process_yield_chart_height
                            )
                            
                            # 그래프 스타일 설정
                            fig_process_yield.update_traces(
                                texttemplate='%{text:.1f}%',
                                textposition='top center',
                                textfont=dict(size=process_yield_label_size, color='black')
                            )
                            fig_process_yield.update_xaxes(
                                type='category',
                                categoryorder='array',
                                categoryarray=sorted(process_yield_data['period'].unique()),
                                title_font_size=process_yield_axis_title_size,
                                tickfont_size=process_yield_axis_tick_size
                            )
                            fig_process_yield.update_yaxes(
                                title_font_size=process_yield_axis_title_size,
                                tickfont_size=process_yield_axis_tick_size,
                                range=[yield_min_range, yield_max_range]
                            )
                            fig_process_yield.update_layout(
                                title_font_size=process_yield_axis_title_size,
                                margin=dict(t=20, b=100),  # 상단 여백 감소, 하단 여백 증가
                                legend=dict(
                                    orientation="h",
                                    yanchor="top",
                                    y=-0.15,  # 범례를 차트 하단으로 완전 이동
                                    xanchor="center",
                                    x=0.5
                                )
                            )
                            
                            st.plotly_chart(fig_process_yield, use_container_width=True)
                        else:
                            st.info("선택된 조건에 해당하는 데이터가 없습니다.")
                    else:
                        st.info("선택된 기간에 해당하는 데이터가 없습니다.")
                else:
                    st.info("선택된 공장/공정 조건에 해당하는 데이터가 없습니다.")
            else:
                st.info("분석할 공장과 공정을 선택해주세요.")
            
            st.divider()
            
            # --- 제품군별 종합 수율 추이 ---
            st.subheader(f"{agg_level} 제품군별 종합 수율 추이")
            
            # 공장 선택 필터
            all_factories = ['전체'] + sorted(df_filtered['공장'].unique())
            selected_factory = st.selectbox(
                "공장 선택", 
                options=all_factories, 
                key="yield_factory_select",
                help="분석할 공장을 선택합니다. '전체' 선택 시 모든 공장의 데이터를 종합하여 분석합니다."
            )

            # 선택된 공장에 따라 데이터 필터링
            if selected_factory == '전체':
                df_yield_factory_filtered = df_filtered.copy()
            else:
                df_yield_factory_filtered = df_filtered[df_filtered['공장'] == selected_factory].copy()
            
            df_resampled_product = get_resampled_data(df_yield_factory_filtered, agg_level, ['총_생산수량', '총_양품수량'], group_by_cols=['period', '신규분류요약', '공정코드'])

            if not df_resampled_product.empty and '신규분류요약' in df_resampled_product.columns:
                # 1. 완제품 실적 (최종 공정 기준)
                last_process = PROCESS_MASTER_ORDER[-1]
                final_prod_count = df_resampled_product[df_resampled_product['공정코드'] == last_process]\
                    .groupby(['period', '신규분류요약'])['총_양품수량'].sum().reset_index()\
                    .rename(columns={'총_양품수량': '완제품_제조개수'})

                # 2. 종합 수율 (전 공정 수율의 곱)
                df_resampled_product_copy = df_resampled_product.copy()
                df_resampled_product_copy['총_생산수량'] = df_resampled_product_copy['총_생산수량'].replace(0, pd.NA)
                with pd.option_context('mode.use_inf_as_na', True):
                    df_resampled_product_copy['개별수율'] = (df_resampled_product_copy['총_양품수량'] / df_resampled_product_copy['총_생산수량']).fillna(1.0)
                
                product_yield_trend = df_resampled_product_copy.groupby(['period', '신규분류요약'])['개별수율'].prod().reset_index()
                product_yield_trend = product_yield_trend.rename(columns={'개별수율': '종합수율(%)'})
                product_yield_trend['종합수율(%)'] *= 100

                # 3. 데이터 병합 및 보정: 완제품 실적이 0이면 수율도 0으로 처리
                product_yield_trend = pd.merge(product_yield_trend, final_prod_count, on=['period', '신규분류요약'], how='left').fillna({'완제품_제조개수': 0})
                product_yield_trend.loc[product_yield_trend['완제품_제조개수'] == 0, '종합수율(%)'] = 0
                
                all_product_groups = sorted(df_resampled_product['신규분류요약'].dropna().unique())

                if not all_product_groups:
                    st.info("선택된 공장에 제품군 데이터가 없습니다.")
                else:
                    for group in all_product_groups:
                        if f"product_group_{group}" not in st.session_state: 
                            st.session_state[f"product_group_{group}"] = True
                    
                    st.markdown("##### 표시할 제품군 선택")
                    btn_cols = st.columns(8)
                    with btn_cols[0]:
                        if st.button("제품군 전체 선택", key="select_all_products_yield", use_container_width=True):
                            for group in all_product_groups: st.session_state[f"product_group_{group}"] = True
                            st.rerun()
                    with btn_cols[1]:
                        if st.button("제품군 전체 해제", key="deselect_all_products_yield", use_container_width=True):
                            for group in all_product_groups: st.session_state[f"product_group_{group}"] = False
                            st.rerun()
                    
                    st.write("")
                    num_cols = 5
                    cols = st.columns(num_cols)
                    selected_product_groups = []
                    for i, group in enumerate(all_product_groups):
                        with cols[i % num_cols]:
                            if st.checkbox(group, key=f"product_group_{group}"):
                                selected_product_groups.append(group)
                    
                    combine_yield = st.checkbox("선택항목 합쳐서 보기", key="combine_product_yield", help="선택한 제품군들의 실적을 합산하여 단일 종합 수율 추이를 분석합니다.")

                    # 그래프 설정 옵션
                    with st.expander("📊 차트 설정", expanded=False):
                        col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                        with col_set1:
                            yield_product_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="yield_product_label_size")
                        with col_set2:
                            yield_product_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="yield_product_axis_title_size")
                        with col_set3:
                            yield_product_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="yield_product_axis_tick_size")
                        with col_set4:
                            yield_product_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="yield_product_chart_height")

                    if selected_product_groups:
                        if combine_yield:
                            df_filtered_for_combine = df_resampled_product[df_resampled_product['신규분류요약'].isin(selected_product_groups)]
                            
                            # 1. 실적 데이터 (최종 공정 기준)
                            last_process = PROCESS_MASTER_ORDER[-1]
                            bar_combined = df_filtered_for_combine[df_filtered_for_combine['공정코드'] == last_process].groupby('period')['총_양품수량'].sum().reset_index().rename(columns={'총_양품수량': '완제품_제조개수'})
                            
                            # 2. 수율 데이터 (개별 공정 수율의 곱)
                            df_yield_combined_base = df_filtered_for_combine.groupby(['period', '공정코드']).agg(총_생산수량=('총_생산수량', 'sum'), 총_양품수량=('총_양품수량', 'sum')).reset_index()
                            df_yield_combined_base['총_생산수량'] = df_yield_combined_base['총_생산수량'].replace(0, pd.NA)
                            with pd.option_context('mode.use_inf_as_na', True):
                                df_yield_combined_base['개별수율'] = (df_yield_combined_base['총_양품수량'] / df_yield_combined_base['총_생산수량']).fillna(1.0)
                            line_combined = df_yield_combined_base.groupby('period')['개별수율'].prod().reset_index(name='종합수율(%)')
                            line_combined['종합수율(%)'] *= 100
                            
                            # 3. 데이터 병합 및 보정
                            df_to_plot = pd.merge(bar_combined, line_combined, on='period', how='outer').fillna(0)
                            df_to_plot.loc[df_to_plot['완제품_제조개수'] == 0, '종합수율(%)'] = 0
                            
                            if not df_to_plot.empty:
                                fig_product_trend = px.line(df_to_plot.sort_values('period'), x='period', y='종합수율(%)', title=f'<b>{agg_level} 선택 제품군 통합 수율 추이 ({selected_factory})</b>', markers=True, text='종합수율(%)', height=yield_product_chart_height)
                                fig_product_trend.update_traces(texttemplate='%{text:.2f}%', textposition='top center', textfont=dict(size=yield_product_label_size, color='black'))
                                fig_product_trend.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(df_to_plot['period'].unique()), title_font_size=yield_product_axis_title_size, tickfont_size=yield_product_axis_tick_size)
                                fig_product_trend.update_yaxes(title_font_size=yield_product_axis_title_size, tickfont_size=yield_product_axis_tick_size)
                                fig_product_trend.update_layout(title_font_size=yield_product_axis_title_size)
                                st.plotly_chart(fig_product_trend, use_container_width=True)
                        else:
                            df_to_plot = product_yield_trend[product_yield_trend['신규분류요약'].isin(selected_product_groups)]
                            if not df_to_plot.empty:
                                fig_product_trend = px.line(df_to_plot.sort_values('period'), x='period', y='종합수율(%)', color='신규분류요약', title=f'<b>{agg_level} 제품군별 종합 수율 추이 ({selected_factory})</b>', markers=True, text='종합수율(%)', height=yield_product_chart_height)
                                fig_product_trend.update_traces(texttemplate='%{text:.2f}%', textposition='top center', textfont=dict(size=yield_product_label_size, color='black'))
                                fig_product_trend.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(df_to_plot['period'].unique()), title_font_size=yield_product_axis_title_size, tickfont_size=yield_product_axis_tick_size)
                                fig_product_trend.update_yaxes(title_font_size=yield_product_axis_title_size, tickfont_size=yield_product_axis_tick_size)
                                fig_product_trend.update_layout(title_font_size=yield_product_axis_title_size)
                                st.plotly_chart(fig_product_trend, use_container_width=True)
                    else:
                        st.info("차트를 표시할 제품군을 선택해주세요.")

            # --- 공장/공정별 평균 수율 ---
            df_total_agg = aggregate_overall_data(df_filtered, 'yield')
            all_factories_in_period = sorted(df_filtered['공장'].unique())
            plot_horizontal_bar_chart_all_processes(df_total_agg, {'rate_col': '평균_수율', 'y_axis_title': '평균 수율', 'chart_title': '공장/공정별 평균 수율'}, all_factories_in_period, PROCESS_MASTER_ORDER)

        with side_col:
            st.markdown(analyze_yield_data(df_total_agg))
            st.divider()
            st.subheader("데이터 원본")
            st.dataframe(df_filtered, use_container_width=True, height=500)
        
        # 다운로드 섹션 추가
        create_download_section(df_filtered, "수율분석", agg_level, start_date, end_date)
elif selected_tab == "불량유형별 분석":
    if df_defect_orig.empty:
        st.info("해당 분석을 위해서는 '불량실적현황(최적화)' 데이터가 필요합니다.")
    else:
        df_defect_filtered, start_date, end_date, agg_level = create_shared_filter_controls(df_defect_orig)

        if df_defect_filtered.empty:
            st.info("선택된 기간에 분석에 필요한 불량 데이터가 없습니다.")
        elif '생산수량' not in df_defect_filtered.columns:
            st.error("불량 데이터 파일에 '생산수량' 컬럼이 없어 불량률을 계산할 수 없습니다.")
        else:
            if '유형별_불량수량' in df_defect_filtered.columns:
                df_defect_filtered['유형별_불량수량'] = pd.to_numeric(df_defect_filtered['유형별_불량수량'], errors='coerce').fillna(0)
            
            main_col, side_col = st.columns([2.8, 1])

            with main_col:
                with st.expander("세부 필터 및 옵션", expanded=True):
                    filter_data_source = df_defect_filtered.copy()
                    filter_options_map = {
                        "공장": "공장",
                        "신규분류요약": "제품군",
                        "사출기계코드": "사출 기계",
                        "공정기계코드": "공정 기계"
                    }
                    available_filters = [k for k in filter_options_map if k in filter_data_source.columns]

                    # 최초 실행 시 모든 필터 전체 선택
                    for key in available_filters:
                        options = sorted(filter_data_source[key].dropna().unique())
                        session_key = f"ms_{key}"
                        if session_key not in st.session_state:
                            st.session_state[session_key] = options

                    # 전체 선택/해제 버튼
                    btn_cols = st.columns(2)
                    with btn_cols[0]:
                        if st.button("세부필터 전체 선택"):
                            for key in available_filters:
                                options = sorted(filter_data_source[key].dropna().unique())
                                st.session_state[f"ms_{key}"] = options
                            st.rerun()
                    with btn_cols[1]:
                        if st.button("세부필터 전체 해제"):
                            for key in available_filters:
                                st.session_state[f"ms_{key}"] = []
                            st.rerun()

                    # 동적 필터링
                    selections = {}
                    filtered_df = filter_data_source.copy()
                    for i, key in enumerate(available_filters):
                        # 앞쪽 필터 선택값에 따라 옵션 제한
                        if i > 0:
                            prev_keys = available_filters[:i]
                            for pk in prev_keys:
                                selected = st.session_state.get(f"ms_{pk}", [])
                                if selected:
                                    filtered_df = filtered_df[filtered_df[pk].isin(selected)]
                        options = sorted(filtered_df[key].dropna().unique())
                        selections[key] = st.multiselect(
                            filter_options_map[key], options, default=st.session_state.get(f"ms_{key}", options),
                            key=f"ms_{key}", label_visibility="collapsed", placeholder=filter_options_map[key]
                        )

                df_display = filtered_df.copy()
                for key, selected_values in selections.items():
                    if selected_values:
                        df_display = df_display[df_display[key].isin(selected_values)]
                
                st.markdown("---")
                st.markdown("<h6>불량 유형 필터</h6>", unsafe_allow_html=True)
                defect_options = sorted(df_display['불량명'].dropna().unique())
                if 'selected_defects' not in st.session_state: st.session_state.selected_defects = defect_options
                
                defect_btn_cols = st.columns(4)
                with defect_btn_cols[0]:
                    if st.button("불량 유형 전체 선택", use_container_width=True): st.session_state.selected_defects = defect_options
                with defect_btn_cols[1]:
                    if st.button("불량 유형 전체 해제", use_container_width=True): st.session_state.selected_defects = []
                
                st.multiselect("표시할 불량 유형 선택", options=defect_options, key='selected_defects', label_visibility="collapsed")
            
            if st.session_state.selected_defects:
                df_display = df_display[df_display['불량명'].isin(st.session_state.selected_defects)]
            else: 
                df_display = df_display[df_display['불량명'].isin([])]
            
            prod_key_cols = ['date', '공장', '신규분류요약', '사출기계코드', '공정기계코드', '생산수량']
            available_prod_key_cols = [col for col in prod_key_cols if col in df_display.columns]
            prod_data_source = df_display[available_prod_key_cols].drop_duplicates()

            st.divider()
            st.subheader("주요 불량 원인 분석 (파레토)", anchor=False)
            if df_display.empty or '유형별_불량수량' not in df_display.columns or df_display['유형별_불량수량'].sum() == 0:
                st.warning("선택된 필터 조건에 해당하는 불량 데이터가 없습니다.")
            else:
                plot_pareto_chart(df_display, title="선택된 조건의 불량유형 파레토 분석", defect_qty_col='유형별_불량수량')

            st.divider()
            st.subheader(f"{agg_level} 총 불량 수량 및 불량률 추이", anchor=False)
            total_defect_resampled = get_resampled_data(df_display, agg_level, ['유형별_불량수량'], group_by_cols=['period'])
            total_prod_resampled = get_resampled_data(prod_data_source, agg_level, ['생산수량'], group_by_cols=['period']).rename(columns={'생산수량': '총_생산수량'})
            
            if not total_defect_resampled.empty:
                combo_data = pd.merge(total_defect_resampled, total_prod_resampled, on='period', how='outer').fillna(0)
                production_for_rate = combo_data['총_생산수량'].replace(0, pd.NA)
                with pd.option_context('mode.use_inf_as_na', True):
                    combo_data['총_불량률(%)'] = (100 * combo_data['유형별_불량수량'] / production_for_rate).fillna(0)
                
                min_rate_val = combo_data['총_불량률(%)'].min()
                max_rate_val = combo_data['총_불량률(%)'].max()
                
                slider_max_bound = max(50.0, max_rate_val * 1.2)
                
                rate_range = st.slider(
                    "총 불량률(%) 축 범위 조절",
                    min_value=0.0,
                    max_value=round(slider_max_bound, -1),
                    value=(float(min_rate_val), float(max_rate_val)),
                    step=1.0,
                    format="%.0f%%"
                )

                fig_combo = make_subplots(specs=[[{"secondary_y": True}]])
                fig_combo.add_trace(go.Bar(x=combo_data['period'], y=combo_data['유형별_불량수량'], name='총 불량 수량', text=combo_data['유형별_불량수량'], texttemplate='%{text:,.0f}', textposition='auto'), secondary_y=False)
                fig_combo.add_trace(go.Scatter(x=combo_data['period'], y=combo_data['총_불량률(%)'], name='총 불량률 (%)', mode='lines+markers+text', text=combo_data['총_불량률(%)'], texttemplate='%{text:.2f}%', textposition='top center', connectgaps=False, textfont=dict(size=16, color='black')), secondary_y=True)
                fig_combo.update_layout(height=600, title_text=f"<b>{agg_level} 총 불량 수량 및 불량률 추이</b>", margin=dict(t=120), legend=dict(orientation="h", yanchor="bottom", y=1.10, xanchor="right", x=1))
                fig_combo.update_yaxes(title_text="<b>총 불량 수량 (개)</b>", secondary_y=False); fig_combo.update_yaxes(title_text="<b>총 불량률 (%)</b>", secondary_y=True, range=rate_range)
                fig_combo.update_xaxes(title_text=f"<b>{agg_level.replace('별', '')}</b>", type='category', categoryorder='array', categoryarray=sorted(combo_data['period'].unique()))
                
                # 자동 라벨 겹침 방지 기능 활성화 (꺾은선 그래프에만 적용)
                fig_combo.update_layout(uniformtext_minsize=12, uniformtext_mode='hide')

                st.plotly_chart(fig_combo, use_container_width=True)
            else:
                st.info("선택된 필터 조건에 해당하는 추이 데이터가 없습니다.")

            st.divider()
            st.subheader(f"{agg_level} 불량 유형별 불량률 추이", anchor=False)
            
            prod_resampled = get_resampled_data(prod_data_source, agg_level, ['생산수량'], group_by_cols=['period']).rename(columns={'생산수량': '기간별_총생산량'})
            defect_resampled = get_resampled_data(df_display, agg_level, ['유형별_불량수량'], group_by_cols=['period', '불량명'])
            
            if not defect_resampled.empty:
                trend_final_data = pd.merge(defect_resampled, prod_resampled, on='period', how='left')
                production_for_rate_ind = trend_final_data['기간별_총생산량'].replace(0, pd.NA)
                with pd.option_context('mode.use_inf_as_na', True):
                    trend_final_data['불량률(%)'] = (100 * trend_final_data['유형별_불량수량'] / production_for_rate_ind).fillna(0)

                chart_option_cols = st.columns([2, 1, 1])
                with chart_option_cols[0]:
                     top_n_defects = st.number_input(
                         "상위 N개 불량 유형 표시", 
                         min_value=1, 
                         max_value=len(trend_final_data['불량명'].unique()), 
                         value=len(trend_final_data['불량명'].unique()), 
                         step=1,
                         help="평균 불량률이 높은 순으로 상위 N개 유형의 추이만 표시합니다."
                     )
                with chart_option_cols[1]:
                    st.markdown("<div style='padding-top: 28px;'></div>", unsafe_allow_html=True)
                    show_labels = st.toggle("차트 라벨 표시", value=True)

                avg_defect_rates = trend_final_data.groupby('불량명')['불량률(%)'].mean().nlargest(top_n_defects).index.tolist()
                trend_final_data_top_n = trend_final_data[trend_final_data['불량명'].isin(avg_defect_rates)]
                
                fig_trend_rate = px.line(trend_final_data_top_n.sort_values('period'), x='period', y='불량률(%)', color='불량명', title=f"<b>{agg_level} 불량 유형별 불량률 추이</b>", markers=True, text='불량률(%)' if show_labels else None, height=600)
                fig_trend_rate.update_traces(texttemplate='%{text:.4f}%', textposition='top center', textfont=dict(size=16, color='black'), connectgaps=False)
                fig_trend_rate.update_layout(legend_title_text='불량 유형', xaxis_title=f"<b>{agg_level.replace('별', '')}</b>", yaxis_title="<b>불량률 (%)</b>")
                fig_trend_rate.update_xaxes(type='category', categoryorder='array', categoryarray=sorted(trend_final_data_top_n['period'].unique()))
                st.plotly_chart(fig_trend_rate, use_container_width=True)
            else:
                st.info("선택된 필터 조건에 해당하는 추이 데이터가 없습니다.")

            with side_col:
                st.markdown(analyze_defect_data(df_defect_filtered))
                st.divider()
                st.subheader("데이터 원본 (필터링됨)")
                st.dataframe(df_display, use_container_width=True, height=500)
        
        # 다운로드 섹션 추가
        create_download_section(df_defect_filtered, "불량유형별분석", agg_level, start_date, end_date)

elif selected_tab == "가동률 분석":
    # 고도화된 가동률 분석 시스템
    df_filtered, start_date, end_date, agg_level = create_shared_filter_controls(df_utilization_orig)
    
    if df_filtered.empty:
        st.info(f"선택된 기간에 해당하는 가동률 데이터가 없습니다.")
    else:
        # 이론상 생산량이 0인 설비 분석
        zero_theory_mask = df_filtered['이론상_총_생산량'] == 0
        valid_data_mask = ~zero_theory_mask
        
        # 가동률 계산이 가능한 데이터만 사용
        df_valid = df_filtered[valid_data_mask].copy()
        df_zero_theory = df_filtered[zero_theory_mask & (df_filtered['총_생산수량'] > 0)].copy()
        
        # 기본 가동률 계산 (유효한 데이터만)
        if not df_valid.empty:
            with pd.option_context('mode.use_inf_as_na', True):
                df_valid['가동률(%)'] = (100 * df_valid['총_생산수량'] / df_valid['이론상_총_생산량']).fillna(0)
        
        # 데이터 현황 알림
        if not df_zero_theory.empty:
            st.warning(f"⚠️ 이론상 생산량이 0이지만 실제 생산이 있는 설비 {df_zero_theory['기계코드'].nunique()}대가 있습니다. 이 설비들은 가동률 계산에서 제외됩니다.")
        
        # 분석 대상 데이터 설정
        df_filtered = df_valid
        
        # 상단 KPI 영역
        st.subheader("📊 가동률 현황 요약")
        kpi_cols = st.columns(4)
        
        avg_utilization = df_filtered['가동률(%)'].mean()
        best_equipment = df_filtered.loc[df_filtered['가동률(%)'].idxmax()] if not df_filtered.empty else None
        worst_equipment = df_filtered.loc[df_filtered['가동률(%)'].idxmin()] if not df_filtered.empty else None
        total_equipment = df_filtered['기계코드'].nunique()
        
        with kpi_cols[0]:
            st.metric("평균 가동률", f"{avg_utilization:.2f}%")
        with kpi_cols[1]:
            st.metric("총 설비 수", f"{total_equipment}대")
        with kpi_cols[2]:
            if best_equipment is not None:
                st.metric("최고 가동률", f"{best_equipment['가동률(%)']:.1f}%", 
                         help=f"설비: {best_equipment['기계코드']}")
        with kpi_cols[3]:
            if worst_equipment is not None:
                st.metric("최저 가동률", f"{worst_equipment['가동률(%)']:.1f}%",
                         help=f"설비: {worst_equipment['기계코드']}")
        
        st.divider()
        
        # 탭 기반 분석 영역
        analysis_tabs = st.tabs(["🏭 전체 현황", "⚙️ 설비 분석", "🏗️ 설비 배치도", "📈 비교 분석", "🎯 성과 분석"])
        
        with analysis_tabs[0]:  # 전체 현황
            col1, col2 = st.columns([2, 1])
            
            with col1:
                # 공장별 가동률 추이 (기존 차트 개선)
                df_resampled_util = get_resampled_data(df_filtered, agg_level, ['총_생산수량', '이론상_총_생산량'], 
                                                     group_by_cols=['period', '공장', '공정코드'])
                if not df_resampled_util.empty:
                    with pd.option_context('mode.use_inf_as_na', True):
                        df_resampled_util['평균_가동률'] = (100 * df_resampled_util['총_생산수량'] / df_resampled_util['이론상_총_생산량']).fillna(0)
                    
                    df_trend = df_resampled_util.groupby(['period', '공장'])['평균_가동률'].mean().reset_index()
                    
                    fig_trend = px.line(df_trend.sort_values('period'), x='period', y='평균_가동률', 
                                      color='공장', title=f'<b>{agg_level} 공장별 가동률 추이</b>', 
                                      markers=True, text='평균_가동률')
                    fig_trend.update_traces(texttemplate='%{text:.2f}%', textposition='top center', 
                                          textfont=dict(size=16, color='black'))
                    fig_trend.update_xaxes(type='category', categoryorder='array', 
                                         categoryarray=sorted(df_trend['period'].unique()))
                    fig_trend.update_layout(height=500)
                    st.plotly_chart(fig_trend, use_container_width=True)
                
                # 공정별 가동률 분포
                st.subheader("공정별 가동률 분포")
                process_util = df_filtered.groupby('공정코드')['가동률(%)'].agg(['mean', 'std', 'count']).reset_index()
                process_util.columns = ['공정코드', '평균_가동률', '표준편차', '설비수']
                
                fig_process = px.bar(process_util, x='공정코드', y='평균_가동률', 
                                   title='<b>공정별 평균 가동률</b>', text='평균_가동률')
                fig_process.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
                fig_process.update_layout(height=400)
                st.plotly_chart(fig_process, use_container_width=True)
            
            with col2:
                # AI 분석 브리핑
                st.markdown(analyze_utilization_data(df_filtered))
                
                # 가동률 분포 히스토그램
                st.subheader("가동률 분포")
                
                # 실제 가동률 범위 계산
                util_min = df_filtered['가동률(%)'].min()
                util_max = df_filtered['가동률(%)'].max()
                
                fig_hist = px.histogram(df_filtered, x='가동률(%)', nbins=20, 
                                      title='<b>설비별 가동률 분포</b>')
                
                # x축 범위를 실제 데이터 범위로 조정 (약간의 여백 추가)
                range_padding = (util_max - util_min) * 0.05  # 5% 여백
                fig_hist.update_xaxes(range=[max(0, util_min - range_padding), util_max + range_padding])
                fig_hist.update_layout(height=300)
                st.plotly_chart(fig_hist, use_container_width=True)
        
        with analysis_tabs[1]:  # 설비 분석
            st.subheader("🔍 설비별 상세 분석")
            
            # 설비 선택 필터
            filter_cols = st.columns(3)
            with filter_cols[0]:
                selected_factory = st.selectbox("공장 선택", 
                                               options=['전체'] + sorted(df_filtered['공장'].unique()),
                                               key='util_factory_filter')
            with filter_cols[1]:
                if selected_factory != '전체':
                    factory_processes = df_filtered[df_filtered['공장'] == selected_factory]['공정코드'].unique()
                else:
                    factory_processes = df_filtered['공정코드'].unique()
                selected_process = st.selectbox("공정 선택",
                                              options=['전체'] + sorted(factory_processes),
                                              key='util_process_filter')
            with filter_cols[2]:
                # 가동률 범위 필터
                min_util, max_util = st.slider("가동률 범위 (%)", 
                                             min_value=0, max_value=100, 
                                             value=(0, 100), key='util_range_filter')
            
            # 필터 적용
            df_equipment = df_filtered.copy()
            if selected_factory != '전체':
                df_equipment = df_equipment[df_equipment['공장'] == selected_factory]
            if selected_process != '전체':
                df_equipment = df_equipment[df_equipment['공정코드'] == selected_process]
            df_equipment = df_equipment[(df_equipment['가동률(%)'] >= min_util) & 
                                      (df_equipment['가동률(%)'] <= max_util)]
            
            if not df_equipment.empty:
                # 설비별 가동률 테이블
                equipment_summary = df_equipment.groupby(['기계코드', '공장', '공정코드']).agg({
                    '가동률(%)': ['mean', 'std', 'min', 'max', 'count'],
                    '총_생산수량': 'sum',
                    '이론상_총_생산량': 'sum'
                }).round(2)
                
                equipment_summary.columns = ['평균_가동률', '가동률_표준편차', '최저_가동률', 
                                           '최고_가동률', '데이터수', '총_생산수량', '이론상_총_생산량']
                equipment_summary = equipment_summary.reset_index()
                equipment_summary = equipment_summary.sort_values('평균_가동률', ascending=False)
                
                st.dataframe(equipment_summary, use_container_width=True, height=400)
                
                # 설비별 가동률 차트
                if len(equipment_summary) <= 20:  # 너무 많으면 차트가 복잡해짐
                    fig_equipment = px.scatter(equipment_summary, x='기계코드', y='평균_가동률',
                                             color='공장', size='총_생산수량',
                                             title='<b>설비별 가동률 및 생산량</b>',
                                             hover_data=['공정코드', '가동률_표준편차'])
                    fig_equipment.update_layout(height=500)
                    st.plotly_chart(fig_equipment, use_container_width=True)
            else:
                st.info("선택한 조건에 해당하는 설비가 없습니다.")
            
            # 이론상 생산량이 0인 설비 별도 표시
            if not df_zero_theory.empty:
                st.subheader("⚠️ 계획 생산량 없이 실제 생산한 설비")
                st.caption("이론상 생산량이 0이지만 실제 생산이 있는 설비들입니다. 가동률 계산이 불가능합니다.")
                
                zero_theory_summary = df_zero_theory.groupby(['기계코드', '공장', '공정코드']).agg({
                    '총_생산수량': 'sum',
                    '이론상_총_생산량': 'sum'
                }).round(2).reset_index()
                zero_theory_summary = zero_theory_summary.sort_values('총_생산수량', ascending=False)
                
                # 컬럼명 변경
                zero_theory_summary.columns = ['기계코드', '공장', '공정코드', '실제_생산량', '계획_생산량']
                
                st.dataframe(zero_theory_summary, use_container_width=True, height=300)
                
                # 이런 설비들의 생산량 분포 차트
                if len(zero_theory_summary) > 0:
                    fig_zero_theory = px.bar(zero_theory_summary.head(15), 
                                           x='기계코드', y='실제_생산량', 
                                           color='공장',
                                           title='<b>계획 없이 생산한 설비별 실제 생산량 (상위 15개)</b>')
                    fig_zero_theory.update_xaxes(tickangle=45)
                    fig_zero_theory.update_layout(height=400)
                    st.plotly_chart(fig_zero_theory, use_container_width=True)
        
        with analysis_tabs[2]:  # 설비 배치도
            st.subheader("🏗️ 설비 배치도")
            
            # 배치도 옵션 선택
            layout_cols = st.columns(4)
            with layout_cols[0]:
                layout_view = st.selectbox("배치도 유형", 
                                         options=["공장별 배치도", "공정별 배치도", "통합 배치도"],
                                         key='layout_view')
            with layout_cols[1]:
                color_metric = st.selectbox("색상 기준", 
                                          options=["가동률", "생산량", "안정성"],
                                          key='color_metric')
            with layout_cols[2]:
                size_metric = st.selectbox("크기 기준", 
                                         options=["생산량", "가동률", "균등"],
                                         key='size_metric')
            with layout_cols[3]:
                show_labels = st.checkbox("설비명 표시", value=True, key='show_labels')
            
            # 배치도별 설비 데이터 준비 (모든 설비 포함)
            # 데이터 정규화 함수 정의
            def normalize_text(text):
                """텍스트 정규화: 공백 제거, 특수문자 통일"""
                if pd.isna(text):
                    return text
                return str(text).strip().replace(' ', '').replace('　', '')  # 일반 공백과 전각 공백 모두 제거
            
            # 1단계: 마스터 설비 리스트 생성 (가동율참고파일에서 로드)
            try:
                # 가동율참고파일 로드
                equipment_master_file = "가동율참고파일(기계별이론상CAPA).xlsx"
                if os.path.exists(equipment_master_file):
                    equipment_master = pd.read_excel(equipment_master_file)
                    # 컬럼명 매핑 (실제 파일 구조에 맞춰 조정)
                    required_cols = ['공장', '공정코드', '기계코드']
                    if all(col in equipment_master.columns for col in required_cols):
                        master_equipment = equipment_master[required_cols].copy()
                    else:
                        # 컬럼 순서가 다를 경우 첫 3개 컬럼 사용
                        master_equipment = equipment_master.iloc[:, :3].copy()
                        master_equipment.columns = required_cols
                    
                    # 데이터 정규화 적용
                    master_equipment['공장_정규화'] = master_equipment['공장'].apply(normalize_text)
                    master_equipment['공정코드_정규화'] = master_equipment['공정코드'].apply(normalize_text)
                    master_equipment['기계코드_정규화'] = master_equipment['기계코드'].apply(normalize_text)
                    master_equipment = master_equipment.drop_duplicates()
                    
                    st.info(f"📋 가동율참고파일에서 {len(master_equipment)}개 설비를 로드했습니다.")
                else:
                    # 파일이 없으면 기존 데이터에서 추출
                    master_equipment = df_utilization_orig[['기계코드', '공장', '공정코드']].drop_duplicates()
                    master_equipment['공장_정규화'] = master_equipment['공장'].apply(normalize_text)
                    master_equipment['공정코드_정규화'] = master_equipment['공정코드'].apply(normalize_text)
                    master_equipment['기계코드_정규화'] = master_equipment['기계코드'].apply(normalize_text)
                    st.warning("⚠️ 가동율참고파일을 찾을 수 없어 기존 데이터를 사용합니다.")
            except Exception as e:
                # 오류 시 기존 방식 사용
                master_equipment = df_utilization_orig[['기계코드', '공장', '공정코드']].drop_duplicates()
                master_equipment['공장_정규화'] = master_equipment['공장'].apply(normalize_text)
                master_equipment['공정코드_정규화'] = master_equipment['공정코드'].apply(normalize_text)
                master_equipment['기계코드_정규화'] = master_equipment['기계코드'].apply(normalize_text)
                st.warning(f"⚠️ 가동율참고파일 로드 오류: 기존 데이터를 사용합니다. ({str(e)})")
            
            # 2단계: 실적 데이터 준비 (배합 공정 제외)
            df_layout = df_utilization_orig[~df_utilization_orig['공정코드'].str.contains('배합', na=False)].copy()
            
            # 날짜 필터 적용
            if 'date' in df_layout.columns:
                df_layout = df_layout[
                    (df_layout['date'] >= pd.to_datetime(start_date)) & 
                    (df_layout['date'] <= pd.to_datetime(end_date))
                ]
            
            # 실적 데이터도 정규화
            df_layout['공장_정규화'] = df_layout['공장'].apply(normalize_text)
            df_layout['공정코드_정규화'] = df_layout['공정코드'].apply(normalize_text)  
            df_layout['기계코드_정규화'] = df_layout['기계코드'].apply(normalize_text)
            
            # 3단계: 실적이 있는 설비 집계 (정규화된 키로 그룹화)
            equipment_with_performance = df_layout.groupby(['기계코드_정규화', '공장_정규화', '공정코드_정규화']).agg({
                '가동률(%)': 'mean',
                '총_생산수량': 'sum',
                '이론상_총_생산량': 'sum',
                '기계코드': 'first',  # 원본 값 유지
                '공장': 'first',
                '공정코드': 'first'
            }).round(2).reset_index()
            
            # 4단계: 마스터 리스트와 실적 데이터 병합 (정규화된 키로 병합)
            equipment_layout_data = master_equipment.merge(
                equipment_with_performance, 
                left_on=['기계코드_정규화', '공장_정규화', '공정코드_정규화'],
                right_on=['기계코드_정규화', '공장_정규화', '공정코드_정규화'],
                how='left',
                suffixes=('_master', '_perf')
            )
            
            # 5단계: 데이터 정리 및 상태 판단
            # 원본 컬럼명 사용 (마스터 파일 우선)
            equipment_layout_data['공장'] = equipment_layout_data['공장_master'].fillna(equipment_layout_data.get('공장_perf', ''))
            equipment_layout_data['공정코드'] = equipment_layout_data['공정코드_master'].fillna(equipment_layout_data.get('공정코드_perf', ''))
            equipment_layout_data['기계코드'] = equipment_layout_data['기계코드_master'].fillna(equipment_layout_data.get('기계코드_perf', ''))
            
            # 실적이 없는 설비들의 값을 0으로 설정
            equipment_layout_data['가동률(%)'] = equipment_layout_data['가동률(%)'].fillna(0)
            equipment_layout_data['총_생산수량'] = equipment_layout_data['총_생산수량'].fillna(0)
            equipment_layout_data['이론상_총_생산량'] = equipment_layout_data['이론상_총_생산량'].fillna(0)
            
            # 6단계: 개선된 실적 상태 구분 로직 (30% 기준)
            def determine_status(row):
                # 실적 데이터가 병합되었는지 확인
                has_performance_data = not pd.isna(row.get('기계코드_perf', None))
                
                if has_performance_data:
                    # 실적 데이터가 있으면 가동률로 판단
                    utilization = row['가동률(%)']
                    if utilization == 0:
                        return '미가동'  # 가동률 0%는 미가동
                    elif utilization <= 30:
                        return '저가동'  # 1~30%는 저가동
                    else:
                        return '정상가동'  # 30% 초과는 정상가동
                else:
                    # 실적 데이터가 없으면 미가동
                    return '미가동'
            
            equipment_layout_data['실적_상태'] = equipment_layout_data.apply(determine_status, axis=1)
            
            # 전체 설비 현황 디버깅 정보 표시
            total_equipment = len(equipment_layout_data)
            active_count = len(equipment_layout_data[equipment_layout_data['실적_상태'] == '정상가동'])
            low_util_count = len(equipment_layout_data[equipment_layout_data['실적_상태'] == '저가동'])
            inactive_count = len(equipment_layout_data[equipment_layout_data['실적_상태'] == '미가동'])
            
            st.info(f"🔍 전체 설비 병합 결과: 전체 {total_equipment}개 | 정상가동 {active_count}개 | 저가동 {low_util_count}개 | 미가동 {inactive_count}개")
            
            # 가동률 안정성 계산 (실적이 있는 설비만 계산)
            equipment_stability = df_layout.groupby('기계코드_정규화')['가동률(%)'].std().fillna(0)
            
            # 안정성 값을 설비 데이터에 매핑 (정규화된 키로)
            stability_df = equipment_stability.reset_index().rename(columns={'가동률(%)': '가동률_편차'})
            equipment_layout_data = equipment_layout_data.merge(
                stability_df,
                on='기계코드_정규화', how='left'
            )
            equipment_layout_data['가동률_편차'] = equipment_layout_data['가동률_편차'].fillna(0)
            equipment_layout_data['가동률_안정성'] = 100 - equipment_layout_data['가동률_편차']
            equipment_layout_data['가동률_안정성'] = equipment_layout_data['가동률_안정성'].clip(lower=0)
            
            if layout_view == "공장별 배치도":
                # 공장별로 분리된 배치도
                factories = equipment_layout_data['공장'].dropna().unique()
                for factory in sorted(factories):
                    factory_data = equipment_layout_data[equipment_layout_data['공장'] == factory]
                    
                    st.subheader(f"📍 {factory} 설비 배치")
                    
                    # 공정별로 배치 (Grid Layout 시뮬레이션)
                    processes = sorted(factory_data['공정코드'].unique())
                    
                    # 색상 및 크기 매핑
                    color_column = {'가동률': '가동률(%)', '생산량': '총_생산수량', '안정성': '가동률_안정성'}[color_metric]
                    size_column = {'생산량': '총_생산수량', '가동률': '가동률(%)', '균등': None}[size_metric]
                    
                    # 각 공정별 서브플롯 생성
                    cols = st.columns(min(len(processes), 3))  # 최대 3열
                    for i, process in enumerate(processes):
                        process_data = factory_data[factory_data['공정코드'] == process]
                        
                        with cols[i % 3]:
                            st.markdown(f"**{process}**")
                            
                            # 설비 배치 시각화 (격자 형태)
                            equipment_count = len(process_data)
                            if equipment_count > 0:
                                # 격자 배치 계산 (겹침 방지)
                                grid_size = int(np.ceil(np.sqrt(equipment_count)))
                                # 최소 간격 보장 (라벨을 위한 여유 공간)
                                spacing = 1.5  # 설비간 간격 증가
                                x_positions = []
                                y_positions = []
                                
                                for idx, (_, equipment) in enumerate(process_data.iterrows()):
                                    x_pos = (idx % grid_size) * spacing
                                    y_pos = (idx // grid_size) * spacing
                                    x_positions.append(x_pos)
                                    y_positions.append(y_pos)
                                
                                process_data = process_data.copy()
                                process_data['x_pos'] = x_positions
                                process_data['y_pos'] = y_positions
                                
                                # 크기 설정 (더 큰 크기로 가시성 향상)
                                if size_column:
                                    sizes = process_data[size_column]
                                    # 정규화 (30-120 범위로 증대)
                                    size_min, size_max = sizes.min(), sizes.max()
                                    if size_max > size_min:
                                        normalized_sizes = 30 + 90 * (sizes - size_min) / (size_max - size_min)
                                    else:
                                        normalized_sizes = [80] * len(sizes)
                                else:
                                    normalized_sizes = [80] * len(process_data)
                                
                                # 배치도 차트 생성 (설비 상태별 구분)
                                # 설비 상태별 분리
                                active_equipment = process_data[process_data['실적_상태'] == '정상가동']
                                low_util_equipment = process_data[process_data['실적_상태'] == '저가동']
                                inactive_equipment = process_data[process_data['실적_상태'] == '미가동']
                                
                                # 기본 차트 생성 (정상가동 설비)
                                if not active_equipment.empty:
                                    fig_layout = px.scatter(
                                        active_equipment, 
                                        x='x_pos', y='y_pos',
                                        color=color_column,
                                        size=normalized_sizes[:len(active_equipment)] if size_column else [80] * len(active_equipment),
                                        hover_data=['기계코드', '가동률(%)', '총_생산수량', '가동률_안정성', '실적_상태'],
                                        color_continuous_scale='RdBu_r' if color_metric == '가동률' else 'Viridis',
                                        title=f"{process} 설비 배치 (전체 {len(process_data)}개 설비)"
                                    )
                                else:
                                    # 정상가동 설비가 없는 경우 빈 차트 생성
                                    fig_layout = go.Figure()
                                    fig_layout.update_layout(title=f"{process} 설비 배치 (전체 {len(process_data)}개 설비)")
                                
                                # 저가동 설비 추가 (노란색 삼각형으로 표시)
                                if not low_util_equipment.empty:
                                    # 저가동 설비 크기 (20 고정)
                                    low_util_size = [20] * len(low_util_equipment)
                                    fig_layout.add_trace(go.Scatter(
                                        x=low_util_equipment['x_pos'],
                                        y=low_util_equipment['y_pos'],
                                        mode='markers',
                                        marker=dict(
                                            size=low_util_size,
                                            color='orange',  # 주황색으로 표시
                                            symbol='triangle-up',  # 삼각형 모양
                                            line=dict(width=2, color='darkorange')
                                        ),
                                        name='저가동 설비',
                                        text=low_util_equipment['기계코드'],
                                        hovertemplate='<b>%{text}</b><br>' +
                                                    '실적_상태: 저가동<br>' +
                                                    '가동률: ' + low_util_equipment['가동률(%)'].astype(str) + '%<br>' +
                                                    '총_생산수량: ' + low_util_equipment['총_생산수량'].astype(str) + '<br>' +
                                                    '<extra></extra>'
                                    ))
                                
                                # 미가동 설비 추가 (다른 색상과 모양으로 표시)
                                if not inactive_equipment.empty:
                                    # 미가동 설비 크기 (20 고정)
                                    inactive_size = [20] * len(inactive_equipment)
                                    fig_layout.add_trace(go.Scatter(
                                        x=inactive_equipment['x_pos'],
                                        y=inactive_equipment['y_pos'],
                                        mode='markers',
                                        marker=dict(
                                            size=inactive_size,
                                            color='lightgray',  # 회색으로 표시
                                            symbol='x',  # X 모양으로 표시
                                            line=dict(width=2, color='red')  # 빨간색 테두리
                                        ),
                                        name='미가동 설비',
                                        text=inactive_equipment['기계코드'],
                                        hovertemplate='<b>%{text}</b><br>' +
                                                    '실적_상태: 미가동<br>' +
                                                    '가동률: 0.0%<br>' +
                                                    '총_생산수량: 0<br>' +
                                                    '<extra></extra>'
                                    ))
                                
                                # 가동률일 때 색상 범위 고정 (더 명확한 대비)
                                if color_metric == '가동률':
                                    fig_layout.update_coloraxes(cmin=0, cmax=100)
                                
                                # 설비명 라벨 추가 (설비 상태별 구분 표시)
                                if show_labels:
                                    for _, row in process_data.iterrows():
                                        # 설비 상태에 따라 색상 및 텍스트 구분
                                        status = row['실적_상태']
                                        if status == '미가동':
                                            label_color = 'red'
                                            rate_color = 'red'
                                            status_text = '[미가동]'
                                            border_width = 2
                                        elif status == '저가동':
                                            label_color = 'darkorange'
                                            rate_color = 'darkorange'  
                                            status_text = '[저가동]'
                                            border_width = 2
                                        else:  # 정상가동
                                            label_color = 'black'
                                            rate_color = 'blue'
                                            status_text = ''
                                            border_width = 1
                                        
                                        # 설비명을 설비 위쪽에 (상태별 색상)
                                        fig_layout.add_annotation(
                                            x=row['x_pos'], y=row['y_pos'] + 0.3,
                                            text=f"{row['기계코드']} {status_text}",
                                            showarrow=False,
                                            font=dict(size=9, color=label_color, family='Arial Black'),
                                            bgcolor='rgba(255,255,255,0.8)',
                                            bordercolor=label_color,
                                            borderwidth=border_width
                                        )
                                        # 가동률을 설비 아래쪽에 (상태별 강조)
                                        fig_layout.add_annotation(
                                            x=row['x_pos'], y=row['y_pos'] - 0.3,
                                            text=f"{row['가동률(%)']:.1f}%",
                                            showarrow=False,
                                            font=dict(size=10, color=rate_color, family='Arial Black'),
                                            bgcolor='rgba(255,255,255,0.9)',
                                            bordercolor=rate_color,
                                            borderwidth=border_width
                                        )
                                
                                # 레이아웃 조정
                                fig_layout.update_layout(
                                    height=400,
                                    xaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
                                    yaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig_layout, use_container_width=True)
                                
                                # 공정별 설비 현황 정보
                                process_total = len(process_data)
                                process_active = len(process_data[process_data['실적_상태'] == '정상가동'])
                                process_low_util = len(process_data[process_data['실적_상태'] == '저가동'])
                                process_inactive = len(process_data[process_data['실적_상태'] == '미가동'])
                                avg_util = process_data['가동률(%)'].mean()
                                total_prod = process_data['총_생산수량'].sum()
                                
                                # 두 줄로 나누어 표시
                                st.caption(f"📊 {process} 설비 현황: 전체 {process_total}개 | 정상가동 {process_active}개 | 저가동 {process_low_util}개 | 미가동 {process_inactive}개")
                                st.caption(f"📈 평균 가동률: {avg_util:.1f}% | 총 생산량: {total_prod:,.0f}")
            
            elif layout_view == "공정별 배치도":
                # 공정별로 모든 공장의 설비 표시
                processes = equipment_layout_data['공정코드'].dropna().unique()
                for process in sorted(processes):
                    process_data = equipment_layout_data[equipment_layout_data['공정코드'] == process]
                    
                    st.subheader(f"⚙️ {process} 전체 설비 배치")
                    
                    # 색상 및 크기 매핑
                    color_column = {'가동률': '가동률(%)', '생산량': '총_생산수량', '안정성': '가동률_안정성'}[color_metric]
                    size_column = {'생산량': '총_생산수량', '가동률': '가동률(%)', '균등': None}[size_metric]
                    
                    # 공장별로 x축 위치 분리 (겹침 방지)
                    factories = process_data['공장'].dropna().unique()
                    factory_positions = {factory: i for i, factory in enumerate(sorted(factories))}
                    spacing = 2.0  # 설비간 간격 증가
                    
                    x_positions = []
                    y_positions = []
                    
                    for factory in sorted(factories):
                        factory_data = process_data[process_data['공장'] == factory]
                        base_x = factory_positions[factory] * 10  # 공장간 간격 증가
                        
                        for idx, (_, equipment) in enumerate(factory_data.iterrows()):
                            x_pos = base_x + (idx % 4) * spacing  # 간격 적용
                            y_pos = (idx // 4) * spacing
                            x_positions.append(x_pos)
                            y_positions.append(y_pos)
                    
                    process_data = process_data.copy()
                    process_data['x_pos'] = x_positions
                    process_data['y_pos'] = y_positions
                    
                    # 크기 설정 (더 큰 크기로 가시성 향상)
                    if size_column:
                        sizes = process_data[size_column]
                        size_min, size_max = sizes.min(), sizes.max()
                        if size_max > size_min:
                            normalized_sizes = 25 + 100 * (sizes - size_min) / (size_max - size_min)
                        else:
                            normalized_sizes = [75] * len(sizes)
                    else:
                        normalized_sizes = [75] * len(process_data)
                    
                    # 배치도 차트 생성 (미가동 설비 구분)
                    # 정상가동 설비와 미가동 설비 분리
                    active_equipment = process_data[process_data['실적_상태'] == '정상가동']
                    inactive_equipment = process_data[process_data['실적_상태'] == '미가동']
                    
                    # 기본 차트 생성 (정상가동 설비)
                    if not active_equipment.empty:
                        fig_process = px.scatter(
                            active_equipment,
                            x='x_pos', y='y_pos',
                            color=color_column,
                            size=normalized_sizes[:len(active_equipment)] if size_column else [75] * len(active_equipment),
                            hover_data=['기계코드', '공장', '가동률(%)', '총_생산수량', '실적_상태'],
                            color_continuous_scale='RdBu_r' if color_metric == '가동률' else 'Viridis',
                            title=f"{process} 전체 설비 배치도 (전체 {len(process_data)}개 설비)"
                        )
                    else:
                        # 정상가동 설비가 없는 경우 빈 차트 생성
                        fig_process = go.Figure()
                        fig_process.update_layout(title=f"{process} 전체 설비 배치도 (전체 {len(process_data)}개 설비)")
                    
                    # 미가동 설비 추가 (다른 색상과 모양으로 표시)
                    if not inactive_equipment.empty:
                        # 미가동 설비 크기 (20 고정)
                        inactive_size = [20] * len(inactive_equipment)
                        fig_process.add_trace(go.Scatter(
                            x=inactive_equipment['x_pos'],
                            y=inactive_equipment['y_pos'],
                            mode='markers',
                            marker=dict(
                                size=inactive_size,
                                color='lightgray',  # 회색으로 표시
                                symbol='x',  # X 모양으로 표시
                                line=dict(width=2, color='red')  # 빨간색 테두리
                            ),
                            name='미가동 설비',
                            text=inactive_equipment['기계코드'],
                            hovertemplate='<b>%{text}</b><br>' +
                                        '공장: ' + inactive_equipment['공장'] + '<br>' +
                                        '실적_상태: 미가동<br>' +
                                        '가동률: 0.0%<br>' +
                                        '총_생산수량: 0<br>' +
                                        '<extra></extra>'
                        ))
                    
                    # 가동률일 때 색상 범위 고정
                    if color_metric == '가동률':
                        fig_process.update_coloraxes(cmin=0, cmax=100)
                    
                    # 설비명 및 가동률 라벨 추가 (미가동 설비 구분)
                    if show_labels:
                        for _, row in process_data.iterrows():
                            # 미가동 설비 여부에 따라 색상 구분
                            is_inactive = row['실적_상태'] == '미가동'
                            label_color = 'red' if is_inactive else 'black'
                            status_text = '[미가동]' if is_inactive else ''
                            
                            # 설비명을 설비 위쪽에 (미가동 시 빨간색)
                            fig_process.add_annotation(
                                x=row['x_pos'], y=row['y_pos'] + 0.3,
                                text=f"{row['기계코드']} {status_text}",
                                showarrow=False,
                                font=dict(size=8, color=label_color, family='Arial Black'),
                                bgcolor='rgba(255,255,255,0.8)',
                                bordercolor=label_color,
                                borderwidth=1
                            )
                            # 가동률을 설비 아래쪽에 (미가동 시 더 강조)
                            rate_color = 'red' if is_inactive else 'blue'
                            fig_process.add_annotation(
                                x=row['x_pos'], y=row['y_pos'] - 0.3,
                                text=f"{row['가동률(%)']:.1f}%",
                                showarrow=False,
                                font=dict(size=9, color=rate_color, family='Arial Black'),
                                bgcolor='rgba(255,255,255,0.9)',
                                bordercolor=rate_color,
                                borderwidth=2 if is_inactive else 1
                            )
                    
                    # 공장 구분선 추가
                    process_factories = process_data['공장'].dropna().unique()
                    for i, factory in enumerate(sorted(process_factories)):
                        x_line = i * 5 - 0.5
                        fig_process.add_vline(x=x_line, line_dash="dash", line_color="gray", opacity=0.5)
                        fig_process.add_annotation(x=i*5 + 1.5, y=-0.5, text=factory, showarrow=False, font=dict(size=12, color='blue'))
                    
                    fig_process.update_layout(
                        height=500,
                        xaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
                        yaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
                        plot_bgcolor='white'
                    )
                    
                    st.plotly_chart(fig_process, use_container_width=True)
                    
                    # 공정별 설비 현황 요약
                    process_total = len(process_data)
                    process_active = len(process_data[process_data['실적_상태'] == '정상가동'])
                    process_low_util = len(process_data[process_data['실적_상태'] == '저가동'])
                    process_inactive = len(process_data[process_data['실적_상태'] == '미가동'])
                    avg_util = process_data['가동률(%)'].mean()
                    total_prod = process_data['총_생산수량'].sum()
                    
                    # 두 줄로 나누어 표시
                    st.caption(f"📊 {process} 설비 현황: 전체 {process_total}개 | 정상가동 {process_active}개 | 저가동 {process_low_util}개 | 미가동 {process_inactive}개")
                    st.caption(f"📈 평균 가동률: {avg_util:.1f}% | 총 생산량: {total_prod:,.0f}")
            
            else:  # 통합 배치도
                st.subheader("🌐 전체 설비 통합 배치도")
                
                # 색상 및 크기 매핑
                color_column = {'가동률': '가동률(%)', '생산량': '총_생산수량', '안정성': '가동률_안정성'}[color_metric]
                size_column = {'생산량': '총_생산수량', '가동률': '가동률(%)', '균등': None}[size_metric]
                
                # 계층적 배치 (공장 > 공정 > 설비) - 겹침 방지
                x_positions = []
                y_positions = []
                spacing = 1.8  # 설비간 간격
                
                factory_offset = 0
                layout_factories = equipment_layout_data['공장'].dropna().unique()
                for factory in sorted(layout_factories):
                    factory_data = equipment_layout_data[equipment_layout_data['공장'] == factory]
                    
                    process_offset = 0
                    factory_processes = factory_data['공정코드'].dropna().unique()
                    for process in sorted(factory_processes):
                        process_data = factory_data[factory_data['공정코드'] == process]
                        
                        for idx, (_, equipment) in enumerate(process_data.iterrows()):
                            x_pos = factory_offset * 25 + process_offset * 6 + (idx % 3) * spacing
                            y_pos = (idx // 3) * spacing
                            x_positions.append(x_pos)
                            y_positions.append(y_pos)
                        
                        process_offset += 1
                    factory_offset += 1
                
                equipment_layout_data = equipment_layout_data.copy()
                equipment_layout_data['x_pos'] = x_positions
                equipment_layout_data['y_pos'] = y_positions
                
                # 크기 설정 (통합 배치도에서도 충분한 크기)
                if size_column:
                    sizes = equipment_layout_data[size_column]
                    size_min, size_max = sizes.min(), sizes.max()
                    if size_max > size_min:
                        normalized_sizes = 20 + 60 * (sizes - size_min) / (size_max - size_min)
                    else:
                        normalized_sizes = [50] * len(sizes)
                else:
                    normalized_sizes = [50] * len(equipment_layout_data)
                
                # 통합 배치도 생성 (미가동 설비 구분)
                # 정상가동 설비와 미가동 설비 분리
                active_equipment = equipment_layout_data[equipment_layout_data['실적_상태'] == '정상가동']
                inactive_equipment = equipment_layout_data[equipment_layout_data['실적_상태'] == '미가동']
                
                # 기본 차트 생성 (정상가동 설비)
                if not active_equipment.empty:
                    fig_integrated = px.scatter(
                        active_equipment,
                        x='x_pos', y='y_pos',
                        color=color_column,
                        size=normalized_sizes[:len(active_equipment)] if size_column else [50] * len(active_equipment),
                        hover_data=['기계코드', '공장', '공정코드', '가동률(%)', '총_생산수량', '실적_상태'],
                        color_continuous_scale='RdBu_r' if color_metric == '가동률' else 'Viridis',
                        title=f"전체 설비 통합 배치도 (전체 {len(equipment_layout_data)}개 설비)"
                    )
                else:
                    # 정상가동 설비가 없는 경우 빈 차트 생성
                    fig_integrated = go.Figure()
                    fig_integrated.update_layout(title=f"전체 설비 통합 배치도 (전체 {len(equipment_layout_data)}개 설비)")
                
                # 미가동 설비 추가 (다른 색상과 모양으로 표시)
                if not inactive_equipment.empty:
                    # 미가동 설비 크기 (20 고정)
                    inactive_size = [20] * len(inactive_equipment)
                    fig_integrated.add_trace(go.Scatter(
                        x=inactive_equipment['x_pos'],
                        y=inactive_equipment['y_pos'],
                        mode='markers',
                        marker=dict(
                            size=inactive_size,
                            color='lightgray',  # 회색으로 표시
                            symbol='x',  # X 모양으로 표시
                            line=dict(width=2, color='red')  # 빨간색 테두리
                        ),
                        name='미가동 설비',
                        text=inactive_equipment['기계코드'],
                        hovertemplate='<b>%{text}</b><br>' +
                                    '공장: ' + inactive_equipment['공장'] + '<br>' +
                                    '공정코드: ' + inactive_equipment['공정코드'] + '<br>' +
                                    '실적_상태: 미가동<br>' +
                                    '가동률: 0.0%<br>' +
                                    '총_생산수량: 0<br>' +
                                    '<extra></extra>'
                    ))
                
                # 가동률일 때 색상 범위 고정
                if color_metric == '가동률':
                    fig_integrated.update_coloraxes(cmin=0, cmax=100)
                
                # 설비명 및 가동률 라벨 추가 (미가동 설비 구분, 선택적)
                if show_labels and len(equipment_layout_data) <= 50:  # 너무 많으면 라벨이 겹침
                    for _, row in equipment_layout_data.iterrows():
                        # 미가동 설비 여부에 따라 색상 구분
                        is_inactive = row['실적_상태'] == '미가동'
                        label_color = 'red' if is_inactive else 'black'
                        status_text = '[미가동]' if is_inactive else ''
                        
                        # 설비명을 설비 위쪽에 (미가동 시 빨간색)
                        fig_integrated.add_annotation(
                            x=row['x_pos'], y=row['y_pos'] + 0.2,
                            text=f"{row['기계코드']} {status_text}",
                            showarrow=False,
                            font=dict(size=7, color=label_color, family='Arial Black'),
                            bgcolor='rgba(255,255,255,0.8)',
                            bordercolor=label_color,
                            borderwidth=1
                        )
                        # 가동률을 설비 아래쪽에 (미가동 시 더 강조)
                        rate_color = 'red' if is_inactive else 'blue'
                        fig_integrated.add_annotation(
                            x=row['x_pos'], y=row['y_pos'] - 0.2,
                            text=f"{row['가동률(%)']:.1f}%",
                            showarrow=False,
                            font=dict(size=8, color=rate_color, family='Arial Black'),
                            bgcolor='rgba(255,255,255,0.9)',
                            bordercolor=rate_color,
                            borderwidth=2 if is_inactive else 1
                        )
                
                fig_integrated.update_layout(
                    height=600,
                    xaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
                    yaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
                    plot_bgcolor='white'
                )
                
                st.plotly_chart(fig_integrated, use_container_width=True)
                
                # 전체 요약 통계 (3단계 설비 현황 포함)
                col1, col2, col3, col4, col5 = st.columns(5)
                
                total_equipment = len(equipment_layout_data)
                active_equipment_count = len(equipment_layout_data[equipment_layout_data['실적_상태'] == '정상가동'])
                low_util_equipment_count = len(equipment_layout_data[equipment_layout_data['실적_상태'] == '저가동'])
                inactive_equipment_count = len(equipment_layout_data[equipment_layout_data['실적_상태'] == '미가동'])
                
                with col1:
                    st.metric("전체 설비 수", f"{total_equipment}대")
                with col2:
                    st.metric("정상가동 설비", f"{active_equipment_count}대", 
                             f"{(active_equipment_count/total_equipment)*100:.1f}%" if total_equipment > 0 else "0%")
                with col3:
                    st.metric("저가동 설비", f"{low_util_equipment_count}대", 
                             f"{(low_util_equipment_count/total_equipment)*100:.1f}%" if total_equipment > 0 else "0%")
                with col4:
                    st.metric("미가동 설비", f"{inactive_equipment_count}대", 
                             f"{(inactive_equipment_count/total_equipment)*100:.1f}%" if total_equipment > 0 else "0%")
                with col5:
                    avg_utilization = equipment_layout_data['가동률(%)'].mean()
                    st.metric("평균 가동률", f"{avg_utilization:.1f}%")
                
                # 추가 정보
                st.caption(f"📍 선택 기간: {start_date} ~ {end_date}")
                st.caption(f"🟢 정상가동(30%초과): 색상별 구분 | 🟠 저가동(1~30%): 주황 삼각형 | 🔴 미가동(0%): 회색 X표시")
                if low_util_equipment_count > 0 or inactive_equipment_count > 0:
                    problem_count = low_util_equipment_count + inactive_equipment_count
                    st.warning(f"⚠️ {problem_count}개 설비에 주의가 필요합니다. (저가동 {low_util_equipment_count}개 | 미가동 {inactive_equipment_count}개) 생산 계획 수립 시 참고하세요.")
        
        with analysis_tabs[3]:  # 비교 분석
            st.subheader("📊 다면 비교 분석")
            
            comparison_type = st.radio("비교 유형 선택", 
                                     options=['공장간 비교', '공정간 비교', '기간별 비교'], 
                                     horizontal=True, key='comparison_type')
            
            if comparison_type == '공장간 비교':
                # 공장별 가동률 박스플롯
                fig_factory_box = px.box(df_filtered, x='공장', y='가동률(%)', 
                                       title='<b>공장별 가동률 분포 비교</b>')
                fig_factory_box.update_layout(height=400)
                st.plotly_chart(fig_factory_box, use_container_width=True)
                
                # 공장별 통계 요약
                factory_stats = df_filtered.groupby('공장')['가동률(%)'].describe().round(2)
                st.subheader("공장별 통계 요약")
                st.dataframe(factory_stats, use_container_width=True)
                
            elif comparison_type == '공정간 비교':
                # 공정별 가동률 바이올린 플롯
                fig_process_violin = px.violin(df_filtered, x='공정코드', y='가동률(%)',
                                             title='<b>공정별 가동률 분포 비교</b>')
                fig_process_violin.update_layout(height=400)
                st.plotly_chart(fig_process_violin, use_container_width=True)
                
                # 공정별 설비 수와 평균 가동률
                process_summary = df_filtered.groupby('공정코드').agg({
                    '가동률(%)': ['mean', 'std', 'count'],
                    '기계코드': 'nunique'
                }).round(2)
                process_summary.columns = ['평균_가동률', '표준편차', '총_데이터수', '설비수']
                process_summary = process_summary.reset_index()
                st.subheader("공정별 요약")
                st.dataframe(process_summary, use_container_width=True)
            
            else:  # 기간별 비교
                if 'date' in df_filtered.columns:
                    # 날짜별 평균 가동률 추이
                    daily_util = df_filtered.groupby('date')['가동률(%)'].mean().reset_index()
                    fig_daily = px.line(daily_util, x='date', y='가동률(%)',
                                      title='<b>일별 평균 가동률 추이</b>')
                    fig_daily.update_layout(height=400)
                    st.plotly_chart(fig_daily, use_container_width=True)
                else:
                    st.info("날짜 정보가 없어 기간별 비교를 할 수 없습니다.")
        
        with analysis_tabs[4]:  # 성과 분석
            st.subheader("🎯 설비 성과 순위")
            
            # 성과 지표 선택
            performance_metric = st.selectbox("성과 지표", 
                                            options=['평균 가동률', '총 생산량', '가동률 안정성'],
                                            key='performance_metric')
            
            # 설비별 성과 계산
            equipment_performance = df_filtered.groupby(['기계코드', '공장', '공정코드']).agg({
                '가동률(%)': ['mean', 'std'],
                '총_생산수량': 'sum'
            }).round(2)
            
            equipment_performance.columns = ['평균_가동률', '가동률_표준편차', '총_생산량']
            # 가동률 안정성 계산 (음수 방지 및 NaN 처리)
            equipment_performance['가동률_안정성'] = (100 - equipment_performance['가동률_표준편차']).clip(lower=0)
            equipment_performance = equipment_performance.fillna(0)  # NaN 값을 0으로 대체
            equipment_performance = equipment_performance.reset_index()
            
            # 선택한 지표에 따라 정렬
            sort_column_map = {
                '평균 가동률': '평균_가동률',
                '총 생산량': '총_생산량', 
                '가동률 안정성': '가동률_안정성'
            }
            sort_column = sort_column_map[performance_metric]
            equipment_performance = equipment_performance.sort_values(sort_column, ascending=False)
            
            # 전체 순위 표시 옵션
            display_option = st.radio("표시 방식", 
                                    options=["전체 순위", "TOP/BOTTOM 10"], 
                                    horizontal=True, key='display_option')
            
            if display_option == "전체 순위":
                st.subheader(f"📊 전체 설비 순위 ({performance_metric})")
                st.caption(f"총 {len(equipment_performance)}개 설비 - 높은 순서대로 정렬")
                st.dataframe(equipment_performance, use_container_width=True, height=600)
            else:
                # TOP 10과 BOTTOM 10 표시
                col1, col2 = st.columns(2)
                
                with col1:
                    st.subheader(f"🏆 TOP 10 ({performance_metric})")
                    top_10 = equipment_performance.head(10)
                    st.dataframe(top_10, use_container_width=True, height=400)
                
                with col2:
                    st.subheader(f"🔻 BOTTOM 10 ({performance_metric})")
                    bottom_10 = equipment_performance.tail(10)
                    st.dataframe(bottom_10, use_container_width=True, height=400)
            
            # 성과 분포 시각화
            # size 값이 유효한지 확인하고 최소값 보장
            equipment_performance['size_value'] = equipment_performance['가동률_안정성'].clip(lower=1)
            
            fig_performance = px.scatter(equipment_performance, 
                                       x='평균_가동률', y='총_생산량',
                                       color='공장', size='size_value',
                                       hover_data=['기계코드', '공정코드', '가동률_안정성'],
                                       title='<b>설비별 성과 매트릭스 (가동률 vs 생산량)</b>')
            fig_performance.update_layout(height=500)
            st.plotly_chart(fig_performance, use_container_width=True)
        
        # 다운로드 섹션 추가
        create_download_section(df_filtered, "가동률분석", agg_level, start_date, end_date)

elif selected_tab == "종합 분석":
    filter_source_df = df_target_orig if not df_target_orig.empty else df_yield_orig
    df_filtered, start_date, end_date, agg_level = create_shared_filter_controls(filter_source_df)
    if df_yield_orig.empty:
        st.info("종합 분석에 필요한 수율 데이터가 없습니다.")
    elif df_filtered.empty:
        st.info("선택된 기간에 분석할 데이터가 부족합니다.")
    else:
        mask_yield = (df_yield_orig['date'].dt.date >= start_date) & (df_yield_orig['date'].dt.date <= end_date)
        df_yield_filt = df_yield_orig[mask_yield].copy()

        # 데이터 처리
        compare_factories = st.session_state.get('compare_factories', False)
        selected_factory = st.session_state.get('overall_factory_select', '전체')
        
        if compare_factories:
            df_yield_filt_factory = df_yield_filt.copy()
            active_factory = '전체'
        else:
            df_yield_filt_factory = df_yield_filt[df_yield_filt['공장'] == selected_factory].copy() if selected_factory != '전체' else df_yield_filt.copy()
            active_factory = selected_factory

        bar_data, line_data = pd.DataFrame(), pd.DataFrame()
        if not df_yield_filt_factory.empty:
            group_by_cols = ['period', '공장', '공정코드'] if compare_factories else ['period', '공정코드']
            df_yield_resampled = get_resampled_data(df_yield_filt_factory, agg_level, ['총_생산수량', '총_양품수량'], group_by_cols=group_by_cols)
            df_final_yield_filtered = df_yield_resampled[df_yield_resampled['공정코드'] == '[80] 누수/규격검사']
            bar_group_cols = ['period', '공장'] if compare_factories else ['period']
            bar_data = df_final_yield_filtered.groupby(bar_group_cols)['총_양품수량'].sum().reset_index().rename(columns={'총_양품수량': '총_생산수량'})
            with pd.option_context('mode.use_inf_as_na', True): df_yield_resampled['개별공정수율'] = (df_yield_resampled['총_양품수량'] / df_yield_resampled['총_생산수량']).fillna(1.0)
            line_group_cols = ['period', '공장'] if compare_factories else ['period']
            line_data = df_yield_resampled.groupby(line_group_cols)['개별공정수율'].prod().reset_index(name='종합수율(%)')
            line_data['종합수율(%)'] *= 100
        else:
            bar_data = pd.DataFrame(columns=['period', '총_생산수량'])
            line_data = pd.DataFrame(columns=['period', '종합수율(%)'])

        if bar_data.empty or line_data.empty: st.info("선택된 기간에 분석할 데이터가 부족합니다.")
        else:
            merge_cols = ['period', '공장'] if compare_factories else ['period']
            combo_data = pd.merge(bar_data, line_data, on=merge_cols, how='outer').sort_values('period').fillna(0)
            combo_data.loc[combo_data['총_생산수량'] == 0, '종합수율(%)'] = 0
            
            st.markdown("---"); st.subheader("차트 옵션 조정", anchor=False)
            
            # 모든 컨트롤을 브리핑 위로 이동
            control_cols_1 = st.columns(3)
            with control_cols_1[0]:
                all_factories = ['전체'] + sorted(df_yield_orig['공장'].unique())
                st.selectbox(
                    "공장 선택", options=all_factories, key="overall_factory_select",
                    disabled=st.session_state.get('compare_factories', False)
                )
            with control_cols_1[1]:
                st.markdown("<div style='padding-top: 28px;'></div>", unsafe_allow_html=True)
                st.checkbox("공장별 함께보기", key="compare_factories")

            control_cols_2 = st.columns(4)
            with control_cols_2[0]: 
                min_yield_val = combo_data['종합수율(%)'].min() if not combo_data.empty else 0
                max_yield_val = combo_data['종합수율(%)'].max() if not combo_data.empty else 100
                buffer = (max_yield_val - min_yield_val) * 0.5 if max_yield_val > min_yield_val else 5.0
                slider_min = max(0.0, min_yield_val - buffer)
                slider_max = min(100.0, max_yield_val + buffer)
                yield_range = st.slider("종합 수율(%) 축 범위", 0.0, 100.0, (slider_min, slider_max), 1.0, format="%.0f%%", key="overall_yield_range")
            with control_cols_2[1]: chart_height = st.slider("차트 높이 조절", 400, 1000, 700, 50, key="overall_chart_height")
            with control_cols_2[2]: show_labels = st.toggle("차트 라벨 표시", value=True, key="overall_show_labels")
            with control_cols_2[3]: 
                # 그래프 설정 옵션
                with st.expander("📊 차트 설정", expanded=False):
                    col_set1, col_set2, col_set3 = st.columns(3)
                    with col_set1:
                        comprehensive_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="comprehensive_label_size")
                    with col_set2:
                        comprehensive_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="comprehensive_axis_title_size")
                    with col_set3:
                        comprehensive_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="comprehensive_axis_tick_size")
            
            # 고차원적 AI 브리핑
            st.subheader("🤖 AI Analyst 종합 분석 브리핑", anchor=False)
            if not combo_data.empty:
                start_period = combo_data['period'].iloc[0]
                end_period = combo_data['period'].iloc[-1] 
                total_prod = combo_data['총_생산수량'].sum()
                avg_yield = combo_data['종합수율(%)'].mean()
                max_yield_row = combo_data.loc[combo_data['종합수율(%)'].idxmax()]
                min_yield_row = combo_data.loc[combo_data['종합수율(%)'].idxmin()]
                
                # 추세 분석
                if len(combo_data) >= 3:
                    recent_3_yield = combo_data['종합수율(%)'].tail(3).mean()
                    early_3_yield = combo_data['종합수율(%)'].head(3).mean()
                    yield_trend = "상승세" if recent_3_yield > early_3_yield + 1 else "하락세" if recent_3_yield < early_3_yield - 1 else "안정세"
                    
                    recent_3_prod = combo_data['총_생산수량'].tail(3).mean()
                    early_3_prod = combo_data['총_생산수량'].head(3).mean()
                    prod_trend = "증가세" if recent_3_prod > early_3_prod * 1.05 else "감소세" if recent_3_prod < early_3_prod * 0.95 else "안정세"
                else:
                    yield_trend = "안정세"
                    prod_trend = "안정세"
                
                # 성과 평가
                performance_level = "우수" if avg_yield >= 90 else "양호" if avg_yield >= 80 else "개선필요"
                
                # 변동성 분석
                yield_std = combo_data['종합수율(%)'].std()
                stability = "안정적" if yield_std <= 3 else "변동성 높음"
                
                st.info(f"""
                **📊 분석 기간:** {start_period} ~ {end_period} | **성과 평가:** {performance_level}
                
                **🎯 핵심 지표**
                • 총 생산량: {total_prod:,.0f}개 ({prod_trend})
                • 평균 종합수율: {avg_yield:.1f}% ({yield_trend}, {stability})
                • 수율 변동폭: {max_yield_row['종합수율(%)']:.1f}% ~ {min_yield_row['종합수율(%)']:.1f}%
                
                **📈 트렌드 분석**
                • 수율 추세: **{yield_trend}** (최근 구간 vs 초기 구간)
                • 생산량 추세: **{prod_trend}** 
                • 최고 성과: {max_yield_row['period']} ({max_yield_row['종합수율(%)']:.1f}%)
                
                **🎯 개선 포인트**
                • {"수율 향상에 집중하여 90% 이상 달성 목표" if avg_yield < 90 else "현재 우수한 수율 수준 유지"}
                • {"생산량 안정화를 통한 품질 일관성 확보" if yield_std > 3 else "안정적인 품질 관리 상태 지속"}
                • 최저 성과 기간({min_yield_row['period']})의 원인 분석 및 재발 방지
                """)
            else:
                st.info("분석할 데이터가 부족합니다.")
            
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            chart_title_prefix = f"{active_factory} " if active_factory != '전체' else ""
            
            if compare_factories:
                for factory_name in sorted(combo_data['공장'].unique()):
                    df_factory = combo_data[combo_data['공장'] == factory_name]
                    
                    # 막대 차트와 꺾은선 그래프의 색상을 별도로 관리
                    bar_color = 'gray'  # 기본값
                    line_color = 'gray' # 기본값
                    for key, color in FACTORY_COLOR_MAP.items():
                        if key in factory_name:
                            bar_color = color
                            line_color = color
                            break
                    
                    # C관의 꺾은선 그래프만 검은색으로 지정
                    if 'C관' in factory_name:
                        line_color = 'black'
                    
                    fig.add_trace(go.Bar(
                        x=df_factory['period'], y=df_factory['총_생산수량'], name=f'{factory_name} 완제품', 
                        legendgroup=factory_name, marker_color=bar_color,
                        text=df_factory['총_생산수량'], texttemplate='<b>%{text:,.0f}</b>',
                        textposition='outside' if show_labels else 'none',
                        textfont=dict(size=comprehensive_label_size, color='black')
                    ), secondary_y=False)

                    # 동적 라벨 위치 설정을 위한 로직
                    df_factory = df_factory.sort_values('period').reset_index(drop=True)
                    text_positions = []
                    for i, row in df_factory.iterrows():
                        # 이전/이후 데이터 포인트와의 관계를 기반으로 위치 결정
                        y_current = row['종합수율(%)']
                        y_prev = df_factory.loc[i - 1, '종합수율(%)'] if i > 0 else -1
                        y_next = df_factory.loc[i + 1, '종합수율(%)'] if i < len(df_factory) - 1 else -1

                        if y_current >= y_prev and y_current >= y_next:
                            position = 'top center'
                        elif y_current < y_prev and y_current < y_next:
                            position = 'bottom center'
                        elif y_current < y_prev:
                            position = 'top center'
                        else:
                            position = 'bottom center'
                        text_positions.append(position)
                    
                    df_factory['text_position'] = text_positions

                    fig.add_trace(go.Scatter(
                        x=df_factory['period'], y=df_factory['종합수율(%)'], name=f'{factory_name} 수율', 
                        legendgroup=factory_name, line=dict(color=line_color), 
                        mode='lines+markers+text' if show_labels else 'lines+markers',
                        text=df_factory['종합수율(%)'], texttemplate='<b>%{text:.2f}%</b>',
                        textposition=df_factory['text_position'],
                        textfont=dict(color='black', size=comprehensive_label_size)
                    ), secondary_y=True)
                fig.update_layout(barmode='group')
            else:
                blue_scale = ['#aed6f1', '#85c1e9', '#5dade2', '#3498db', '#2e86c1', '#2874a6', '#21618c', '#1b4f72', '#153d5a', '#102e48', '#0b1e34', '#071323']
                bar_colors = [blue_scale[i % len(blue_scale)] for i in range(len(combo_data))]
                fig.add_trace(go.Bar(x=combo_data['period'], y=combo_data['총_생산수량'], name='완제품 제조 개수', text=combo_data['총_생산수량'], texttemplate='<b>%{text:,.0f}</b>', textposition='outside' if show_labels else 'none', textfont=dict(size=comprehensive_label_size), marker_color=bar_colors), secondary_y=False)
                fig.add_trace(go.Scatter(x=combo_data['period'], y=combo_data['종합수율(%)'], name=f'{agg_level} 종합 수율', mode='lines+markers+text' if show_labels else 'lines+markers', line=dict(color='crimson', width=3), marker=dict(color='crimson', size=8), text=combo_data['종합수율(%)'], texttemplate='<b>%{text:.2f}%</b>', textposition='top center', textfont=dict(color='black', size=comprehensive_label_size, family="Arial, sans-serif")), secondary_y=True)

            max_bar_val = combo_data['총_생산수량'].max() if not combo_data.empty else 0

            fig.update_layout(height=chart_height, title_text=f'<b>{chart_title_prefix}{agg_level} 완제품 제조 실적 및 종합 수율</b>', title_font_size=comprehensive_axis_title_size, margin=dict(t=120), legend=dict(orientation="h", yanchor="bottom", y=1.10, xanchor="right", x=1, font_size=16))
            fig.update_yaxes(title_text="<b>완제품 제조 개수</b>", secondary_y=False, title_font_size=comprehensive_axis_title_size, tickfont_size=comprehensive_axis_tick_size, range=[0, max_bar_val * 1.15])
            fig.update_yaxes(title_text="<b>종합 수율 (%)</b>", secondary_y=True, title_font_size=comprehensive_axis_title_size, tickfont_size=comprehensive_axis_tick_size, range=yield_range)
            fig.update_xaxes(title_text=f"<b>{agg_level.replace('별', '')}</b>", type='category', categoryorder='array', categoryarray=sorted(combo_data['period'].unique()), title_font_size=comprehensive_axis_title_size, tickfont_size=comprehensive_axis_tick_size)
            st.plotly_chart(fig, use_container_width=True)


            # --- 제품군별 종합 실적 분석 ---
            st.divider()
            st.subheader(f"{agg_level} 제품군별 완제품 제조 실적 및 종합 수율", anchor=False)

            # 공장 선택 필터
            pg_all_factories = ['전체'] + sorted(df_yield_orig['공장'].unique())
            pg_selected_factory = st.selectbox(
                "분석 공장 선택", 
                options=pg_all_factories, 
                key="pg_factory_select",
                help="제품군별 분석을 수행할 공장을 선택합니다. '전체' 선택 시 모든 공장의 데이터를 종합하여 분석합니다."
            )

            # 선택된 공장에 따라 데이터 필터링
            if pg_selected_factory == '전체':
                df_yield_pg_filtered = df_yield_filt.copy()
            else:
                df_yield_pg_filtered = df_yield_filt[df_yield_filt['공장'] == pg_selected_factory].copy()
            
            if '신규분류요약' in df_yield_pg_filtered.columns:
                all_product_groups_pg = sorted(df_yield_pg_filtered['신규분류요약'].dropna().unique())

                if not all_product_groups_pg:
                    st.warning("선택된 공장에 제품군 데이터가 없습니다.")
                else:
                    for group in all_product_groups_pg:
                        if f"pg_product_group_{group}" not in st.session_state: st.session_state[f"pg_product_group_{group}"] = True
                    
                    st.markdown("##### 표시할 제품군 선택")
                    btn_cols_pg = st.columns(8)
                    with btn_cols_pg[0]:
                        if st.button("제품군 전체 선택", key="pg_select_all", use_container_width=True):
                            for group in all_product_groups_pg: st.session_state[f"pg_product_group_{group}"] = True
                            st.rerun()
                    with btn_cols_pg[1]:
                        if st.button("제품군 전체 해제", key="pg_deselect_all", use_container_width=True):
                            for group in all_product_groups_pg: st.session_state[f"pg_product_group_{group}"] = False
                            st.rerun()
                    
                    st.write("")
                    num_cols_pg = 5
                    cols_pg = st.columns(num_cols_pg)
                    selected_product_groups_pg = []
                    for i, group in enumerate(all_product_groups_pg):
                        with cols_pg[i % num_cols_pg]:
                            if st.checkbox(group, key=f"pg_product_group_{group}"):
                                selected_product_groups_pg.append(group)
                    
                    combine_pg = st.checkbox("선택항목 합쳐서 보기", key="pg_combine_yield", help="선택한 제품군들의 실적을 합산하여 단일 종합 수율 및 생산 실적 추이를 분석합니다.")

                    if selected_product_groups_pg:
                        df_resampled_pg = get_resampled_data(df_yield_pg_filtered, agg_level, ['총_생산수량', '총_양품수량'], group_by_cols=['period', '신규분류요약', '공정코드'])
                        df_resampled_pg_filtered = df_resampled_pg[df_resampled_pg['신규분류요약'].isin(selected_product_groups_pg)]

                        if not df_resampled_pg_filtered.empty:
                            df_to_plot_pg = pd.DataFrame()
                            
                            last_process = PROCESS_MASTER_ORDER[-1]

                            if combine_pg:
                                # 1. 실적 데이터 (최종 공정 기준)
                                bar_combined = df_resampled_pg_filtered[df_resampled_pg_filtered['공정코드'] == last_process].groupby('period')['총_양품수량'].sum().reset_index().rename(columns={'총_양품수량': '완제품_제조개수'})
                                
                                # 2. 수율 데이터 (개별 공정 수율의 곱)
                                df_yield_combined_base = df_resampled_pg_filtered.groupby(['period', '공정코드']).agg(총_생산수량=('총_생산수량', 'sum'), 총_양품수량=('총_양품수량', 'sum')).reset_index()
                                df_yield_combined_base['총_생산수량'] = df_yield_combined_base['총_생산수량'].replace(0, pd.NA)
                                with pd.option_context('mode.use_inf_as_na', True):
                                    df_yield_combined_base['개별수율'] = (df_yield_combined_base['총_양품수량'] / df_yield_combined_base['총_생산수량']).fillna(1.0)
                                line_combined = df_yield_combined_base.groupby('period')['개별수율'].prod().reset_index(name='종합수율(%)')
                                line_combined['종합수율(%)'] *= 100
                                
                                # 3. 데이터 병합 및 보정
                                df_to_plot_pg = pd.merge(bar_combined, line_combined, on='period', how='outer').fillna(0)
                                df_to_plot_pg.loc[df_to_plot_pg['완제품_제조개수'] == 0, '종합수율(%)'] = 0
                                df_to_plot_pg['신규분류요약'] = "선택항목 종합"
                            else:
                                # 1. 실적 데이터 (최종 공정 기준)
                                bar_data_pg = df_resampled_pg_filtered[df_resampled_pg_filtered['공정코드'] == last_process].groupby(['period', '신규분류요약'])['총_양품수량'].sum().reset_index().rename(columns={'총_양품수량': '완제품_제조개수'})
                                
                                # 2. 수율 데이터 (개별 공정 수율의 곱)
                                df_resampled_pg_filtered_copy = df_resampled_pg_filtered.copy()
                                df_resampled_pg_filtered_copy['총_생산수량'] = df_resampled_pg_filtered_copy['총_생산수량'].replace(0, pd.NA)
                                with pd.option_context('mode.use_inf_as_na', True):
                                    df_resampled_pg_filtered_copy['개별공정수율'] = (df_resampled_pg_filtered_copy['총_양품수량'] / df_resampled_pg_filtered_copy['총_생산수량']).fillna(1.0)
                                line_data_pg = df_resampled_pg_filtered_copy.groupby(['period', '신규분류요약'])['개별공정수율'].prod().reset_index(name='종합수율(%)')
                                line_data_pg['종합수율(%)'] *= 100
                                
                                # 3. 데이터 병합 및 보정
                                df_to_plot_pg = pd.merge(bar_data_pg, line_data_pg, on=['period', '신규분류요약'], how='outer').sort_values('period').fillna(0)
                                df_to_plot_pg.loc[df_to_plot_pg['완제품_제조개수'] == 0, '종합수율(%)'] = 0

                            if not df_to_plot_pg.empty:
                                # 라벨 겹침 방지를 위한 동적 위치 조정
                                df_to_plot_pg = df_to_plot_pg.sort_values(['period', '종합수율(%)'], ascending=[True, False])
                                positions = ['top center', 'bottom center', 'middle right', 'middle left', 'top right', 'bottom right', 'top left', 'bottom left']
                                df_to_plot_pg['text_position'] = df_to_plot_pg.groupby('period').cumcount().apply(lambda i: positions[i % len(positions)])

                                fig_pg = make_subplots(specs=[[{"secondary_y": True}]])
                                
                                colors = px.colors.qualitative.Plotly
                                group_col = '신규분류요약'
                                
                                unique_groups = sorted(df_to_plot_pg[group_col].unique())
                                for i, group_name in enumerate(unique_groups):
                                    df_group = df_to_plot_pg[df_to_plot_pg[group_col] == group_name]
                                    color = colors[i % len(colors)]
                                    
                                    fig_pg.add_trace(go.Bar(
                                        x=df_group['period'], y=df_group['완제품_제조개수'], 
                                        name=f'{group_name} 완제품', legendgroup=group_name, marker_color=color,
                                        text=df_group['완제품_제조개수'], texttemplate='<b>%{text:,.0f}</b>', 
                                        textposition='outside',
                                        textfont=dict(size=18, color='black'),
                                        cliponaxis=False
                                    ), secondary_y=False)
                                    fig_pg.add_trace(go.Scatter(
                                        x=df_group['period'], y=df_group['종합수율(%)'], 
                                        name=f'{group_name} 수율', legendgroup=group_name, 
                                        mode='lines+markers+text', line=dict(color=color),
                                        text=df_group['종합수율(%)'], texttemplate='<b>%{text:.2f}%</b>', 
                                        textposition=df_group['text_position'],
                                        textfont=dict(size=16, color='black'),
                                        cliponaxis=False
                                    ), secondary_y=True)

                                factory_title = f"({pg_selected_factory})" if pg_selected_factory != '전체' else '(전체 공장)'
                                fig_pg.update_layout(
                                    height=600, 
                                    title_text=f'<b>{agg_level} 제품군별 완제품 제조 실적 및 종합 수율 {factory_title}</b>', 
                                    barmode='group', 
                                    legend_title_text='범례',
                                    uniformtext_minsize=12,
                                    uniformtext_mode='hide'
                                )
                                max_bar_val_pg = df_to_plot_pg['완제품_제조개수'].max() if not df_to_plot_pg.empty else 0
                                # 라벨 표시 공간 확보를 위해 y축 범위 상향 조정
                                fig_pg.update_yaxes(title_text="<b>완제품 제조 개수</b>", secondary_y=False, range=[0, max_bar_val_pg * 1.25])
                                fig_pg.update_yaxes(title_text="<b>종합 수율 (%)</b>", secondary_y=True, range=[0, 101])
                                fig_pg.update_xaxes(
                                    title_text=f"<b>{agg_level.replace('별', '')}</b>", 
                                    type='category', 
                                    categoryorder='array', 
                                    categoryarray=sorted(df_to_plot_pg['period'].unique()),
                                    automargin=True
                                )
                                
                                st.plotly_chart(fig_pg, use_container_width=True)
                            else:
                                st.info("선택된 조건에 해당하는 데이터가 없습니다.")
                        else:
                            st.info("선택된 제품군에 대한 데이터가 없습니다.")
                    else:
                        st.info("차트를 표시할 제품군을 선택해주세요.")
            else:
                st.warning("수율 데이터에 '신규분류요약' 컬럼이 없어 제품군별 분석을 제공할 수 없습니다.")
            
            # 다운로드 섹션 추가
            create_download_section(combo_data, "종합분석", agg_level, start_date, end_date)


elif selected_tab == "생산실적 상세조회":
    df_raw, start_date, end_date, agg_level = create_shared_filter_controls(df_yield_orig)
    
    if df_raw.empty:
        st.info("상세 조회에 필요한 '생산실적현황(간편)(수율).xlsx'와 유사한 파일 또는 선택된 기간 내 데이터가 없습니다.")
    else:
        st.markdown("### 🔍 생산실적 다방면 정밀 분석")
        
        # 분석 범위 선택
        scope_col1, scope_col2 = st.columns(2)
        with scope_col1:
            analysis_scope = st.selectbox(
                "분석 범위",
                ["완제품 기준 (최종공정)", "전체 공정 기준", "공정별 비교"]
            )
        with scope_col2:
            analysis_depth = st.selectbox(
                "분석 깊이",
                ["기본 분석", "심화 분석", "상세 분석"]
            )
        
        # 데이터 준비
        if analysis_scope == "완제품 기준 (최종공정)":
            df_base = df_raw[df_raw['공정코드'] == '[80] 누수/규격검사'].copy()
        elif analysis_scope == "전체 공정 기준":
            df_base = df_raw.copy()
        else:  # 공정별 비교
            df_base = df_raw.copy()
        
        # 공장별 수율 계산을 위한 전체 공정 데이터 준비 (종합 수율용)
        df_all_process = df_raw.copy()  # 모든 공정의 데이터를 유지
        
        if df_base.empty:
            st.warning("선택된 분석 범위에 데이터가 없습니다.")
        else:
            # 필터링 시스템
            with st.expander("🎛️ 필터링 옵션 (실시간 연동)", expanded=True):
                col1, col2, col3 = st.columns(3)
                
                # 1단계: 공장 선택
                with col1:
                    st.markdown("**기본 필터**")
                    factories = sorted(df_base['공장'].dropna().unique()) if '공장' in df_base.columns else []
                    sel_factories = st.multiselect("공장", factories, key="detail_factories")
                    
                    # 공장 선택에 따른 데이터 필터링
                    temp_data_1 = df_base.copy()
                    if sel_factories:
                        temp_data_1 = temp_data_1[temp_data_1['공장'].isin(sel_factories)]
                    
                    if analysis_scope == "공정별 비교":
                        processes = sorted(temp_data_1['공정코드'].dropna().unique()) if '공정코드' in temp_data_1.columns else []
                        sel_processes = st.multiselect("공정", processes, key="detail_processes")
                    else:
                        sel_processes = []
                
                # 2단계: 제품 선택 (공장/공정 고려)
                with col2:
                    st.markdown("**제품 필터**")
                    temp_data_2 = temp_data_1.copy()
                    if sel_processes:
                        temp_data_2 = temp_data_2[temp_data_2['공정코드'].isin(sel_processes)]
                    
                    categories = sorted(temp_data_2['신규분류요약'].dropna().unique()) if '신규분류요약' in temp_data_2.columns else []
                    sel_categories = st.multiselect("제품군", categories, key="detail_categories")
                    
                    # 제품군 선택에 따른 제품 필터링
                    temp_data_3 = temp_data_2.copy()
                    if sel_categories:
                        temp_data_3 = temp_data_3[temp_data_3['신규분류요약'].isin(sel_categories)]
                    
                    products = sorted(temp_data_3['품명'].dropna().unique()) if '품명' in temp_data_3.columns else []
                    product_search = st.text_input("제품명 검색", key="detail_search")
                    if product_search:
                        products = [p for p in products if product_search.lower() in str(p).lower()]
                    sel_products = st.multiselect("제품", products, key="detail_products")
                
                # 3단계: 설비 선택 (유연한 필터링)
                with col3:
                    st.markdown("**설비 필터**")
                    # 제품이 선택되지 않은 경우 공장/공정/제품군 기준으로만 필터링
                    temp_data_4 = temp_data_2.copy()  # 공장/공정/제품군까지만 적용
                    if sel_categories:
                        temp_data_4 = temp_data_4[temp_data_4['신규분류요약'].isin(sel_categories)]
                    # 제품이 선택된 경우에만 제품 필터링 추가 적용
                    if sel_products:
                        temp_data_4 = temp_data_4[temp_data_4['품명'].isin(sel_products)]
                    
                    machines = sorted(temp_data_4['기계코드'].dropna().unique()) if '기계코드' in temp_data_4.columns else []
                    st.info(f"💡 {len(machines)}개의 설비가 선택 가능합니다")
                    sel_machines = st.multiselect("설비", machines, key="detail_machines")
                    
                    auto_select = st.selectbox("자동 선택", ["없음", "수율 상위 10개", "생산량 상위 10개"])
                    if auto_select != "없음" and '총_생산수량' in df_base.columns and '총_양품수량' in df_base.columns:
                        temp_data = df_base.copy()
                        if sel_factories:
                            temp_data = temp_data[temp_data['공장'].isin(sel_factories)]
                        if sel_categories:
                            temp_data = temp_data[temp_data['신규분류요약'].isin(sel_categories)]
                        
                        if auto_select == "수율 상위 10개":
                            auto_agg = temp_data.groupby('품명').agg({'총_생산수량': 'sum', '총_양품수량': 'sum'}).reset_index()
                            auto_agg['수율'] = auto_agg['총_양품수량'] * 100 / auto_agg['총_생산수량'].replace(0, 1)
                            top_products = auto_agg.nlargest(10, '수율')['품명'].tolist()
                            if not sel_products:
                                sel_products = top_products
                        else:  # 생산량 상위 10개
                            auto_agg = temp_data.groupby('품명')['총_생산수량'].sum().reset_index()
                            top_products = auto_agg.nlargest(10, '총_생산수량')['품명'].tolist()
                            if not sel_products:
                                sel_products = top_products
            
            # 필터 적용
            df_filtered = df_base.copy()
            if sel_factories:
                df_filtered = df_filtered[df_filtered['공장'].isin(sel_factories)]
            if sel_processes:
                df_filtered = df_filtered[df_filtered['공정코드'].isin(sel_processes)]
            if sel_categories:
                df_filtered = df_filtered[df_filtered['신규분류요약'].isin(sel_categories)]
            if sel_products:
                df_filtered = df_filtered[df_filtered['품명'].isin(sel_products)]
            if sel_machines:
                df_filtered = df_filtered[df_filtered['기계코드'].isin(sel_machines)]
            
            if df_filtered.empty:
                st.warning("선택한 조건에 해당하는 데이터가 없습니다.")
            else:
                # 수율 계산
                if '수율(%)' not in df_filtered.columns and '총_생산수량' in df_filtered.columns and '총_양품수량' in df_filtered.columns:
                    df_filtered['수율(%)'] = (df_filtered['총_양품수량'] * 100 / df_filtered['총_생산수량'].replace(0, pd.NA)).fillna(0)
                
                # KPI 요약
                st.divider()
                st.markdown("### 📊 핵심 성과 지표")
                
                kpi1, kpi2, kpi3, kpi4 = st.columns(4)
                
                # KPI용 최종공정 생산수량 계산
                final_process_kpi = df_all_process[df_all_process['공정코드'] == '[80] 누수/규격검사'].copy()
                # 필터 적용
                if sel_processes:
                    final_process_kpi = final_process_kpi[final_process_kpi['공정코드'].isin(sel_processes)]
                if sel_categories:
                    final_process_kpi = final_process_kpi[final_process_kpi['신규분류요약'].isin(sel_categories)]
                if sel_products:
                    final_process_kpi = final_process_kpi[final_process_kpi['품명'].isin(sel_products)]
                if sel_machines:
                    final_process_kpi = final_process_kpi[final_process_kpi['기계코드'].isin(sel_machines)]
                if sel_factories:
                    final_process_kpi = final_process_kpi[final_process_kpi['공장'].isin(sel_factories)]
                
                total_prod = final_process_kpi['총_생산수량'].sum() if not final_process_kpi.empty else 0
                total_good = final_process_kpi['총_양품수량'].sum() if not final_process_kpi.empty else 0
                
                # KPI용 종합수율 계산 (공정별 수율의 곱)
                all_process_kpi = df_all_process.copy()
                # 필터 적용
                if sel_processes:
                    all_process_kpi = all_process_kpi[all_process_kpi['공정코드'].isin(sel_processes)]
                if sel_categories:
                    all_process_kpi = all_process_kpi[all_process_kpi['신규분류요약'].isin(sel_categories)]
                if sel_products:
                    all_process_kpi = all_process_kpi[all_process_kpi['품명'].isin(sel_products)]
                if sel_machines:
                    all_process_kpi = all_process_kpi[all_process_kpi['기계코드'].isin(sel_machines)]
                if sel_factories:
                    all_process_kpi = all_process_kpi[all_process_kpi['공장'].isin(sel_factories)]
                
                if not all_process_kpi.empty:
                    # 전체 공정별 수율 계산
                    process_yields_kpi = all_process_kpi.groupby('공정코드').agg({
                        '총_생산수량': 'sum',
                        '총_양품수량': 'sum'
                    }).reset_index()
                    process_yields_kpi['공정수율'] = process_yields_kpi['총_양품수량'] / process_yields_kpi['총_생산수량'].replace(0, 1)
                    
                    # 전체 공정수율의 곱으로 종합수율 계산
                    avg_yield = process_yields_kpi['공정수율'].prod() * 100
                else:
                    avg_yield = 0
                
                with kpi1:
                    st.metric("총 생산수량", f"{total_prod:,.0f}")
                with kpi2:
                    st.metric("총 양품수량", f"{total_good:,.0f}")
                with kpi3:
                    st.metric("종합 수율", f"{avg_yield:.1f}%")
                with kpi4:
                    data_count = len(df_filtered)
                    st.metric("데이터 건수", f"{data_count:,}")
                
                # 분석 탭
                st.divider()
                if analysis_depth == "기본 분석":
                    tab1, tab2, tab3, tab4 = st.tabs(["🏭 공장별", "📦 제품별", "⚙️ 설비별", "📈 시계열"])
                elif analysis_depth == "심화 분석":
                    tab1, tab2, tab3, tab4 = st.tabs(["🏭 공장별", "📦 제품별", "⚙️ 설비별", "📈 시계열"])
                else:  # 상세 분석
                    tab1, tab2, tab3, tab4, tab5 = st.tabs(["🏭 공장별", "📦 제품별", "⚙️ 설비별", "📈 시계열", "🔬 고급분석"])
                
                with tab1:
                    st.markdown("#### 공장별 생산 성과 (종합 수율)")
                    if '공장' in df_filtered.columns:
                        # 1. 최종공정의 총 생산수량 계산
                        final_process_base = df_all_process[df_all_process['공정코드'] == '[80] 누수/규격검사'].copy()
                        # 필터 적용
                        if sel_categories:
                            final_process_base = final_process_base[final_process_base['신규분류요약'].isin(sel_categories)]
                        if sel_products:
                            final_process_base = final_process_base[final_process_base['품명'].isin(sel_products)]
                        if sel_machines:
                            final_process_base = final_process_base[final_process_base['기계코드'].isin(sel_machines)]
                        
                        # 공장별 최종공정 생산수량
                        factory_production = final_process_base.groupby('공장')['총_생산수량'].sum().reset_index()
                        
                        # 2. 공정별 수율의 곱 계산
                        all_process_base = df_all_process.copy()
                        # 필터 적용
                        if sel_categories:
                            all_process_base = all_process_base[all_process_base['신규분류요약'].isin(sel_categories)]
                        if sel_products:
                            all_process_base = all_process_base[all_process_base['품명'].isin(sel_products)]
                        if sel_machines:
                            all_process_base = all_process_base[all_process_base['기계코드'].isin(sel_machines)]
                        
                        # 공정별 수율 계산
                        process_yields = all_process_base.groupby(['공장', '공정코드']).agg({
                            '총_생산수량': 'sum',
                            '총_양품수량': 'sum'
                        }).reset_index()
                        process_yields['공정수율'] = process_yields['총_양품수량'] / process_yields['총_생산수량'].replace(0, 1)
                        
                        # 공장별 공정수율의 곱 계산
                        factory_compound_yield = process_yields.groupby('공장')['공정수율'].prod().reset_index()
                        factory_compound_yield['종합수율(%)'] = factory_compound_yield['공정수율'] * 100
                        factory_compound_yield = factory_compound_yield[['공장', '종합수율(%)']]
                        
                        # 3. 최종 데이터 병합
                        factory_data = factory_production.merge(factory_compound_yield, on='공장', how='left')
                        factory_data['종합수율(%)'] = factory_data['종합수율(%)'].fillna(0)
                        
                        # 그래프 설정 옵션
                        with st.expander("📊 차트 설정", expanded=False):
                            col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                            with col_set1:
                                factory_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="factory_label_size")
                            with col_set2:
                                factory_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="factory_axis_title_size")
                            with col_set3:
                                factory_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="factory_axis_tick_size")
                            with col_set4:
                                factory_performance_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=550, step=50, key="factory_performance_chart_height")

                        col1, col2 = st.columns(2)
                        with col1:
                            fig = px.bar(factory_data, x='공장', y='총_생산수량', title="공장별 생산수량 (최종공정)", text='총_생산수량')
                            fig.update_traces(texttemplate='%{text:,.0f}', textposition='outside', textfont_size=factory_label_size, textfont_color='black')
                            # Y축 범위 조정 (라벨 공간 확보)
                            max_val = factory_data['총_생산수량'].max()
                            fig.update_layout(
                                yaxis_title="생산수량",
                                yaxis=dict(range=[0, max_val * 1.2]),  # 20% 여유 공간
                                margin=dict(t=80, b=80, l=80, r=80),
                                height=factory_performance_chart_height,
                                title_font_size=factory_axis_title_size
                            )
                            fig.update_xaxes(title_font_size=factory_axis_title_size, tickfont_size=factory_axis_tick_size)
                            fig.update_yaxes(title_font_size=factory_axis_title_size, tickfont_size=factory_axis_tick_size)
                            st.plotly_chart(fig, use_container_width=True)
                        with col2:
                            fig = px.bar(factory_data, x='공장', y='종합수율(%)', title="공장별 종합수율 (공정수율의 곱)", color='종합수율(%)', text='종합수율(%)')
                            fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside', textfont_size=factory_label_size, textfont_color='black')
                            # Y축 범위 조정 (라벨 공간 확보)
                            max_val = factory_data['종합수율(%)'].max()
                            fig.update_layout(
                                yaxis_title="종합수율(%)",
                                yaxis=dict(range=[0, max_val * 1.2]),  # 20% 여유 공간
                                margin=dict(t=80, b=80, l=80, r=80),
                                height=factory_performance_chart_height,
                                title_font_size=factory_axis_title_size
                            )
                            fig.update_xaxes(title_font_size=factory_axis_title_size, tickfont_size=factory_axis_tick_size)
                            fig.update_yaxes(title_font_size=factory_axis_title_size, tickfont_size=factory_axis_tick_size)
                            st.plotly_chart(fig, use_container_width=True)
                        
                        st.dataframe(factory_data, use_container_width=True)
                    else:
                        st.info("공장 정보가 없습니다.")
                
                with tab2:
                    st.markdown("#### 제품별 생산 성과 (종합수율)")
                    if sel_products and '품명' in df_filtered.columns:
                        # 1. 제품별 최종공정 생산수량 계산
                        final_product_base = df_all_process[df_all_process['공정코드'] == '[80] 누수/규격검사'].copy()
                        # 필터 적용
                        if sel_categories:
                            final_product_base = final_product_base[final_product_base['신규분류요약'].isin(sel_categories)]
                        if sel_products:
                            final_product_base = final_product_base[final_product_base['품명'].isin(sel_products)]
                        if sel_machines:
                            final_product_base = final_product_base[final_product_base['기계코드'].isin(sel_machines)]
                        if sel_factories:
                            final_product_base = final_product_base[final_product_base['공장'].isin(sel_factories)]
                        
                        # 제품별 최종공정 생산수량
                        product_production = final_product_base.groupby('품명')[['총_생산수량', '총_양품수량']].sum().reset_index()
                        
                        # 2. 제품별 공정수율의 곱 계산
                        all_product_base = df_all_process.copy()
                        # 필터 적용
                        if sel_categories:
                            all_product_base = all_product_base[all_product_base['신규분류요약'].isin(sel_categories)]
                        if sel_products:
                            all_product_base = all_product_base[all_product_base['품명'].isin(sel_products)]
                        if sel_machines:
                            all_product_base = all_product_base[all_product_base['기계코드'].isin(sel_machines)]
                        if sel_factories:
                            all_product_base = all_product_base[all_product_base['공장'].isin(sel_factories)]
                        
                        # 제품별 공정수율 계산
                        product_process_yields = all_product_base.groupby(['품명', '공정코드']).agg({
                            '총_생산수량': 'sum',
                            '총_양품수량': 'sum'
                        }).reset_index()
                        product_process_yields['공정수율'] = product_process_yields['총_양품수량'] / product_process_yields['총_생산수량'].replace(0, 1)
                        
                        # 제품별 공정수율의 곱 계산
                        product_compound_yield = product_process_yields.groupby('품명')['공정수율'].prod().reset_index()
                        product_compound_yield['종합수율(%)'] = product_compound_yield['공정수율'] * 100
                        
                        # 3. 최종 데이터 병합
                        product_data = product_production.merge(product_compound_yield[['품명', '종합수율(%)']], on='품명', how='left')
                        product_data['종합수율(%)'] = product_data['종합수율(%)'].fillna(0)
                        
                        # 그래프 설정 옵션
                        with st.expander("📊 차트 설정", expanded=False):
                            col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                            with col_set1:
                                product_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="product_label_size")
                            with col_set2:
                                product_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="product_axis_title_size")
                            with col_set3:
                                product_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="product_axis_tick_size")
                            with col_set4:
                                product_performance_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=650, step=50, key="product_performance_chart_height")

                        # 성과 매트릭스
                        fig = px.scatter(product_data, x='총_생산수량', y='종합수율(%)', 
                                       size='총_양품수량', hover_name='품명',
                                       title="제품별 성과 매트릭스 (종합수율)", text='품명', height=product_performance_chart_height)
                        fig.update_traces(textposition='top center', textfont_size=product_label_size, textfont_color='black')
                        fig.update_layout(xaxis_title="총 생산수량 (최종공정)", yaxis_title="종합수율(%)", title_font_size=product_axis_title_size)
                        fig.update_xaxes(title_font_size=product_axis_title_size, tickfont_size=product_axis_tick_size)
                        fig.update_yaxes(title_font_size=product_axis_title_size, tickfont_size=product_axis_tick_size)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # 순위 차트
                        sort_by = st.selectbox("정렬 기준", ["총_생산수량", "총_양품수량", "종합수율(%)"])
                        sorted_data = product_data.sort_values(sort_by, ascending=False).head(10)
                        
                        fig = px.bar(sorted_data, x='품명', y=sort_by, title=f"제품별 {sort_by} Top 10", text=sort_by)
                        if sort_by == '종합수율(%)':
                            fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside', textfont_size=product_label_size, textfont_color='black')
                        else:
                            fig.update_traces(texttemplate='%{text:,.0f}', textposition='outside', textfont_size=product_label_size, textfont_color='black')
                        # Y축 범위 조정 (라벨 공간 확보)
                        max_val = sorted_data[sort_by].max()
                        fig.update_layout(
                            xaxis_tickangle=-45, 
                            yaxis_title=sort_by,
                            yaxis=dict(range=[0, max_val * 1.25]),  # 25% 여유 공간
                            margin=dict(t=80, b=180, l=80, r=80),
                            height=product_performance_chart_height,
                            title_font_size=product_axis_title_size
                        )
                        fig.update_xaxes(title_font_size=product_axis_title_size, tickfont_size=product_axis_tick_size)
                        fig.update_yaxes(title_font_size=product_axis_title_size, tickfont_size=product_axis_tick_size)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        st.dataframe(product_data.sort_values(sort_by, ascending=False), use_container_width=True)
                    else:
                        st.info("제품을 선택해주세요.")
                
                with tab3:
                    st.markdown("#### 설비별 생산 성과 (종합수율)")
                    if '기계코드' in df_filtered.columns:
                        # 1. 설비별 최종공정 생산수량 계산
                        final_machine_base = df_all_process[df_all_process['공정코드'] == '[80] 누수/규격검사'].copy()
                        # 필터 적용
                        if sel_categories:
                            final_machine_base = final_machine_base[final_machine_base['신규분류요약'].isin(sel_categories)]
                        if sel_products:
                            final_machine_base = final_machine_base[final_machine_base['품명'].isin(sel_products)]
                        if sel_machines:
                            final_machine_base = final_machine_base[final_machine_base['기계코드'].isin(sel_machines)]
                        if sel_factories:
                            final_machine_base = final_machine_base[final_machine_base['공장'].isin(sel_factories)]
                        
                        # 설비별 최종공정 생산수량 및 제품수
                        machine_production = final_machine_base.groupby('기계코드').agg({
                            '총_생산수량': 'sum',
                            '총_양품수량': 'sum',
                            '품명': 'nunique'
                        }).reset_index()
                        machine_production.rename(columns={'품명': '생산제품수'}, inplace=True)
                        
                        # 2. 설비별 공정수율의 곱 계산
                        all_machine_base = df_all_process.copy()
                        # 필터 적용
                        if sel_categories:
                            all_machine_base = all_machine_base[all_machine_base['신규분류요약'].isin(sel_categories)]
                        if sel_products:
                            all_machine_base = all_machine_base[all_machine_base['품명'].isin(sel_products)]
                        if sel_machines:
                            all_machine_base = all_machine_base[all_machine_base['기계코드'].isin(sel_machines)]
                        if sel_factories:
                            all_machine_base = all_machine_base[all_machine_base['공장'].isin(sel_factories)]
                        
                        # 설비별 공정수율 계산
                        machine_process_yields = all_machine_base.groupby(['기계코드', '공정코드']).agg({
                            '총_생산수량': 'sum',
                            '총_양품수량': 'sum'
                        }).reset_index()
                        machine_process_yields['공정수율'] = machine_process_yields['총_양품수량'] / machine_process_yields['총_생산수량'].replace(0, 1)
                        
                        # 설비별 공정수율의 곱 계산
                        machine_compound_yield = machine_process_yields.groupby('기계코드')['공정수율'].prod().reset_index()
                        machine_compound_yield['종합수율(%)'] = machine_compound_yield['공정수율'] * 100
                        
                        # 3. 최종 데이터 병합
                        machine_data = machine_production.merge(machine_compound_yield[['기계코드', '종합수율(%)']], on='기계코드', how='left')
                        machine_data['종합수율(%)'] = machine_data['종합수율(%)'].fillna(0)
                        
                        # 그래프 설정 옵션
                        with st.expander("📊 차트 설정", expanded=False):
                            col_set1, col_set2, col_set3, col_set4 = st.columns(4)
                            with col_set1:
                                machine_label_size = st.slider("라벨 크기", min_value=8, max_value=30, value=18, step=1, key="machine_label_size")
                            with col_set2:
                                machine_axis_title_size = st.slider("축 제목 크기", min_value=10, max_value=30, value=18, step=1, key="machine_axis_title_size")
                            with col_set3:
                                machine_axis_tick_size = st.slider("축 서식 크기", min_value=8, max_value=30, value=18, step=1, key="machine_axis_tick_size")
                            with col_set4:
                                machine_performance_chart_height = st.slider("차트 높이", min_value=400, max_value=1000, value=600, step=50, key="machine_performance_chart_height")

                        # 효율성 분석
                        fig = px.scatter(machine_data, x='총_생산수량', y='종합수율(%)', 
                                       size='생산제품수', hover_name='기계코드',
                                       title="설비별 효율성 분석 (종합수율)", text='기계코드', height=machine_performance_chart_height)
                        fig.update_traces(textposition='top center', textfont_size=machine_label_size, textfont_color='black')
                        fig.update_layout(xaxis_title="총 생산수량 (최종공정)", yaxis_title="종합수율(%)", title_font_size=machine_axis_title_size)
                        fig.update_xaxes(title_font_size=machine_axis_title_size, tickfont_size=machine_axis_tick_size)
                        fig.update_yaxes(title_font_size=machine_axis_title_size, tickfont_size=machine_axis_tick_size)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Top 설비
                        top_machines = machine_data.nlargest(10, '총_생산수량')
                        col1, col2 = st.columns(2)
                        with col1:
                            fig = px.bar(top_machines, x='기계코드', y='총_생산수량', title="생산수량 Top 10", text='총_생산수량')
                            fig.update_traces(texttemplate='%{text:,.0f}', textposition='outside', textfont_size=machine_label_size, textfont_color='black')
                            # Y축 범위 조정 (라벨 공간 확보)
                            max_val = top_machines['총_생산수량'].max()
                            fig.update_layout(
                                xaxis_tickangle=-45, 
                                yaxis_title="생산수량",
                                yaxis=dict(range=[0, max_val * 1.25]),  # 25% 여유 공간
                                margin=dict(t=80, b=180, l=80, r=80),
                                height=machine_performance_chart_height,
                                title_font_size=machine_axis_title_size
                            )
                            fig.update_xaxes(title_font_size=machine_axis_title_size, tickfont_size=machine_axis_tick_size)
                            fig.update_yaxes(title_font_size=machine_axis_title_size, tickfont_size=machine_axis_tick_size)
                            st.plotly_chart(fig, use_container_width=True)
                        with col2:
                            fig = px.bar(top_machines, x='기계코드', y='종합수율(%)', title="종합수율 Top 10", color='종합수율(%)', text='종합수율(%)')
                            fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside', textfont_size=machine_label_size, textfont_color='black')
                            # Y축 범위 조정 (라벨 공간 확보)
                            max_val = top_machines['종합수율(%)'].max()
                            fig.update_layout(
                                xaxis_tickangle=-45, 
                                yaxis_title="종합수율(%)",
                                yaxis=dict(range=[0, max_val * 1.25]),  # 25% 여유 공간
                                margin=dict(t=80, b=180, l=80, r=80),
                                height=machine_performance_chart_height,
                                title_font_size=machine_axis_title_size
                            )
                            fig.update_xaxes(title_font_size=machine_axis_title_size, tickfont_size=machine_axis_tick_size)
                            fig.update_yaxes(title_font_size=machine_axis_title_size, tickfont_size=machine_axis_tick_size)
                            st.plotly_chart(fig, use_container_width=True)
                        
                        st.dataframe(machine_data.sort_values('총_생산수량', ascending=False), use_container_width=True)
                    else:
                        st.info("설비 정보가 없습니다.")
                
                # 시계열 탭은 모든 분석 깊이에서 표시
                if analysis_depth in ["기본 분석", "심화 분석", "상세 분석"]:
                    with tab4:
                        st.markdown("#### 시계열 추이 분석")
                        
                        time_basis = st.selectbox("분석 기준", ["제품별", "공장별", "설비별"])
                        
                        if time_basis == "제품별" and sel_products:
                            group_col = '품명'
                            time_data = df_filtered[df_filtered['품명'].isin(sel_products)]
                        elif time_basis == "공장별":
                            group_col = '공장'
                            # 공장별 종합 수율을 위해 전체 공정 데이터 사용
                            time_data = df_all_process.copy()
                            # 필터 적용 (공장 필터 제외)
                            if sel_processes:
                                time_data = time_data[time_data['공정코드'].isin(sel_processes)]
                            if sel_categories:
                                time_data = time_data[time_data['신규분류요약'].isin(sel_categories)]
                            if sel_products:
                                time_data = time_data[time_data['품명'].isin(sel_products)]
                            if sel_machines:
                                time_data = time_data[time_data['기계코드'].isin(sel_machines)]
                            if sel_factories:  # 공장 필터는 마지막에 적용
                                time_data = time_data[time_data['공장'].isin(sel_factories)]
                        elif time_basis == "설비별" and sel_machines:
                            group_col = '기계코드'
                            time_data = df_filtered[df_filtered['기계코드'].isin(sel_machines)]
                        else:
                            st.info(f"{time_basis} 분석을 위한 데이터를 선택해주세요.")
                            time_data = pd.DataFrame()
                        
                        if not time_data.empty and group_col in time_data.columns:
                            # 시계열 집계
                            trend_data = get_resampled_data(
                                time_data, agg_level, 
                                ['총_생산수량', '총_양품수량'], 
                                group_by_cols=['period', group_col]
                            )
                            
                            if not trend_data.empty:
                                if time_basis == "공장별":
                                    trend_data['종합수율(%)'] = trend_data['총_양품수량'] * 100 / trend_data['총_생산수량'].replace(0, 1)
                                    yield_col = '종합수율(%)'
                                    yield_title = '종합수율'
                                else:
                                    trend_data['수율(%)'] = trend_data['총_양품수량'] * 100 / trend_data['총_생산수량'].replace(0, 1)
                                    yield_col = '수율(%)'
                                    yield_title = '수율'
                                
                                # 듀얼 축 차트
                                fig = make_subplots(specs=[[{"secondary_y": True}]])
                                
                                groups = sorted(trend_data[group_col].unique())
                                colors = px.colors.qualitative.Plotly
                                
                                for i, group in enumerate(groups[:6]):
                                    group_data = trend_data[trend_data[group_col] == group]
                                    color = colors[i % len(colors)]
                                    
                                    # 생산량 바
                                    fig.add_trace(go.Bar(
                                        x=group_data['period'], y=group_data['총_양품수량'],
                                        name=f'{group} 양품수량', marker_color=color,
                                        text=group_data['총_양품수량'],
                                        texttemplate='%{text:,.0f}',
                                        textposition='outside',
                                        textfont_size=16
                                    ), secondary_y=False)
                                    
                                    # 수율 라인
                                    fig.add_trace(go.Scatter(
                                        x=group_data['period'], y=group_data[yield_col],
                                        name=f'{group} {yield_title}(%)', line=dict(color=color),
                                        mode='lines+markers+text',
                                        text=group_data[yield_col],
                                        texttemplate='%{text:.1f}%',
                                        textposition='top center',
                                        textfont_size=16
                                    ), secondary_y=True)
                                
                                fig.update_layout(title=f"{agg_level} {time_basis} 추이", height=600, barmode='group')
                                fig.update_yaxes(title_text="양품수량", secondary_y=False)
                                fig.update_yaxes(title_text=f"{yield_title} (%)", secondary_y=True)
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # 데이터 테이블
                                pivot_data = trend_data.pivot(index='period', columns=group_col, values=['총_양품수량', yield_col])
                                st.dataframe(pivot_data, use_container_width=True)
                            else:
                                st.info("시계열 분석할 데이터가 없습니다.")
                
                if analysis_depth == "상세 분석":
                    with tab5:
                        st.markdown("#### 고급 분석")
                        
                        analysis_type = st.selectbox("고급 분석 유형", ["성과 비교", "이상치 분석", "상관관계 분석"])
                        
                        if analysis_type == "성과 비교" and sel_products and len(sel_products) >= 2:
                            st.markdown("##### 제품간 성과 비교")
                            
                            compare_data = df_filtered.groupby('품명').agg({
                                '총_생산수량': 'sum',
                                '총_양품수량': 'sum'
                            }).reset_index()
                            compare_data['수율(%)'] = compare_data['총_양품수량'] * 100 / compare_data['총_생산수량'].replace(0, 1)
                            
                            # 레이더 차트
                            fig = go.Figure()
                            
                            for product in sel_products[:5]:
                                product_row = compare_data[compare_data['품명'] == product]
                                if not product_row.empty:
                                    # 정규화된 값
                                    prod_norm = product_row['총_생산수량'].iloc[0] / compare_data['총_생산수량'].max() * 100
                                    good_norm = product_row['총_양품수량'].iloc[0] / compare_data['총_양품수량'].max() * 100
                                    yield_val = product_row['수율(%)'].iloc[0]
                                    
                                    fig.add_trace(go.Scatterpolar(
                                        r=[prod_norm, good_norm, yield_val, prod_norm],
                                        theta=['생산수량', '양품수량', '수율(%)', '생산수량'],
                                        fill='toself',
                                        name=product,
                                        text=[f'{prod_norm:.1f}', f'{good_norm:.1f}', f'{yield_val:.1f}%', f'{prod_norm:.1f}'],
                                        textposition='middle center',
                                        textfont_size=16
                                    ))
                            
                            fig.update_layout(
                                polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                                title="제품별 성과 비교 (레이더 차트)"
                            )
                            st.plotly_chart(fig, use_container_width=True)
                            
                        elif analysis_type == "이상치 분석":
                            st.markdown("##### 수율 이상치 분석")
                            
                            if '수율(%)' in df_filtered.columns and sel_products:
                                outlier_data = df_filtered.groupby('품명')['수율(%)'].agg(['mean', 'std']).reset_index()
                                outlier_data['변동계수'] = outlier_data['std'] / outlier_data['mean']
                                outlier_data = outlier_data.sort_values('변동계수', ascending=False)
                                
                                fig = px.bar(outlier_data.head(10), x='품명', y='변동계수', 
                                           title="수율 변동성이 높은 제품 Top 10", text='변동계수')
                                fig.update_traces(texttemplate='%{text:.3f}', textposition='outside', textfont_size=16, textfont_color='black')
                                # Y축 범위 조정 (라벨 공간 확보)
                                max_val = outlier_data.head(10)['변동계수'].max()
                                fig.update_layout(
                                    xaxis_tickangle=-45, 
                                    yaxis_title="변동계수",
                                    yaxis=dict(range=[0, max_val * 1.25]),  # 25% 여유 공간
                                    margin=dict(t=80, b=180, l=80, r=80),
                                    height=650,
                                    title_font_size=14
                                )
                                st.plotly_chart(fig, use_container_width=True)
                                
                                st.dataframe(outlier_data, use_container_width=True)
                        
                        else:  # 상관관계 분석
                            st.markdown("##### 생산수량 vs 수율 상관관계")
                            
                            if sel_products and '총_생산수량' in df_filtered.columns and '수율(%)' in df_filtered.columns:
                                corr_data = df_filtered.groupby('품명').agg({
                                    '총_생산수량': 'sum',
                                    '수율(%)': 'mean'
                                }).reset_index()
                                
                                correlation = corr_data['총_생산수량'].corr(corr_data['수율(%)'])
                                
                                fig = px.scatter(corr_data, x='총_생산수량', y='수율(%)', 
                                               hover_name='품명', trendline="ols",
                                               title=f"생산수량 vs 수율 상관관계 (상관계수: {correlation:.3f})",
                                               text='품명')
                                fig.update_traces(textposition='top center', textfont_size=9, textfont_color='black')
                                fig.update_layout(xaxis_title="총 생산수량", yaxis_title="수율(%)")
                                st.plotly_chart(fig, use_container_width=True)
                                
                                if abs(correlation) > 0.5:
                                    if correlation > 0:
                                        st.success(f"강한 양의 상관관계 (r={correlation:.3f}): 생산수량이 많을수록 수율이 높습니다.")
                                    else:
                                        st.warning(f"강한 음의 상관관계 (r={correlation:.3f}): 생산수량이 많을수록 수율이 낮습니다.")
                                else:
                                    st.info(f"약한 상관관계 (r={correlation:.3f}): 생산수량과 수율 간 뚜렷한 관계가 없습니다.")
                
                # 원본 데이터
                st.divider()
                with st.expander("📋 원본 데이터", expanded=False):
                    if not df_filtered.empty:
                        cols = df_filtered.columns.tolist()
                        default_cols = ['date', '공장', '공정코드', '품명', '신규분류요약', '기계코드', '총_생산수량', '총_양품수량', '수율(%)']
                        display_cols = [col for col in default_cols if col in cols]
                        
                        selected_cols = st.multiselect("표시할 컬럼", cols, default=display_cols)
                        
                        if selected_cols:
                            display_df = df_filtered[selected_cols]
                            sort_cols = [col for col in ['date', '공장', '공정코드'] if col in selected_cols]
                            if sort_cols:
                                display_df = display_df.sort_values(sort_cols)
                            st.dataframe(display_df, use_container_width=True, height=400)
        
        # 다운로드 섹션 추가
        create_download_section(df_filtered, "생산실적상세조회", agg_level, start_date, end_date)

